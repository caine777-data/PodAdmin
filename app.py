#!/usr/bin/env python3
"""
app.py — PodAdmin (interface graphique)
Université de Toulouse — MFCA

Console d'administration Esup-Pod. Reprend le téléversement par lot de
« Pod Téléverseur » et y ajoute des modules d'admin :
  • Comptes — statut « équipe » (is_staff)
  • (à venir) Réaffectation de propriétaire, Nettoyage/Modération,
    Inventaire/Stats, Chaînes & thèmes.
Nécessite un token de compte SUPERUTILISATEUR.
"""

from __future__ import annotations

__author__      = "Cédric MONNA"
__contact__     = "cedricmonna@gmail.com"
__institution__ = "Université de Toulouse — MFCA"
from __version__ import __version__   # source unique (voir __version__.py)
__date__        = "2026"
__copyright__   = "© Copyright 2026 Cédric MONNA"
__license__     = ("Tous droits réservés — réutilisation, diffusion ou "
                   "adaptation soumises à l'autorisation de l'auteur.")


import os
import sys
import subprocess
import threading
import time
import urllib.parse
from datetime import datetime

import customtkinter as ctk
from tkinter import filedialog, messagebox, simpledialog

import config as cfg
import maj                                  # vérification des mises à jour
from pod_api import (PodAPI, PodAPIError,
                     SUBTITLE_LANGS, SUBTITLE_KINDS)
# Moteur de téléversement par morceaux via session web (gros fichiers > seuil).
from pod_chunked import PodChunkedSession, PodChunkedError

# Pillow (fourni avec customtkinter) — pour afficher le logo
try:
    from PIL import Image as PILImage
    HAS_PIL = True
except Exception:
    HAS_PIL = False


def resource_path(rel: str) -> str:
    """Chemin d'une ressource, compatible PyInstaller (--onefile) et exécution directe."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)

# Glisser-déposer (optionnel — l'appli fonctionne sans, via les boutons)
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DND = True
except Exception:
    HAS_DND = False

APP_TITLE = "PodAdmin — Université de Toulouse"
APP_VERSION = __version__      # affichée dans la barre latérale et « À propos »

# Délai (ms) d'attente après la dernière frappe avant de reconstruire une liste
# filtrée. Assez court pour rester réactif, assez long pour éviter de refaire
# tout l'affichage à chaque caractère saisi.
FILTER_DELAY_MS = 250

# Délai (ms) avant de redessiner les listes après une modification (suppression,
# changement de statut, affectation à une chaîne…). Ce court sursis laisse le
# temps de voir les ✔/✗ posés sur les lignes traitées avant que la liste ne se
# nettoie. Il sert aussi à REGROUPER les rafraîchissements : pendant un lot de
# 300 vidéos, un seul redessin a lieu, à la fin.
REFRESH_DELAY_MS = 1500

# Nom du champ « vidéo 360° » dans l'API Esup-Pod. « is_360 » est le nom
# standard ; centralisé ici pour n'avoir qu'un seul point à corriger si une
# instance le nommait autrement (à confirmer via verifier_champ_360.py).
FIELD_360 = "is_360"

# ══════════════════════════════════════════════════════════════════════════
#  PALETTE — une couleur = un sens
# ══════════════════════════════════════════════════════════════════════════
#
# Les couleurs étaient écrites en hexadécimal directement dans les appels de
# widgets : 24 teintes distinctes, dont #f59e0b répété 75 fois. Changer la
# nuance d'avertissement demandait 75 modifications, et le même vert servait
# tantôt à signaler une réussite, tantôt à marquer un bouton d'action — rôle
# que le bleu tenait ailleurs.
#
# Chaque constante est un COUPLE (mode clair, mode sombre). CustomTkinter
# choisit automatiquement selon le thème actif : c'est ce qui rend le
# basculement clair/sombre possible. Une couleur unique resterait identique
# dans les deux modes, et donnerait par endroits du texte pâle sur fond clair.

# — Couleurs d'action ——————————————————————————————————————————————————
C_ACTION       = ("#2563eb", "#2563eb")   # UNIQUE action principale de l'écran
C_ACTION_SURV  = ("#1d4ed8", "#1d4ed8")
# ⚠️ Les UTILITAIRES ne prennent PAS cette teinte : « Rafraîchir », « Recharger »,
# « Parcourir… » sont en C_NEUTRE. Un écran ne porte qu'un seul bouton coloré,
# celui de son action propre. Le commentaire citait « Rafraîchir » en exemple
# alors que les quatre boutons de ce nom sont neutres depuis la 1.5.2 : c'est
# ainsi qu'un onglet ajouté plus tard aurait repris le bleu.
C_SUCCES       = ("#15803d", "#16a34a")   # validation : Lancer, Enregistrer
C_SUCCES_SURV  = ("#166534", "#15803d")
C_ALERTE       = ("#b45309", "#b45309")   # avertissement, interruption
C_ALERTE_SURV  = ("#92400e", "#92400e")
C_ERREUR       = ("#b91c1c", "#ef4444")   # échec, message d'erreur
C_DESTRUCTIF   = ("#7f1d1d", "#7f1d1d")   # suppression définitive UNIQUEMENT
C_DESTR_SURV   = ("#991b1b", "#991b1b")
C_NEUTRE       = ("gray65", "gray35")     # secondaire : Annuler, Fermer
C_NEUTRE_SURV  = ("gray55", "gray28")
# Un bouton gris gardait le texte BLANC par défaut de CustomTkinter : en mode
# clair, blanc sur gris65 donne un bouton qui paraît DÉSACTIVÉ. Le texte des
# boutons secondaires doit donc être foncé en clair, pâle en sombre. Un bouton
# réellement désactivé reste distinct : CustomTkinter lui applique sa propre
# teinte « text_color_disabled ».
T_SUR_NEUTRE   = ("gray10", "gray95")
C_ACCENT       = ("#6d28d9", "#7c3aed")   # actions particulières : Habillage
C_ACCENT_SURV  = ("#5b21b6", "#6d28d9")

# — Échelle de SURFACES ————————————————————————————————————————————————
# Douze gris différents cohabitaient pour désigner quatre choses seulement.
# « gray85 » et « gray86 » servaient au même usage sans qu'on puisse dire
# lequel était le bon, et un panneau changeait de teinte selon l'onglet.
#
# Quatre niveaux suffisent, du plus proche du fond au plus détaché :
#   S_CARTE      un panneau posé sur le fond de la fenêtre
#   S_LIGNE      une ligne de liste, un en-tête de tableau, un encart
#   S_SELECTION  l'élément actif ou sélectionné
#   S_PUCE       une pastille ou un aperçu à l'intérieur d'une ligne
#
# Toute nouvelle surface doit reprendre l'un de ces quatre niveaux. En ajouter
# un cinquième « juste pour cet écran » est précisément ce qui a produit les
# douze précédents.
# Le fond de la fenêtre et la barre latérale tombaient sur la MÊME teinte en
# mode clair (219,219,219), et `S_LIGNE` valait exactement le fond : d'où une
# impression de gris uniforme, où rien ne se détachait de rien.
#
# L'échelle suit désormais une élévation. En mode clair, une surface posée est
# plus CLAIRE que son support ; en mode sombre, plus claire aussi — mais on
# part du bas. Le fond de fenêtre est le niveau le plus bas des deux côtés.
S_FOND       = ("gray88", "gray11")   # fond de la fenêtre
S_BARRE      = ("gray92", "gray15")   # barre latérale
S_CARTE      = ("gray96", "gray18")   # panneau posé sur le fond
S_LIGNE      = ("gray91", "gray22")   # ligne de liste, en-tête, encart
S_LIGNE_ALT  = ("gray95", "gray19")   # zébrure : alterne avec S_LIGNE
S_SELECTION  = ("gray78", "gray30")   # élément actif
S_PUCE       = ("gray84", "gray26")   # pastille ou aperçu dans une ligne
# Un filet de séparation : « gray30 » était écrit seul, donc appliqué tel quel
# dans les DEUX modes — un trait presque noir en travers d'un panneau clair.
S_FILET      = ("gray80", "gray32")

# — Contrôles de saisie ————————————————————————————————————————————————
# Les listes déroulantes prenaient le bleu par défaut de CustomTkinter, c'est-
# à-dire EXACTEMENT la teinte des boutons d'action. Sur l'onglet Vidéos, cinq
# filtres, le bouton « Rafraîchir » et le bouton « Appliquer » criaient donc
# aussi fort : l'œil n'avait aucun point d'entrée.
#
# Un filtre n'est pas une action, c'est un réglage. Il doit se voir sans
# appeler. D'où ces teintes neutres, plus claires que C_NEUTRE (réservé aux
# boutons secondaires) pour rester distinguables d'un bouton désactivé.
C_CHAMP        = ("gray80", "gray30")     # fond d'une liste déroulante
C_CHAMP_BOUTON = ("gray70", "gray38")     # sa partie cliquable (la flèche)
C_CHAMP_SURV   = ("gray62", "gray45")
T_CHAMP        = ("gray10", "gray90")     # texte d'un contrôle

# Style complet d'une liste déroulante, à déplier avec ** dans chaque appel :
#     ctk.CTkOptionMenu(parent, values=[...], **STYLE_CHAMP)
# Passer par un dictionnaire unique évite qu'un nouveau filtre ajouté plus tard
# reprenne le bleu par défaut sans que personne ne le remarque — un test vérifie
# qu'aucun CTkOptionMenu n'est créé sans lui.
STYLE_CHAMP = {
    "fg_color": C_CHAMP,
    "button_color": C_CHAMP_BOUTON,
    "button_hover_color": C_CHAMP_SURV,
    "text_color": T_CHAMP,
}

# Une zone de liste (CTkComboBox) est ÉDITABLE : son fond doit rester clair,
# comme un champ de saisie. Seule sa flèche est neutralisée.
STYLE_ZONE = {
    "button_color": C_CHAMP_BOUTON,
    "button_hover_color": C_CHAMP_SURV,
}

# — Couleurs de TEXTE ——————————————————————————————————————————————————
# `gray` (3,93:1) passait sous le minimum de lisibilité WCAG AA (4,5:1) et
# était pourtant la teinte secondaire la plus employée : 101 occurrences.
T_SECONDAIRE   = ("gray45", "gray70")     # mesuré à 7,40:1 sur fond sombre
T_DISCRET      = ("gray50", "gray60")     # mentions de bas de panneau
T_SUCCES       = ("#15803d", "#22c55e")
T_ALERTE       = ("#b45309", "#f59e0b")
T_ERREUR       = ("#b91c1c", "#ef4444")

# — Tailles de police ——————————————————————————————————————————————————
# Les tailles allaient de 9 à 26 px sans échelle. Cinq niveaux suffisent.
T_TITRE   = 20    # titre d'onglet
T_SOUS    = 16    # titre de section
T_CORPS   = 13    # texte courant
T_PETIT   = 12    # libellés de formulaire
T_MINI    = 11    # mentions, compteurs, aide en ligne

# Page Moodle où les enseignants téléchargent l'application et le tutoriel.
# Reprise dans le message de délivrance du jeton (onglet Comptes).
MOODLE_URL = "https://moodle.utoulouse.fr/course/section.php?id=72329"

# Adresse du support, mise en COPIE CACHÉE des messages de délivrance de jeton :
# cela garde une trace de l'envoi sans exposer l'adresse au destinataire.
SUPPORT_MAIL = "support-pod@utoulouse.fr"


# ════════════════════════════════════════════════════════════════════════════
#  ALIGNEMENT DES LIBELLÉS DE NAVIGATION
# ════════════════════════════════════════════════════════════════════════════

def _prefixe_aligne(police, icone: str, colonne: int) -> str:
    """Renvoie l'icône suivie de l'espacement qui amène le texte à `colonne`.

    Les glyphes d'icônes n'ont pas tous la même largeur. Un espacement en dur
    — trois espaces après chaque icône — décalait donc les libellés jusqu'à
    11 px les uns des autres. Pire : cette largeur dépend de la police du
    système, si bien que les entrées fautives changeaient d'un poste à
    l'autre. Aucun réglage figé ne pouvait convenir.

    On mesure donc le glyphe DANS LA POLICE RÉELLEMENT UTILISÉE, puis on
    complète avec des espaces ordinaires et, pour le reliquat, des espaces
    fins (U+2009) — un espace ordinaire étant trop large pour ajuster
    finement. L'écart résiduel tombe à 2 px.

    `police` doit être l'instance passée au bouton : mesurer avec une autre
    police donnerait un résultat faux.
    """
    try:
        large = police.measure(" ") or 4
        # Certaines polices ne possèdent pas l'espace fin et renvoient 0, ce
        # qui provoquerait une division par zéro : on se rabat sur l'espace
        # ordinaire, l'alignement restant meilleur qu'avec un espacement figé.
        fine = police.measure("\u2009") or large
        manque = max(large, colonne - police.measure(icone))
        entiers = int(manque // large)
        reliquat = manque - entiers * large
        fins = int(round(reliquat / fine)) if fine else 0
        return icone + " " * entiers + "\u2009" * fins
    except Exception:
        # Un défaut d'alignement ne doit jamais empêcher la barre de se
        # construire.
        return icone + "   "


def bloc_filtre(parent, libelle: str):
    """Crée un conteneur portant `libelle` AU-DESSUS ; renvoie ce conteneur.

    Les quatre premiers filtres portaient leur intitulé DANS leur valeur par
    défaut (« Tous statuts », « Toutes chaînes »). Dès qu'on filtrait, ces mots
    disparaissaient : l'écran affichait « Public », « MFCA », « Cours » sans
    plus rien dire de ce que chaque valeur filtrait. L'information se perdait
    exactement au moment où le filtre devenait actif.

    Le libellé ne peut pas être posé À GAUCHE du menu : mesuré, cela coûterait
    139 px sur une rangée qui n'en a que 3 de marge en fenêtre minimale, et le
    champ de recherche serait écrasé. Au-dessus, il ne coûte que ~16 px de
    hauteur et vaut pour les deux rangées, ce qui uniformise du même coup les
    deux conventions qui cohabitaient.
    """
    bloc = ctk.CTkFrame(parent, fg_color="transparent")
    ctk.CTkLabel(bloc, text=libelle, font=ctk.CTkFont(size=10),
                 text_color=T_DISCRET, anchor="w").pack(fill="x", padx=2)
    return bloc


# ════════════════════════════════════════════════════════════════════════════
#  INFOBULLE
# ════════════════════════════════════════════════════════════════════════════

def ajouter_infobulle(widget, texte: str, delai_ms: int = 450):
    """Affiche `texte` dans une petite étiquette au survol prolongé de `widget`.

    Nécessaire dès qu'un bouton perd son libellé au profit d'une icône seule :
    une poubelle se comprend, mais rien n'annonce alors ce qui sera supprimé.
    Le survol rend le mot, sans le laisser occuper la ligne en permanence.

    Le délai évite que l'étiquette clignote quand la souris ne fait que
    traverser la rangée pour atteindre un autre bouton.

    Prudence d'implémentation :
      • la fenêtre est créée au survol et DÉTRUITE à la sortie, jamais gardée
        en attente — un conteneur vide en réserve est précisément ce qui se
        dessinait en carré noir sur macOS ;
      • aucune transparence n'est employée, son rendu variant selon le système ;
      • tout est encadré par `try`, une infobulle ne devant jamais empêcher un
        clic.
    """
    etat = {"fenetre": None, "apres": None}

    def _montrer():
        etat["apres"] = None
        try:
            if etat["fenetre"] is not None or not widget.winfo_exists():
                return
            fen = ctk.CTkToplevel(widget)
            fen.overrideredirect(True)          # ni titre ni bordure
            fen.attributes("-topmost", True)
            cadre = ctk.CTkFrame(fen, fg_color=S_SELECTION, corner_radius=4)
            cadre.pack()
            ctk.CTkLabel(cadre, text=texte, font=ctk.CTkFont(size=T_MINI),
                         text_color=T_CHAMP).pack(padx=8, pady=3)
            # Sous le widget, légèrement décalée : au-dessus, le curseur la
            # masquerait.
            #
            # Le décalage horizontal est BORNÉ à l'écran : ces boutons sont en
            # bout de rangée, donc collés au bord droit. Aligné naïvement sur
            # le bord gauche du bouton, le texte sortait de l'écran et se
            # lisait « Supprimer ce… ».
            fen.update_idletasks()
            largeur = fen.winfo_reqwidth()
            x = widget.winfo_rootx()
            marge = 8
            debord = x + largeur - widget.winfo_screenwidth() + marge
            if debord > 0:
                x -= debord
            fen.geometry(f"+{max(marge, x)}"
                         f"+{widget.winfo_rooty() + widget.winfo_height() + 4}")
            etat["fenetre"] = fen
        except Exception:
            etat["fenetre"] = None

    def _cacher(_evt=None):
        if etat["apres"] is not None:
            try:
                widget.after_cancel(etat["apres"])
            except Exception:
                pass
            etat["apres"] = None
        if etat["fenetre"] is not None:
            try:
                etat["fenetre"].destroy()
            except Exception:
                pass
            etat["fenetre"] = None

    def _entrer(_evt=None):
        _cacher()
        try:
            etat["apres"] = widget.after(delai_ms, _montrer)
        except Exception:
            pass

    widget.bind("<Enter>", _entrer, add="+")
    widget.bind("<Leave>", _cacher, add="+")
    # Un clic ouvre une fenêtre de confirmation : l'infobulle ne doit pas
    # rester posée par-dessus.
    widget.bind("<Button-1>", _cacher, add="+")
    widget.bind("<Destroy>", _cacher, add="+")
    return widget


# ════════════════════════════════════════════════════════════════════════════
#  MODÈLE : une entrée de la file d'attente
# ════════════════════════════════════════════════════════════════════════════

class UploadItem:
    """Une vidéo dans la file de téléversement (fichier, titre, état, slug, URL)."""
    def __init__(self, path: str):
        """Crée une entrée de la file d'upload à partir d'un chemin de fichier."""
        self.path = path
        self.filename = os.path.basename(path)
        # Titre par défaut = nom de fichier sans extension, nettoyé
        base = os.path.splitext(self.filename)[0]
        self.title = base.replace("_", " ").replace("-", " ").strip()
        self.status = "en attente"     # libellé AFFICHÉ (peut contenir un émoji)
        # Indicateur de réussite, distinct du libellé affiché. Comparer le texte
        # du statut était fragile : le libellé posé après un envoi réussi est
        # « ✅ terminé », alors que le test portait sur « terminé » — l'égalité
        # échouait toujours, et une vidéo déjà envoyée repartait à chaque clic
        # sur « Lancer le téléversement », créant des doublons.
        self.done = False
        self.slug = ""
        self.video_url = ""
        self.error = ""
        # widgets (remplis à l'affichage)
        self.row = None
        self.title_var = None
        self.status_lbl = None


# ════════════════════════════════════════════════════════════════════════════
#  APPLICATION
# ════════════════════════════════════════════════════════════════════════════

# Base conditionnelle : mixe le moteur de glisser-déposer si disponible
if HAS_DND:
    class _AppBase(ctk.CTk, TkinterDnD.DnDWrapper):
        """Classe de base de la fenêtre (avec glisser-déposer si disponible)."""
        pass
else:
    class _AppBase(ctk.CTk):
        """Classe de base de la fenêtre (avec glisser-déposer si disponible)."""
        pass


class App(_AppBase):
    def __init__(self):
        """Initialise la fenêtre, charge config + token, construit l'UI et tente une connexion auto."""
        super().__init__()
        # Initialiser le moteur glisser-déposer (tkdnd)
        self.dnd_ok = False
        if HAS_DND:
            try:
                self.TkdndVersion = TkinterDnD._require(self)
                self.dnd_ok = True
            except Exception:
                self.dnd_ok = False

        self.title(APP_TITLE)
        self.geometry("1180x760")
        self.minsize(1000, 660)
        # Fond de la fenêtre : niveau le plus bas de l'échelle d'élévation.
        # Sans cela, CustomTkinter donne au fond la teinte par défaut d'un
        # cadre, c'est-à-dire exactement celle de la barre latérale.
        self.configure(fg_color=S_FOND)

        self.config_data = cfg.load_config()
        self.token = cfg.load_token()
        # Identifiants du compte VÉHICULE (session web pour le chunké des gros
        # fichiers). Deux sources, dans cet ordre :
        #   1. ceux saisis dans l'onglet Configuration (coffre-fort de l'OS) —
        #      prioritaires, pour pouvoir utiliser un autre compte au besoin ;
        #   2. à défaut, le compte DEPOT embarqué dans config.py → la bascule
        #      sur les gros fichiers fonctionne SANS aucune saisie préalable.
        self.vehicle_username, self.vehicle_password = cfg.load_vehicle_credentials()
        if not (self.vehicle_username and self.vehicle_password):
            self.vehicle_username = getattr(cfg, "VEHICLE_USERNAME", "")
            self.vehicle_password = getattr(cfg, "VEHICLE_PASSWORD", "")
        self.vehicle_owner_url = ""      # URL Pod du véhicule (résolue à la connexion)
        self.api: PodAPI | None = None

        self.types: list[dict] = []
        self.type_map: dict[str, str] = {}     # titre → url

        # ══ MAGASIN DE VIDÉOS PARTAGÉ ══════════════════════════════════════
        # Source de vérité UNIQUE pour les onglets qui manipulent des vidéos
        # (Vidéos, Explorateur, Chaînes). Auparavant chaque onglet chargeait sa
        # propre liste : la même vidéo existait en plusieurs exemplaires, et
        # modifier l'un ne modifiait pas les autres (d'où les désynchronisations).
        # Avec une liste unique, ce problème devient impossible : il n'y a qu'un
        # seul objet par vidéo, partagé par tous les onglets.
        self.videos: list[dict] = []            # la liste unique
        self.videos_loaded_at = None            # datetime du dernier chargement
        self.videos_loading = False             # un scan est-il déjà en cours ?
        self._videos_waiters: list = []         # callbacks à servir en fin de scan
        # VERROU du magasin. Plusieurs threads peuvent vouloir remplir ou muter
        # `self.videos` en même temps (l'onglet Chaînes charge en synchrone
        # pendant qu'un scan asynchrone tourne, un lot supprime des vidéos…).
        # Sans verrou, deux scans complets pouvaient s'écrire l'un sur l'autre.
        self._videos_lock = threading.RLock()
        self.site_urls: list[str] = []         # sites (requis à l'upload)
        self.access_groups: list[dict] = []    # groupes d'accès {code_name, display_name, url}
        # Vrai une fois l'onglet Groupes réellement chargé (groupes + table des
        # comptes propriétaires). `access_groups` ne suffit pas comme indicateur :
        # la connexion le remplit déjà, sans charger le reste ni afficher.
        self._groups_loaded = False
        self._auto_loaded: set = set()         # onglets déjà auto-chargés cette session
        self.items: list[UploadItem] = []
        self.all_users: list[dict] = []        # liste complète Pod (pour sélection owner)
        self.additional_owner_urls: list[str] = []
        self.upload_owner_url: str = ""        # propriétaire explicite du lot (obligatoire)
        self.upload_owner_label: str = ""
        self.additional_owner_map: dict[str, str] = {}   # url → libellé (pour ré-ouverture)

        self._build_ui()
        self._show_tab("upload")

        # Connexion auto si token déjà présent
        if self.config_data.get("url") and self.token:
            self._run(self._auto_connect)

        # Vérification d'une nouvelle version, différée de 2 secondes pour ne
        # pas concurrencer le chargement initial, puis menée en arrière-plan.
        self.after(2000, self._verifier_maj)

    # ── Threading helpers ────────────────────────────────────────────────

    def _run(self, fn, *a):
        """Lance une fonction dans un thread d'arrière-plan (pour ne pas geler l'interface)."""
        threading.Thread(target=fn, args=a, daemon=True).start()

    def _ui(self, fn, *a, **kw):
        """Planifie une mise à jour d'interface dans le thread principal Tk (thread-safe)."""
        self.after(0, lambda: fn(*a, **kw))

    # ── Construction de l'interface ──────────────────────────────────────

    def _build_ui(self):
        """Construit la barre latérale (logo, état, navigation) et la zone de contenu."""
        # Sidebar
        self.sidebar = ctk.CTkFrame(self, width=220, corner_radius=0,
                                    fg_color=S_BARRE)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        # En-tête : logo Université de Toulouse sur bandeau blanc (repli texte si absent)
        logo_loaded = False
        if HAS_PIL:
            try:
                logo_path = resource_path(os.path.join("assets", "logo_ut.png"))
                if os.path.exists(logo_path):
                    pil = PILImage.open(logo_path)
                    W = 156
                    H = round(W * pil.height / pil.width)
                    self.logo_img = ctk.CTkImage(light_image=pil, dark_image=pil, size=(W, H))
                    card = ctk.CTkFrame(self.sidebar, fg_color="white", corner_radius=8)
                    card.pack(padx=12, pady=(10, 4), fill="x")
                    ctk.CTkLabel(card, image=self.logo_img, text="").pack(padx=8, pady=7)
                    logo_loaded = True
            except Exception:
                logo_loaded = False

        if not logo_loaded:
            ctk.CTkLabel(self.sidebar, text="Université de Toulouse",
                         font=ctk.CTkFont(size=13, weight="bold")).pack(pady=(20, 0), padx=14)

        ctk.CTkLabel(self.sidebar, text="🛠️  PodAdmin",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(2, 0), padx=14)

        # État connexion
        box = ctk.CTkFrame(self.sidebar, fg_color=S_LIGNE,
                           corner_radius=8)
        box.pack(padx=12, pady=14, fill="x")
        self.status_dot = ctk.CTkLabel(box, text="⚫", font=ctk.CTkFont(size=13))
        self.status_dot.pack(side="left", padx=8, pady=6)
        self.status_lbl = ctk.CTkLabel(box, text="Non connecté",
                                       font=ctk.CTkFont(size=11), text_color=T_SECONDAIRE)
        self.status_lbl.pack(side="left")

        # Agent identifié — « Dépôt au nom de : … »
        #
        # Ce libellé reste VIDE tant qu'aucun dépôt délégué n'est choisi, c'est-
        # à-dire la plupart du temps. Il occupait pourtant ses 34 px en
        # permanence, pris directement sur la zone de navigation. Il n'est donc
        # placé que lorsqu'il a quelque chose à dire (voir `_definir_agent`).
        self.agent_lbl = ctk.CTkLabel(self.sidebar, text="", font=ctk.CTkFont(size=11),
                                      text_color=T_SECONDAIRE, wraplength=190, justify="left")
        self.agent_visible = False

        self.sidebar_separateur = ctk.CTkFrame(self.sidebar, height=1, fg_color=S_FILET)
        self.sidebar_separateur.pack(fill="x", padx=12, pady=3)

        # Version épinglée en bas (hors zone défilante)
        ctk.CTkLabel(self.sidebar, text=f"v{APP_VERSION}",
                     font=ctk.CTkFont(size=9), text_color=T_DISCRET).pack(side="bottom", pady=(0, 8))

        # Bascule clair / sombre, juste au-dessus du numéro de version.
        # Le choix est mémorisé : l'application rouvre dans le mode retenu.
        self.theme_btn = ctk.CTkButton(
            self.sidebar, text="", width=150, height=26,
            font=ctk.CTkFont(size=11), fg_color=C_NEUTRE,
            hover_color=C_NEUTRE_SURV, command=self._basculer_theme, text_color=T_SUR_NEUTRE)
        self.theme_btn.pack(side="bottom", pady=(6, 2))
        self._maj_libelle_theme()

        # Bandeau « nouvelle version disponible » : AUCUN widget n'est créé ici.
        #
        # Une première version utilisait un cadre conteneur transparent de
        # hauteur nulle, inséré d'avance. Deux ennuis en ont découlé : sur
        # Windows il amputait la barre latérale, et sur macOS un cadre
        # transparent de taille nulle se dessinait en CARRÉ NOIR.
        #
        # Le bandeau est donc créé de toutes pièces au moment où une mise à jour
        # est détectée, puis détruit ensuite : pas de conteneur intermédiaire,
        # donc pas de comportement dépendant du système.
        self.maj_bandeau = None

        # Zone de navigation DÉFILANTE : sur petit écran, les onglets défilent
        # au lieu de déborder hors de la fenêtre (le haut et le bas restent fixes).
        nav_scroll = ctk.CTkScrollableFrame(self.sidebar, fg_color="transparent",
                                            width=200)
        nav_scroll.pack(side="top", fill="both", expand=True, padx=0, pady=0)

        # Navigation groupée par MÉTIER plutôt qu'en liste continue.
        #
        # Douze entrées à plat obligeaient l'œil à tout relire pour trouver la
        # bonne, et la dernière (« À propos ») sortait de l'écran sur un
        # portable en 800 px de haut. Quatre blocs de trois ou quatre entrées se
        # parcourent d'un coup d'œil.
        #
        # L'ordre suit la fréquence d'usage : « Vidéos » est l'onglet du
        # quotidien, il remonte ; « Comptes » sert à délivrer un jeton de temps
        # en temps, il rejoint le bloc des accès.
        #
        # Les boutons passent de 40 à 32 px : les intitulés de section ajoutent
        # de la hauteur, cette réduction la compense et fait tenir l'ensemble
        # sans défilement dans une fenêtre courante.
        #
        # ICÔNE ET LIBELLÉ SONT SÉPARÉS, et l'espacement calculé à l'exécution.
        # Auparavant chaque entrée portait trois espaces en dur après son
        # icône : les glyphes n'ayant pas la même largeur, les libellés se
        # décalaient jusqu'à 11 px les uns des autres. Et comme cette largeur
        # dépend de la police du système, les entrées fautives n'étaient même
        # pas les mêmes d'un poste à l'autre — donc impossible à corriger par
        # un réglage figé.
        #
        # `_prefixe_aligne` mesure le glyphe et complète jusqu'à une colonne
        # fixe, ce qui ramène l'écart à 2 px sur toute plateforme.
        COLONNE_ICONE = 34      # largeur réservée à l'icône, en pixels
        NAVIGATION = [
            ("CONTENUS", [
                ("📂", "Téléversement", "upload"),
                ("🎬", "Encodage", "encode"),
                ("🎞️", "Vidéos", "browse"),
                ("🔄", "Réaffectation", "reassign"),
            ]),
            ("ORGANISATION", [
                ("📺", "Chaînes", "ct"),
                ("📊", "Inventaire", "stats"),
            ]),
            ("ACCÈS", [
                ("👤", "Comptes", "comptes"),
                ("🔐", "Groupes d'accès", "groups"),
            ]),
            ("SYSTÈME", [
                ("⚙️", "Configuration", "config"),
                ("📋", "Journal", "log"),
            ]),
        ]

        # « Aide » et « À propos » sortent du flux défilant, épinglés en pied
        # sur UNE SEULE LIGNE à deux colonnes.
        #
        # La navigation débordait à la taille par défaut (1180×760) : « À
        # propos » était hors champ dès l'ouverture, pour tout le monde. Le
        # défaut avait échappé aux mesures précédentes, faites à tort en
        # 1280×800.
        #
        # Les épingler sans les compacter n'aurait rien rapporté — ils
        # occuperaient la même hauteur ailleurs. C'est la mise sur une seule
        # ligne qui rend les 32 px : ce sont les deux entrées les moins
        # consultées, et deux mots courts y tiennent sans être tronqués.
        EPINGLES = [("❓", "Aide", "help"), ("ℹ️", "À propos", "about")]

        self.nav_btns = {}
        for i, (section, entrees) in enumerate(NAVIGATION):
            # Intitulé de section : volontairement discret (petit, gris). Il
            # sépare sans attirer l'œil — ce n'est pas un élément cliquable et
            # il ne doit pas se confondre avec un onglet.
            ctk.CTkLabel(nav_scroll, text=section, anchor="w",
                         font=ctk.CTkFont(size=10, weight="bold"),
                         text_color=T_DISCRET).pack(
                             fill="x", padx=10, pady=((6 if i else 2), 0))
            for icone, libelle, key in entrees:
                police = ctk.CTkFont(size=13)
                b = ctk.CTkButton(
                    nav_scroll,
                    text=_prefixe_aligne(police, icone, COLONNE_ICONE) + libelle,
                    anchor="w", height=32,
                    fg_color="transparent", text_color=("gray10", "gray90"),
                    hover_color=("gray75", "gray28"),
                    font=police,
                    command=lambda k=key: self._show_tab(k))
                b.pack(fill="x", padx=4)
                self.nav_btns[key] = b

        # Rangée épinglée, posée AVANT le bouton de thème dans l'ordre `bottom`
        # pour se retrouver juste au-dessus de lui.
        rangee_bas = ctk.CTkFrame(self.sidebar, fg_color="transparent", height=30)
        rangee_bas.pack(side="bottom", fill="x", padx=4, pady=(2, 0))
        # GRID et non pack : avec `expand=True, fill="x"`, chaque bouton
        # conservait sa largeur naturelle (140 px pour CTkButton par défaut),
        # si bien que « Aide » occupait 140 px et « À propos » 68 — deux
        # colonnes visiblement inégales. `uniform` force la stricte moitié.
        rangee_bas.columnconfigure(0, weight=1, uniform="epingles")
        rangee_bas.columnconfigure(1, weight=1, uniform="epingles")
        for colonne, (icone, libelle, key) in enumerate(EPINGLES):
            police = ctk.CTkFont(size=12)
            b = ctk.CTkButton(
                rangee_bas,
                # Même préfixe calculé que dans le flux : sans lui, ces deux
                # libellés étaient CENTRÉS alors que les dix autres sont
                # alignés à gauche — l'œil voyait un décalage.
                # Colonne plus étroite qu'en flux : sur une demi-largeur, l'écart
                # de 34 px laissait l'icône seule à gauche et le mot très loin.
                text=_prefixe_aligne(police, icone, COLONNE_ICONE - 14) + libelle,
                anchor="w", width=1, height=28,
                fg_color="transparent", text_color=("gray10", "gray90"),
                hover_color=("gray75", "gray28"), font=police,
                command=lambda k=key: self._show_tab(k))
            b.grid(row=0, column=colonne, sticky="ew", padx=1)
            # Ils restent dans `nav_btns` : le surlignage de l'onglet actif et
            # tout le reste continuent de fonctionner sans cas particulier.
            self.nav_btns[key] = b


        # Zone principale
        # Filet de séparation entre la barre latérale et le contenu.
        #
        # L'élévation seule ne suffit pas en mode CLAIR : l'écart barre/fond
        # est de 11 niveaux de gris, soit 26 % en sombre mais seulement 4,7 %
        # en clair. À forte luminance, l'œil exige un écart bien plus grand
        # pour percevoir la même différence — la frontière ne se lisait donc
        # pas. Un filet de 1 px pose la limite sans toucher à l'échelle.
        ctk.CTkFrame(self, width=1, corner_radius=0,
                     fg_color=S_FILET).pack(side="left", fill="y")

        self.content = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.content.pack(side="right", fill="both", expand=True, padx=14, pady=14)

        self.tabs = {}
        self._build_tab_upload()
        self._build_tab_encode()
        self._build_tab_comptes()
        self._build_tab_browse()
        self._build_tab_reassign()
        self._build_tab_stats()
        self._build_tab_ct()
        self._build_tab_groups()
        self._build_tab_config()
        self._build_tab_help()
        self._build_tab_log()
        self._build_tab_about()

    def _basculer_theme(self):
        """Passe du mode sombre au mode clair, et inversement.

        Les couleurs de la palette sont des COUPLES (clair, sombre) :
        CustomTkinter choisit la bonne teinte automatiquement, sans qu'aucun
        widget n'ait à être reconstruit. C'est la raison d'être des couples —
        une couleur unique resterait identique dans les deux modes."""
        nouveau = "light" if ctk.get_appearance_mode().lower() == "dark" else "dark"
        ctk.set_appearance_mode(nouveau)
        self.config_data["theme"] = nouveau
        try:
            cfg.save_config(self.config_data)      # mémorisé pour la prochaine fois
        except Exception:
            pass
        self._maj_libelle_theme()
        self._log(f"Thème : mode {'clair' if nouveau == 'light' else 'sombre'}.")

    def _maj_libelle_theme(self):
        """Le bouton annonce le mode vers lequel il fait basculer."""
        sombre = ctk.get_appearance_mode().lower() == "dark"
        try:
            self.theme_btn.configure(text="☀  Mode clair" if sombre else "🌙  Mode sombre")
        except Exception:
            pass

    def _show_tab(self, key: str):
        """Affiche l'onglet `key` et met en surbrillance son bouton de navigation."""
        for f in self.tabs.values():
            f.pack_forget()
        self.tabs[key].pack(fill="both", expand=True)
        for k, b in self.nav_btns.items():
            b.configure(fg_color=S_SELECTION if k == key else "transparent")
        # Chargement paresseux à la première ouverture de l'onglet Groupes
        if key == "groups":
            self._show_groups_tab_hook()
        # Auto-chargement des onglets à listes, à leur PREMIÈRE ouverture de la
        # session. Ensuite le cache est réutilisé (affichage instantané) ; le
        # bouton « 🔄 Rafraîchir » de chaque onglet force une relecture serveur
        # (utile si des modifs ont été faites via le site web Pod).
        # On MÉMORISE l'onglet courant : si l'utilisateur ouvre un onglet à
        # liste AVANT que la connexion (asynchrone, 1 à 5 s) n'ait abouti, le
        # chargement automatique ne peut pas partir. Sans cette mémoire, il ne
        # partirait JAMAIS tant qu'il reste sur cet onglet — l'écran resterait
        # vide, ce qui ressemble à une panne de l'instance.
        # `_on_connexion_etablie()` rejoue alors le chargement.
        self.onglet_courant = key
        self._auto_charger(key)

    # Onglets qui chargent leur liste à la première ouverture. Ensuite le cache
    # est réutilisé ; le bouton « 🔄 Rafraîchir » force une relecture serveur.
    def _auto_loaders(self) -> dict:
        """Table des chargements automatiques, par onglet."""
        return {
            "encode": self._encode_scan,
            "browse": self._browse_load,
            "stats":  self._stats_scan,
            "ct":     self._ct_load,
        }

    def _auto_charger(self, key: str):
        """Déclenche le chargement automatique d'un onglet, si possible.

        Ne fait rien tant que la connexion n'est pas établie : c'est
        `_on_connexion_etablie()` qui rappellera cette méthode."""
        auto = self._auto_loaders()
        if key in auto and self.api and key not in self._auto_loaded:
            self._auto_loaded.add(key)
            auto[key]()

    def _on_connexion_etablie(self):
        """Appelée dès que la connexion à l'instance aboutit.

        Rejoue le chargement de l'onglet actuellement affiché : sans cela, un
        onglet ouvert pendant la connexion resterait définitivement vide."""
        courant = getattr(self, "onglet_courant", None)
        if courant:
            self._auto_charger(courant)

    # ═════════════════════════════════════════════════════════════════════
    #  ONGLET TÉLÉVERSEMENT
    # ═════════════════════════════════════════════════════════════════════

    def _build_tab_upload(self):
        """Construit l'onglet Téléversement (sélection, réglages communs, liste, lancement)."""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tabs["upload"] = frame

        ctk.CTkLabel(frame, text="📂  Téléversement par lot",
                     font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=(0, 10))

        # — Barre de sélection —
        sel = ctk.CTkFrame(frame, fg_color="transparent")
        sel.pack(fill="x", pady=(0, 6))

        ctk.CTkButton(sel, text="➕  Ajouter des fichiers", width=190,
                      fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=self._add_files, text_color=T_SUR_NEUTRE).pack(side="left", padx=(0, 8))
        ctk.CTkButton(sel, text="📁  Ajouter un dossier", width=190,
                      fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=self._add_folder, text_color=T_SUR_NEUTRE).pack(side="left", padx=(0, 8))
        ctk.CTkButton(sel, text="🗑  Vider la liste", width=140,
                      fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=self._clear_items, text_color=T_SUR_NEUTRE).pack(side="left")

        self.count_lbl = ctk.CTkLabel(sel, text="0 vidéo(s)", text_color=T_SECONDAIRE,
                                      font=ctk.CTkFont(size=11))
        self.count_lbl.pack(side="right")

        # — Réglages communs (appliqués à tout le lot) —
        common = ctk.CTkFrame(frame, fg_color=S_CARTE)
        common.pack(fill="x", pady=(0, 8))

        ctk.CTkLabel(common, text="Réglages communs au lot",
                     font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=4,
                                                           padx=12, pady=(10, 4), sticky="w")

        ctk.CTkLabel(common, text="Type :").grid(row=1, column=0, padx=(12, 4), pady=8, sticky="e")
        self.type_combo = ctk.CTkComboBox(common, values=["(chargement…)"], width=200,
                                          **STYLE_ZONE)
        self.type_combo.grid(row=1, column=1, padx=4, pady=8, sticky="w")

        ctk.CTkLabel(common, text="Visibilité :").grid(row=1, column=2, padx=(20, 4), pady=8, sticky="e")
        self.visibility_combo = ctk.CTkComboBox(
            common, width=200, values=["Brouillon / Privé", "Public"],
            **STYLE_ZONE)
        self.visibility_combo.set("Brouillon / Privé")
        self.visibility_combo.grid(row=1, column=3, padx=4, pady=8, sticky="w")

        self.encode_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(common, text="Lancer l'encodage après le téléversement",
                        variable=self.encode_var).grid(row=2, column=0, columnspan=2,
                                                        padx=12, pady=(0, 6), sticky="w")

        # Propriétaire des vidéos (OBLIGATOIRE — choix explicite avant l'envoi)
        ctk.CTkLabel(common, text="Propriétaire :").grid(row=4, column=0, padx=(12, 4),
                                                         pady=8, sticky="e")
        ctk.CTkButton(common, text="🎯  Choisir le propriétaire…", width=200,
                      fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=self._upload_pick_owner, text_color=T_SUR_NEUTRE).grid(row=4, column=1, padx=4,
                                                            pady=8, sticky="w")
        self.upload_owner_lbl = ctk.CTkLabel(
            common, text="⚠️  à définir avant l'envoi", text_color=T_ALERTE,
            font=ctk.CTkFont(size=11, weight="bold"))
        self.upload_owner_lbl.grid(row=4, column=2, columnspan=2, padx=12, pady=8, sticky="w")

        # Propriétaires additionnels communs
        ctk.CTkButton(common, text="👥  Propriétaires additionnels…", width=240,
                      fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=self._edit_additional_owners, text_color=T_SUR_NEUTRE).grid(
            row=2, column=2, columnspan=2, padx=12, pady=(0, 6), sticky="w")
        self.add_owners_lbl = ctk.CTkLabel(common, text="aucun", text_color=T_SECONDAIRE,
                                           font=ctk.CTkFont(size=11))
        self.add_owners_lbl.grid(row=3, column=2, columnspan=2, padx=12, pady=(0, 8), sticky="w")

        common.columnconfigure(3, weight=1)

        # — Tableau des vidéos (titres éditables) —
        hint = ("Vérifiez / corrigez les titres avant l'envoi  —  "
                "💡 vous pouvez aussi glisser-déposer fichiers et dossiers ci-dessous :"
                if getattr(self, "dnd_ok", False) else
                "Vérifiez / corrigez les titres avant l'envoi :")
        ctk.CTkLabel(frame, text=hint, font=ctk.CTkFont(size=12)).pack(anchor="w", pady=(2, 2))

        self.list_frame = ctk.CTkScrollableFrame(frame, height=240, fg_color=S_CARTE)
        self.list_frame.pack(fill="both", expand=True)

        # Activer le glisser-déposer sur la zone de liste
        if getattr(self, "dnd_ok", False):
            try:
                self.list_frame.drop_target_register(DND_FILES)
                self.list_frame.dnd_bind("<<Drop>>", self._on_drop)
            except Exception:
                pass

        empty_text = ("Aucune vidéo.\nGlissez-déposez ici des fichiers ou des dossiers,\n"
                      "ou utilisez les boutons ci-dessus."
                      if getattr(self, "dnd_ok", False) else
                      "Aucune vidéo.\nUtilisez « Ajouter des fichiers » ou « Ajouter un dossier ».")
        self._empty_hint = ctk.CTkLabel(self.list_frame, text=empty_text, text_color=T_SECONDAIRE)
        self._empty_hint.pack(pady=40)

        # — Lancement + progression —
        launch = ctk.CTkFrame(frame, fg_color="transparent")
        launch.pack(fill="x", pady=(8, 0))

        self.launch_btn = ctk.CTkButton(
            launch, text="🚀  Lancer le téléversement", height=40,
            fg_color=C_SUCCES, hover_color=C_SUCCES_SURV,
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._start_upload)
        self.launch_btn.pack(side="left")

        # Relancer uniquement les vidéos en échec (sans re-sélectionner le lot).
        # Masqué tant qu'il n'y a pas d'échec à relancer.
        self.retry_btn = ctk.CTkButton(
            launch, text="🔄  Relancer les échecs", height=40,
            fg_color=C_ALERTE, hover_color=C_ALERTE_SURV,
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._retry_failed)
        # (pas de .pack ici : affiché seulement s'il y a des échecs)

        self.global_msg = ctk.CTkLabel(launch, text="", text_color=T_SECONDAIRE,
                                       font=ctk.CTkFont(size=12))
        self.global_msg.pack(side="left", padx=14)

        # — Progression —
        #
        # Les deux barres restaient affichées en permanence, à zéro : deux
        # traits inertes en bas de l'écran, qui n'informaient de rien tant
        # qu'aucun envoi n'était en cours.
        #
        # Elles sont donc CRÉÉES ici mais pas placées : `_afficher_progression()`
        # les fait apparaître au lancement du lot, `_masquer_progression()` les
        # retire à la fin.
        #
        # ⚠️ Les widgets sont créés une fois pour toutes et masqués par
        # `pack_forget`. Surtout PAS de cadre conteneur vide en attente : un
        # CTkFrame vide occupe 200 px par défaut, et transparent en hauteur
        # nulle il se dessine en CARRÉ NOIR sur macOS. C'est exactement l'ennui
        # qu'avait connu le bandeau de mise à jour.
        self.file_progress = ctk.CTkProgressBar(frame)
        self.file_progress.set(0)
        self.file_progress_lbl = ctk.CTkLabel(frame, text="", text_color=T_SECONDAIRE,
                                              font=ctk.CTkFont(size=10))
        self.batch_progress = ctk.CTkProgressBar(frame, progress_color=C_SUCCES)
        self.batch_progress.set(0)
        self.progression_visible = False

    def _definir_agent(self, texte: str):
        """Renseigne le libellé « Dépôt au nom de : … » et le place s'il y a lieu.

        Un libellé vide est retiré de l'affichage plutôt que laissé en place :
        sa hauteur était prise sur la navigation, où elle manquait.

        Le widget n'est jamais détruit, seulement retiré : le recréer à chaque
        changement exposerait aux ennuis d'affichage déjà rencontrés sur macOS."""
        self.agent_lbl.configure(text=texte)
        if texte and not self.agent_visible:
            # `before` : sans lui, `pack` placerait le libellé en toute fin de
            # barre latérale, sous la navigation, et non à sa place logique.
            self.agent_lbl.pack(padx=14, pady=(0, 6), anchor="w",
                                before=self.sidebar_separateur)
            self.agent_visible = True
        elif not texte and self.agent_visible:
            self.agent_lbl.pack_forget()
            self.agent_visible = False

    def _afficher_progression(self):
        """Fait apparaître les deux barres de progression (début d'un lot).

        Idempotent : appelée deux fois, elle ne place rien en double."""
        if self.progression_visible:
            return
        self.file_progress.pack(fill="x", pady=(8, 0))
        self.file_progress_lbl.pack(anchor="w")
        self.batch_progress.pack(fill="x", pady=(4, 0))
        self.progression_visible = True

    def _masquer_progression(self):
        """Retire les barres de l'affichage (fin d'un lot).

        Les widgets ne sont pas détruits : ils resserviront au lot suivant, et
        les recréer exposerait aux ennuis de cadres transparents sur macOS."""
        if not self.progression_visible:
            return
        self.batch_progress.pack_forget()
        self.file_progress_lbl.pack_forget()
        self.file_progress.pack_forget()
        self.progression_visible = False

    # ── Ajout de fichiers / dossier ──────────────────────────────────────

    def _add_files(self):
        """Ouvre un sélecteur de fichiers et ajoute les vidéos choisies à la file."""
        paths = filedialog.askopenfilenames(
            title="Choisir des vidéos",
            filetypes=[("Vidéos", "*.mp4 *.mov *.avi *.mkv *.webm *.m4v *.wmv *.flv *.mpg *.mpeg"),
                       ("Tous les fichiers", "*.*")])
        self._add_paths(paths)

    def _add_folder(self):
        """Scanne un dossier (récursivement) et ajoute toutes les vidéos trouvées à la file."""
        folder = filedialog.askdirectory(title="Choisir un dossier de vidéos")
        if not folder:
            return
        found = []
        for root, _dirs, files in os.walk(folder):
            for name in files:
                if os.path.splitext(name)[1].lower() in cfg.VIDEO_EXTENSIONS:
                    found.append(os.path.join(root, name))
        if not found:
            self.global_msg.configure(text="Aucune vidéo trouvée dans ce dossier.",
                                      text_color=T_ALERTE)
            return
        self._add_paths(sorted(found))

    def _on_drop(self, event):
        """Glisser-déposer : ajoute les vidéos des fichiers/dossiers déposés."""
        try:
            paths = self.tk.splitlist(event.data)
        except Exception:
            paths = [event.data]
        found = []
        for p in paths:
            p = p.strip().strip("{}")
            if not p:
                continue
            if os.path.isdir(p):
                for root, _d, names in os.walk(p):
                    for n in names:
                        if os.path.splitext(n)[1].lower() in cfg.VIDEO_EXTENSIONS:
                            found.append(os.path.join(root, n))
            elif os.path.isfile(p) and os.path.splitext(p)[1].lower() in cfg.VIDEO_EXTENSIONS:
                found.append(p)
        if found:
            self._show_tab("upload")
            self._add_paths(sorted(found))
            self.global_msg.configure(
                text=f"{len(found)} vidéo(s) ajoutée(s) par glisser-déposer.", text_color=T_SUCCES)
        else:
            self.global_msg.configure(
                text="Aucune vidéo reconnue dans les éléments déposés.", text_color=T_ALERTE)

    def _add_paths(self, paths):
        """Ajoute des chemins à la file en évitant les doublons, puis rafraîchit l'affichage."""
        existing = {it.path for it in self.items}
        added = 0
        for p in paths:
            if p not in existing:
                self.items.append(UploadItem(p))
                existing.add(p)          # éviter les doublons dans un même lot
                added += 1
        if added:
            self._refresh_list()
            self._log(f"{added} vidéo(s) ajoutée(s) à la file.")

    def _clear_items(self):
        """Vide la file d'attente et rafraîchit l'affichage."""
        self.items.clear()
        self._refresh_list()

    def _refresh_list(self):
        """Reconstruit le tableau des vidéos en attente (nom, titre éditable, état)."""
        for w in self.list_frame.winfo_children():
            w.destroy()

        if not self.items:
            empty_text = ("Aucune vidéo.\nGlissez-déposez ici des fichiers ou des dossiers,\n"
                          "ou utilisez les boutons ci-dessus."
                          if getattr(self, "dnd_ok", False) else
                          "Aucune vidéo.\nUtilisez « Ajouter des fichiers » ou « Ajouter un dossier ».")
            ctk.CTkLabel(self.list_frame, text=empty_text, text_color=T_SECONDAIRE).pack(pady=40)
            self.count_lbl.configure(text="0 vidéo(s)")
            return

        # En-tête
        hdr = ctk.CTkFrame(self.list_frame, fg_color=S_LIGNE, corner_radius=4)
        hdr.pack(fill="x", pady=(0, 2))
        ctk.CTkLabel(hdr, text="Fichier", width=230, anchor="w",
                     font=ctk.CTkFont(size=11, weight="bold")).pack(side="left", padx=8, pady=4)
        ctk.CTkLabel(hdr, text="Titre (éditable)", anchor="w",
                     font=ctk.CTkFont(size=11, weight="bold")).pack(side="left", padx=8, expand=True, fill="x")
        ctk.CTkLabel(hdr, text="État", width=110,
                     font=ctk.CTkFont(size=11, weight="bold")).pack(side="right", padx=8)

        for i, it in enumerate(self.items):
            row = ctk.CTkFrame(self.list_frame,
                               fg_color=S_LIGNE_ALT if i % 2 == 0 else S_LIGNE,
                               corner_radius=4)
            row.pack(fill="x", pady=1)
            it.row = row

            # nom de fichier (tronqué) + taille
            try:
                size_mb = os.path.getsize(it.path) / (1024 * 1024)
                size_txt = f"{size_mb:.0f} Mo"
            except OSError:
                size_txt = "?"
            fname = it.filename if len(it.filename) <= 28 else it.filename[:25] + "…"
            ctk.CTkLabel(row, text=f"{fname}\n{size_txt}", width=230, anchor="w",
                         justify="left", font=ctk.CTkFont(size=11)).pack(side="left", padx=8, pady=4)

            # titre éditable
            it.title_var = ctk.StringVar(value=it.title)
            it.title_var.trace_add("write",
                                    lambda *_x, item=it: setattr(item, "title", item.title_var.get()))
            ctk.CTkEntry(row, textvariable=it.title_var).pack(
                side="left", padx=8, pady=6, expand=True, fill="x")

            # bouton supprimer
            ctk.CTkButton(row, text="✕", width=28, height=26,
                          fg_color=C_NEUTRE, hover_color=C_DESTRUCTIF,
                          text_color=T_SUR_NEUTRE,
                          command=lambda item=it: self._remove_item(item)).pack(side="right", padx=4)

            # état
            it.status_lbl = ctk.CTkLabel(row, text=it.status, width=100,
                                         text_color=T_DISCRET, font=ctk.CTkFont(size=11))
            it.status_lbl.pack(side="right", padx=6)

        self.count_lbl.configure(text=f"{len(self.items)} vidéo(s)")

    def _remove_item(self, item: UploadItem):
        """Retire une vidéo de la file et rafraîchit l'affichage."""
        if item in self.items:
            self.items.remove(item)
            self._refresh_list()

    def _set_item_status(self, item: UploadItem, status: str, color=None):
        """Met à jour le libellé d'état d'une vidéo dans la liste.

        `color` valait « gray60 » par défaut : une teinte unique, appliquée
        telle quelle aux deux thèmes, à peine lisible sur fond clair. Sans
        couleur explicite, on retombe désormais sur la teinte secondaire de la
        palette, qui s'adapte au mode actif."""
        item.status = status
        if item.status_lbl:
            item.status_lbl.configure(text=status,
                                      text_color=color or T_SECONDAIRE)

    # ── Propriétaires additionnels communs ───────────────────────────────

    def _edit_additional_owners(self):
        """Ouvre OwnerPicker pour choisir les co-propriétaires communs au lot."""
        if not self.api:
            self.global_msg.configure(text="Connectez-vous d'abord (onglet Configuration).",
                                      text_color=T_ALERTE)
            return
        OwnerPicker(self, on_done=self._on_owners_picked,
                    preselected=dict(self.additional_owner_map))

    def _on_owners_picked(self, urls: list[str], labels: list[str]):
        """Callback d'OwnerPicker : mémorise les co-propriétaires choisis et met à jour le libellé."""
        self.additional_owner_urls = urls
        self.additional_owner_map = dict(zip(urls, labels))
        if urls:
            self.add_owners_lbl.configure(text=", ".join(labels)[:60], text_color=T_SUCCES)
        else:
            self.add_owners_lbl.configure(text="aucun", text_color=T_SECONDAIRE)

    # ── Lancement du téléversement ───────────────────────────────────────

    def _upload_pick_owner(self):
        """Ouvre le sélecteur de compte (mono) pour définir le propriétaire du
        lot. L'agent est présélectionné (cas le plus fréquent), mais le choix
        reste explicite et modifiable."""
        if not self.api:
            self.global_msg.configure(text="Connectez-vous d'abord.", text_color=T_ALERTE)
            return
        # Présélection : l'agent déposant (compte du token), s'il est connu
        agent_url = self.config_data.get("agent_owner_url", "")
        agent_lbl = self.config_data.get("agent_owner_label", "")
        pre = {agent_url: agent_lbl} if agent_url else None
        OwnerPicker(self, on_done=lambda *a: None, single=True,
                    on_single=self._upload_set_owner,
                    title="Propriétaire des vidéos du lot", preselected=pre)

    def _upload_set_owner(self, u: dict):
        """Enregistre le propriétaire choisi pour le lot."""
        self.upload_owner_url = u.get("url", "")
        self.upload_owner_label = self._user_label(u)
        self.upload_owner_lbl.configure(text=f"✅  {self.upload_owner_label}",
                                        text_color=T_SUCCES)

    def _start_upload(self):
        """Vérifie les prérequis (connexion, agent, type) puis lance le lot en arrière-plan."""
        if not self.api:
            self.global_msg.configure(text="Non connecté. Voir l'onglet Configuration.",
                                      text_color=T_ERREUR)
            return
        if not self.items:
            self.global_msg.configure(text="Aucune vidéo à téléverser.", text_color=T_ALERTE)
            return
        # Propriétaire : choix EXPLICITE obligatoire (plus de détection auto
        # silencieuse, qui pouvait déposer sur le mauvais compte).
        owner_url = getattr(self, "upload_owner_url", "")
        if not owner_url:
            self.global_msg.configure(
                text="⚠️  Choisissez d'abord le propriétaire des vidéos.",
                text_color=T_ALERTE)
            self._upload_pick_owner()      # ouvre le sélecteur (agent présélectionné)
            return
        type_title = self.type_combo.get()
        type_url = self.type_map.get(type_title, "")
        if not type_url:
            self.global_msg.configure(text="Sélectionnez un type valide.", text_color=T_ALERTE)
            return

        # Au-delà du seuil, la bascule chunkée se fait automatiquement via le
        # compte véhicule embarqué (DEPOT) : rien à configurer. On ne bloque donc
        # que dans le cas improbable où aucun véhicule ne serait disponible.
        if (any(self._file_size(it.path) > cfg.CHUNK_THRESHOLD_BYTES for it in self.items)
                and not (self.vehicle_username and self.vehicle_password)):
            self.global_msg.configure(
                text="⚠️  Compte véhicule indisponible : impossible de téléverser un gros "
                     "fichier. Contactez le support.",
                text_color=T_ALERTE)
            return

        self.launch_btn.configure(state="disabled")
        self._afficher_progression()
        self.batch_progress.set(0)
        # Lecture des widgets ICI (thread principal), puis passage en arguments.
        is_draft = self.visibility_combo.get().startswith("Brouillon")
        do_encode = self.encode_var.get()
        self._run(self._do_batch_upload, owner_url, type_url, is_draft, do_encode)

    @staticmethod
    def _file_size(path: str) -> int:
        """Taille d'un fichier en octets (0 si illisible)."""
        try:
            return os.path.getsize(path)
        except Exception:
            return 0

    @staticmethod
    def _search_term_for(filename: str) -> str:
        """Terme de recherche pour retrouver une vidéo créée par chunké (Pod la
        titre d'après le nom de fichier ASCII envoyé)."""
        base = os.path.splitext(os.path.basename(filename))[0]
        return PodChunkedSession._ascii_filename(base)

    def _verify_chunked_creation(self, search_term: str, pre_ids: set, creator_owner_url: str):
        """(Thread) Après un 504 à la finalisation, Pod termine la création côté
        serveur. On sonde l'API jusqu'à voir une vidéo NOUVELLE (id absent de
        pre_ids) correspondant au fichier, créée par le compte VÉHICULE. Renvoie
        le dict vidéo, ou None après expiration de la fenêtre de vérification."""
        import time as _t
        deadline = _t.time() + cfg.CHUNK_VERIFY_TIMEOUT_S
        while _t.time() < deadline:
            try:
                cands = self.api.search_videos({"search": search_term, "limit": 25})
            except Exception:
                cands = []
            for v in cands:
                if v.get("id") in pre_ids:
                    continue
                own = v.get("owner")
                own_str = own if isinstance(own, str) else (
                    own.get("url", "") if isinstance(own, dict) else "")
                if creator_owner_url and own_str and creator_owner_url.rstrip("/") not in own_str.rstrip("/"):
                    continue
                self._ui(self._log, f"✓ Vidéo apparue après finalisation serveur : {v.get('slug')}")
                return v
            remaining = max(0, int(deadline - _t.time()))
            self._ui(self.global_msg.configure,
                     text=f"⏳ Finalisation côté serveur (gros fichier)… vérification, "
                          f"{remaining//60} min {remaining%60}s restantes",
                     text_color=T_ALERTE)
            _t.sleep(cfg.CHUNK_VERIFY_INTERVAL_S)
        return None

    @staticmethod
    def _est_coupure_reseau(err: Exception) -> bool:
        """Cette erreur vient-elle d'une coupure de connexion (et non d'un refus
        du serveur) ?

        On ne veut replier sur l'envoi par morceaux QUE dans ce cas. Un refus
        métier (400 champ manquant, 403 droits insuffisants…) échouerait de la
        même façon en chunké : le rejouer ne ferait que perdre du temps et
        risquerait de créer un doublon.

        Signature typique de la coupure par la passerelle :
        « SSLEOFError: EOF occurred in violation of protocol ».
        """
        texte = f"{getattr(err, 'body', '')} {err}".lower()
        indices = ("sslerror", "ssleoferror", "eof occurred",
                   "connection aborted", "connection reset",
                   "max retries exceeded", "connectionerror",
                   "remotedisconnected", "broken pipe")
        # `status` vaut 0 quand aucune réponse HTTP n'a été reçue (vraie coupure).
        sans_reponse = getattr(err, "status", 0) in (0, 502, 503, 504)
        return sans_reponse and any(i in texte for i in indices)

    def _replier_sur_chunked(self, it, owner_url: str, type_url: str,
                             is_draft: bool, progress, on_retry):
        """Renvoie le fichier par MORCEAUX après l'échec de l'envoi direct.

        Déroulé : ouverture d'une session avec le compte véhicule → envoi
        découpé → la vidéo naît au nom du véhicule → réattribution au
        propriétaire choisi et pose des métadonnées.

        Renvoie le dictionnaire de la vidéo créée. Toute erreur est propagée à
        l'appelant, qui l'affichera comme un échec normal."""
        chunked = PodChunkedSession(self.config_data.get("url", ""),
                                    self.vehicle_username, self.vehicle_password)
        chunked.login()
        try:
            slug = chunked.upload_video_chunked(
                it.path, chunk_size=cfg.CHUNK_SIZE_BYTES,
                progress_cb=progress, retry_cb=on_retry)
        except PodChunkedError as ce:
            # La passerelle a coupé la finalisation : Pod termine côté serveur.
            # On attend que la vidéo apparaisse plutôt que de conclure à l'échec.
            if ce.status in (502, 503, 504):
                self._ui(self._log,
                         f"⏳ {it.title} : finalisation coupée (HTTP {ce.status}) — "
                         "Pod termine côté serveur, vérification en cours…")
                video = self._verify_chunked_creation(it, owner_url)
                if not video:
                    raise
                slug = video.get("slug", "")
            else:
                raise
        finally:
            chunked.close()

        video = self.api.get_video_by_slug(slug)
        if not video:
            raise PodAPIError(f"Vidéo envoyée (slug={slug}) mais introuvable via l'API.", 0, "")

        # Réattribution au propriétaire choisi + métadonnées. Si elle échoue, la
        # vidéo reste au nom du véhicule : on le signale FORT (jamais en silence).
        patch = {
            "owner": owner_url,
            "title": it.title or it.filename,
            "type": type_url,
            "is_draft": is_draft,
            "main_lang": self.config_data.get("main_lang", "fr"),
            "cursus": self.config_data.get("cursus", "0"),
        }
        if self.additional_owner_urls:
            patch["additional_owners"] = list(self.additional_owner_urls)
        try:
            self.api.patch_video(video, patch)
        except Exception as e:
            it.error = f"réattribution échouée : {e}"
            self._ui(self._set_item_status, it, "⚠️ NON réattribuée", "#ef4444")
            self._ui(self._log,
                     f"⚠️⚠️ {it.title} : vidéo créée (slug={slug}) mais NON réattribuée "
                     f"à {owner_url} — RESTE au nom du véhicule ! Réattribuez-la à la "
                     f"main (onglet Réaffectation). Détail : {e}")
        else:
            self._ui(self._log,
                     f"✅ {it.title} : envoi par morceaux réussi (slug={slug}).")
        return video

    def _do_batch_upload(self, owner_url: str, type_url: str,
                         is_draft: bool, do_encode: bool):
        """(Thread) Téléverse chaque vidéo, ajoute les crédits, lance l'encodage, suit la progression."""
        # `is_draft` et `do_encode` sont reçus en ARGUMENTS : ils ont été lus
        # dans le thread principal par l'appelant. Lire un widget Tk depuis un
        # thread de travail n'est pas fiable (Tcl n'est pas thread-safe) et
        # provoque des plantages aléatoires « main thread is not in main loop ».
        total = len(self.items)
        ok = 0
        chunked = None      # session véhicule, ouverte à la 1re nécessité

        for idx, it in enumerate(self.items, 1):
            # On saute les vidéos DÉJÀ envoyées avec succès : sans cela, ajouter
            # un fichier à une file déjà traitée renverrait tout le lot.
            if it.done:
                ok += 1
                self._ui(self.batch_progress.set, idx / total)
                continue

            self._ui(self._set_item_status, it, "en cours", "#3b82f6")
            self._ui(self.file_progress.set, 0)
            self._ui(self.global_msg.configure,
                     text=f"Téléversement {idx}/{total} : {it.title}", text_color=T_SECONDAIRE)

            def progress(sent, tot, item=it):
                """Callback de progression de l'envoi (met à jour la barre du fichier)."""
                frac = sent / tot if tot else 0
                self._ui(self.file_progress.set, frac)
                self._ui(self.file_progress_lbl.configure,
                         text=f"{item.filename} — {sent/1024/1024:.0f} / {tot/1024/1024:.0f} Mo")

            def on_retry(attempt, total_try, err, item=it):
                """Callback de relance : trace la nouvelle tentative dans le Journal."""
                self._ui(self._log,
                         f"⟳ Nouvelle tentative {attempt}/{total_try} pour {item.title} "
                         f"(coupure réseau)…")
                self._ui(self._set_item_status, item, f"⟳ essai {attempt+1}", "#f59e0b")

            big = self._file_size(it.path) > cfg.CHUNK_THRESHOLD_BYTES
            try:
                if big:
                    # ── Gros fichier : téléversement par MORCEAUX via le VÉHICULE ──
                    if chunked is None:
                        chunked = PodChunkedSession(
                            self.config_data.get("url", ""),
                            self.vehicle_username, self.vehicle_password)
                        chunked.login()
                        self._ui(self._log, "Session véhicule ouverte (upload chunké).")
                    self._ui(self._log,
                             f"Gros fichier (> {cfg.CHUNK_THRESHOLD_BYTES//1024//1024} Mo) : "
                             f"bascule chunkée pour {it.title}.")
                    # Repères pour la récupération après un éventuel 504.
                    search_term = self._search_term_for(it.filename)
                    try:
                        pre_ids = {v.get("id") for v in
                                   self.api.search_videos({"search": search_term, "limit": 25})}
                    except Exception:
                        pre_ids = set()
                    # 1) Envoi par morceaux → vidéo créée au nom du VÉHICULE.
                    video = None
                    try:
                        slug = chunked.upload_video_chunked(
                            it.path, chunk_size=cfg.CHUNK_SIZE_BYTES,
                            progress_cb=progress, retry_cb=on_retry)
                    except PodChunkedError as ce:
                        if ce.status in (502, 503, 504):
                            self._ui(self._log,
                                     f"⏳ Finalisation coupée par la passerelle (HTTP {ce.status}) "
                                     "— Pod termine côté serveur, vérification en cours…")
                            self._ui(self._set_item_status, it, "⏳ finalisation serveur", "#f59e0b")
                            video = self._verify_chunked_creation(
                                search_term, pre_ids, self.vehicle_owner_url)
                            if not video:
                                raise
                            slug = video.get("slug", "")
                        else:
                            raise
                    it.slug = slug
                    if video is None:
                        video = self.api.get_video_by_slug(slug)
                    it.video_url = video.get("url", "") if isinstance(video, dict) else ""
                    # 2) RÉATTRIBUTION au propriétaire choisi + métadonnées (par token).
                    #    Point critique : si le PATCH owner échoue, la vidéo reste au
                    #    nom du véhicule → on le signale FORT (jamais en silence).
                    if video:
                        patch = {
                            "owner": owner_url,                      # ← propriétaire CHOISI
                            "title": it.title or it.filename,
                            "type": type_url,
                            "is_draft": is_draft,
                            "main_lang": self.config_data.get("main_lang", "fr"),
                            "cursus": self.config_data.get("cursus", "0"),
                        }
                        if self.additional_owner_urls:
                            patch["additional_owners"] = list(self.additional_owner_urls)
                        try:
                            self.api.patch_video(video, patch)
                        except Exception as e:
                            # Échec de réattribution : la vidéo existe mais reste au
                            # nom du véhicule. Erreur BRUYANTE (pas de faux succès).
                            it.error = f"réattribution échouée : {e}"
                            self._ui(self._set_item_status, it,
                                     "⚠️ NON réattribuée", "#ef4444")
                            self._ui(self._log,
                                     f"⚠️⚠️ {it.title} : vidéo créée (slug={slug}) mais NON "
                                     f"réattribuée à {owner_url} — RESTE au nom du véhicule ! "
                                     f"Réattribuez-la à la main (onglet Réaffectation). Détail : {e}")
                            self._ui(self.batch_progress.set, idx / total)
                            continue      # on n'enchaîne pas encodage/crédits sur un état douteux
                    else:
                        self._ui(self._log,
                                 f"⚠️ Vidéo créée (slug={slug}) mais introuvable via l'API pour "
                                 "réattribution/métadonnées — à vérifier côté web.")
                else:
                    # ── Fichier sous le seuil : upload classique par TOKEN ──
                    try:
                        video = self.api.upload_video(
                            it.path, it.title or it.filename, owner_url, type_url,
                            main_lang=self.config_data.get("main_lang", "fr"),
                            cursus=self.config_data.get("cursus", "0"),
                            is_draft=is_draft,
                            additional_owner_urls=self.additional_owner_urls,
                            site_urls=self.site_urls,
                            progress_cb=progress,
                            retry_cb=on_retry,
                        )
                    except PodAPIError as e:
                        # REPLI AUTOMATIQUE SUR L'ENVOI PAR MORCEAUX.
                        #
                        # L'envoi monobloc peut être coupé par la passerelle même
                        # sous le seuil : au-delà d'environ une minute de transfert,
                        # nginx ferme la connexion (erreur SSL « EOF occurred in
                        # violation of protocol »). Le seuil en octets ne suffit
                        # donc pas : ce qui compte est la DURÉE de l'envoi, qui
                        # dépend du débit montant.
                        #
                        # Réessayer à l'identique échoue invariablement (constaté :
                        # 3 tentatives, puis 3 autres après relance manuelle). On
                        # bascule donc sur la voie chunkée, conçue pour résister
                        # à ces coupures, plutôt que d'abandonner.
                        if not self._est_coupure_reseau(e):
                            raise
                        self._ui(self._log,
                                 f"⚠️ {it.title} : envoi direct coupé par le serveur. "
                                 "Bascule automatique sur l'envoi par morceaux…")
                        self._ui(self._set_item_status, it, "⟳ envoi par morceaux", "#f59e0b")
                        video = self._replier_sur_chunked(
                            it, owner_url, type_url, is_draft, progress, on_retry)
                    it.slug = video.get("slug", "") if isinstance(video, dict) else ""
                    it.video_url = video.get("url", "") if isinstance(video, dict) else ""

                # Encodage
                if do_encode and it.slug:
                    try:
                        self.api.launch_encoding(it.slug)
                    except Exception as e:
                        self._ui(self._log, f"Encodage non lancé ({it.title}) : {e}")

                ok += 1
                it.done = True            # marque le succès : ne sera pas relancé
                self._ui(self._set_item_status, it, "✅ terminé", "#22c55e")
                self._ui(self._log,
                         f"Téléversé{' (chunké)' if big else ''} : {it.title}  (slug={it.slug})")

            except PodChunkedError as e:
                it.error = f"{e} — {e.body}"
                self._ui(self._set_item_status, it, "❌ échec", "#ef4444")
                self._ui(self._log, f"ÉCHEC chunké {it.title} : {e} | {e.body[:200]}")
            except PodAPIError as e:
                it.error = f"{e} — {e.body}"
                self._ui(self._set_item_status, it, "❌ échec", "#ef4444")
                self._ui(self._log, f"ÉCHEC {it.title} : {e} | {e.body[:200]}")
            except Exception as e:
                it.error = str(e)
                self._ui(self._set_item_status, it, "❌ échec", "#ef4444")
                self._ui(self._log, f"ÉCHEC {it.title} : {e}")

            self._ui(self.batch_progress.set, idx / total)

        # Fermeture propre de la session véhicule si elle a été ouverte.
        if chunked is not None:
            chunked.close()

        self._ui(self._on_batch_done, ok, total)

    def _on_batch_done(self, ok: int, total: int):
        """Réactive l'interface et affiche le bilan une fois le lot terminé."""
        self.launch_btn.configure(state="normal")
        self.file_progress.set(0)
        self.file_progress_lbl.configure(text="")
        self._masquer_progression()
        color = "#22c55e" if ok == total else "#f59e0b"
        self.global_msg.configure(text=f"Terminé : {ok}/{total} vidéo(s) téléversée(s).", text_color=color)
        self._log(f"Lot terminé : {ok}/{total} réussis.")
        # Afficher le bouton « Relancer les échecs » s'il reste des échecs
        self._update_retry_button()

        # Les vidéos qui viennent d'être déposées n'existent PAS dans le magasin
        # partagé : il a été rempli avant l'envoi. Sans relecture, les onglets
        # Vidéos, Encodage et Inventaire continueraient d'ignorer ces nouvelles
        # vidéos jusqu'à un rafraîchissement manuel.
        #
        # C'est le seul cas où il faut vraiment relire le serveur : une création
        # ne peut pas se déduire de ce qu'on a en mémoire, contrairement à une
        # modification ou une suppression.
        if ok:
            self._ui(self._recharger_apres_depot)

    def _recharger_apres_depot(self):
        """Relit l'instance après un dépôt réussi, puis rafraîchit les onglets.

        Une CRÉATION est le seul cas qui impose une relecture serveur : les
        modifications et suppressions se répercutent en mémoire, mais une vidéo
        nouvelle est par définition absente du magasin."""
        self._log("🔄 Mise à jour de la liste après le dépôt…")
        self.ensure_videos(force=True, on_ready=self._refresh_video_views)

    def _update_retry_button(self):
        """Affiche le bouton de relance uniquement s'il y a des vidéos en échec."""
        # On s'appuie sur l'indicateur `done` plutôt que sur le libellé affiché :
        # une vidéo non réussie et non en cours est à relancer, quel que soit le
        # texte de son statut.
        n_fail = sum(1 for it in self.items
                     if not it.done and str(it.status).endswith("échec"))
        if n_fail:
            self.retry_btn.configure(text=f"🔄  Relancer les échecs ({n_fail})")
            if not self.retry_btn.winfo_ismapped():
                self.retry_btn.pack(side="left", padx=8)
        else:
            if self.retry_btn.winfo_ismapped():
                self.retry_btn.pack_forget()

    def _retry_failed(self):
        """Relance le téléversement des seules vidéos en échec.
        Le lot saute automatiquement les vidéos déjà « terminé », donc on peut
        réutiliser le même parcours : on remet les échecs « en attente » et on
        relance. Le propriétaire et le type déjà choisis sont réutilisés."""
        if not self.api:
            self.global_msg.configure(text="Non connecté.", text_color=T_ERREUR)
            return
        owner_url = getattr(self, "upload_owner_url", "")
        if not owner_url:
            self.global_msg.configure(
                text="⚠️  Choisissez d'abord le propriétaire des vidéos.", text_color=T_ALERTE)
            self._upload_pick_owner()
            return
        type_url = self.type_map.get(self.type_combo.get(), "")
        if not type_url:
            self.global_msg.configure(text="Sélectionnez un type valide.", text_color=T_ALERTE)
            return
        # Réinitialiser le statut des échecs pour qu'ils soient re-tentés
        failed = [it for it in self.items
                  if not it.done and str(it.status).endswith("échec")]
        if not failed:
            self.global_msg.configure(text="Aucune vidéo en échec.", text_color=T_SECONDAIRE)
            return
        for it in failed:
            self._set_item_status(it, "en attente")
        self._log(f"Relance de {len(failed)} vidéo(s) en échec…")
        self.launch_btn.configure(state="disabled")
        self.retry_btn.pack_forget()
        # La relance n'emprunte PAS `_start_upload` : sans cet appel, les barres
        # resteraient masquées pendant tout le renvoi.
        self._afficher_progression()
        # Lecture des widgets ICI (thread principal), puis passage en arguments.
        is_draft = self.visibility_combo.get().startswith("Brouillon")
        do_encode = self.encode_var.get()
        self._run(self._do_batch_upload, owner_url, type_url, is_draft, do_encode)

    # ═════════════════════════════════════════════════════════════════════
    #  (L'onglet Co-auteurs a été SUPPRIMÉ : aucun contributeur n'était utilisé
    #   sur l'instance. `pod_api.add_contributor()` reste disponible si l'usage
    #   revenait un jour.)
    # ═════════════════════════════════════════════════════════════════════

    # ═════════════════════════════════════════════════════════════════════
    #  ONGLET GROUPES D'ACCÈS
    # ═════════════════════════════════════════════════════════════════════

    @staticmethod
    def _is_manual_group(g: dict) -> bool:
        """Un groupe est « manuel » (modifiable) s'il commence par 'grp_'.
        Les autres (student, staff, employee…) sont synchronisés par l'annuaire
        SSO et ne doivent pas être modifiés via l'API (risque d'écrasement)."""
        return str(g.get("code_name", "")).lower().startswith("grp_")

    def _build_tab_groups(self):
        """Onglet de gestion des groupes d'accès : lister, créer, gérer les
        membres et supprimer. Seuls les groupes manuels (préfixe 'grp_') sont
        modifiables ; les groupes SSO sont affichés en lecture seule."""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tabs["groups"] = frame

        ctk.CTkLabel(frame, text="🔐  Groupes d'accès",
                     font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(
            frame,
            text="Gérez les groupes qui peuvent accéder aux vidéos restreintes. "
                 "Seuls les groupes manuels (préfixe « grp_ ») sont modifiables ici ; "
                 "les groupes synchronisés par l'annuaire (student, staff…) sont en "
                 "lecture seule pour éviter qu'ils soient écrasés à la synchro.",
            text_color=T_SECONDAIRE, font=ctk.CTkFont(size=12),
            justify="left", wraplength=860).pack(anchor="w", pady=(0, 10))

        # — Barre d'actions —
        bar = ctk.CTkFrame(frame, fg_color="transparent")
        bar.pack(fill="x")
        ctk.CTkButton(bar, text="🔄  Recharger", width=130, fg_color=C_ACTION,
                      hover_color=C_ACTION_SURV, command=self._groups_reload).pack(side="left")
        ctk.CTkButton(bar, text="➕  Nouveau groupe manuel", width=200, fg_color=C_NEUTRE,
                      command=self._groups_create_dialog, text_color=T_SUR_NEUTRE).pack(side="left", padx=8)
        self.groups_status = ctk.CTkLabel(bar, text="", text_color=T_SECONDAIRE,
                                          font=ctk.CTkFont(size=11))
        self.groups_status.pack(side="left", padx=10)

        # — Liste des groupes —
        self.groups_list = ctk.CTkScrollableFrame(frame, label_text="Groupes d'accès", fg_color=S_CARTE, label_anchor="w",
                                                  label_font=ctk.CTkFont(size=12, weight="bold"))
        self.groups_list.pack(fill="both", expand=True, pady=(8, 0))

        self.groups_owners_map = {}   # url compte → url owner (chargé à la demande)

    def _show_groups_tab_hook(self):
        """Prépare l'onglet Groupes d'accès à chaque ouverture.

        ATTENTION au piège corrigé ici : la connexion remplit déjà
        `self.access_groups`. L'ancienne version ne déclenchait le chargement
        QUE si cette liste était vide — donc, une fois connecté, elle ne faisait
        rien… et l'affichage n'était jamais construit, puisque le rendu n'avait
        lieu qu'à l'intérieur du chargement. Résultat : un onglet vide alors que
        les groupes étaient bien en mémoire.

        On distingue donc deux cas :
          • données pas encore chargées → chargement complet (groupes + table
            des comptes propriétaires, nécessaire pour ajouter des membres) ;
          • données déjà en mémoire → simple affichage, sans appel réseau.
        """
        if not self.api:
            return
        if not getattr(self, "_groups_loaded", False):
            self._groups_reload()
        else:
            # Déjà chargé : on se contente de (re)construire l'affichage.
            self._render_groups_list()
            self.groups_status.configure(
                text=f"{len(self.access_groups)} groupe(s).", text_color=T_SECONDAIRE)

    def _groups_reload(self):
        """Recharge la liste des groupes d'accès (avec leurs URLs) en arrière-plan."""
        if not self.api:
            self.groups_status.configure(text="Connectez-vous d'abord.", text_color=T_ALERTE)
            return
        self.groups_status.configure(text="⏳  Chargement…", text_color=T_SECONDAIRE)
        self._run(self._do_groups_reload)

    def _do_groups_reload(self):
        """(Thread) Recharge groupes + table owners, puis réaffiche la liste."""
        try:
            self.access_groups = self.api.get_access_groups()
            # Table compte→owner (pour convertir lors de l'ajout de membres)
            self.groups_owners_map = self.api.get_owners_map()
            self._groups_loaded = True     # évite un rechargement à chaque ouverture
            self._ui(self.groups_status.configure,
                     text=f"{len(self.access_groups)} groupe(s).", text_color=T_SECONDAIRE)
            self._ui(self._render_groups_list)
        except Exception as e:
            self._ui(self.groups_status.configure, text=f"❌  {e}", text_color=T_ERREUR)

    def _render_groups_list(self):
        """Affiche une ligne par groupe (manuel = modifiable, SSO = verrouillé)."""
        for w in self.groups_list.winfo_children():
            w.destroy()
        if not self.access_groups:
            ctk.CTkLabel(self.groups_list, text="Aucun groupe. Cliquez sur « Recharger ».",
                         text_color=T_SECONDAIRE).pack(anchor="w", padx=8, pady=8)
            return
        for g in self.access_groups:
            manual = self._is_manual_group(g)
            n_members = len(g.get("users") or [])
            row = ctk.CTkFrame(self.groups_list,
                               fg_color=S_LIGNE_ALT if manual else S_LIGNE,
                               corner_radius=6)
            row.pack(fill="x", pady=2)
            tag = "✏️ manuel" if manual else "🔒 annuaire"
            ctk.CTkLabel(row, text=f"{g.get('code_name')}   ({n_members} membre·s)  ·  {tag}",
                         anchor="w", font=ctk.CTkFont(size=12)).pack(
                side="left", padx=10, pady=8, fill="x", expand=True)
            if manual:
                # `side="right"` empile de DROITE à GAUCHE : le premier posé se
                # retrouve le plus à droite. On pose donc la suppression en
                # premier pour qu'elle soit en bout de rangée, loin de l'action
                # courante — auparavant elle précédait « Membres » à la lecture.
                btn_g = ctk.CTkButton(row, text="🗑", width=34,
                                      fg_color=C_NEUTRE, hover_color=C_DESTRUCTIF,
                                      text_color=T_SUR_NEUTRE,
                                      font=ctk.CTkFont(size=T_CORPS),
                                      command=lambda gg=g: self._groups_delete(gg))
                btn_g.pack(side="right", padx=4)
                ajouter_infobulle(btn_g, "Supprimer ce groupe")
                ctk.CTkFrame(row, width=1, height=22,
                             fg_color=S_PUCE).pack(side="right", padx=6)
                ctk.CTkButton(row, text="👥 Membres", width=90, fg_color=C_NEUTRE,
                              command=lambda gg=g: self._groups_manage_members(gg), text_color=T_SUR_NEUTRE).pack(side="right", padx=4)
            else:
                ctk.CTkLabel(row, text="lecture seule", text_color=T_DISCRET,
                             font=ctk.CTkFont(size=11)).pack(side="right", padx=12)

    # — Création d'un groupe manuel —
    def _groups_create_dialog(self):
        """Ouvre la fenêtre de création d'un groupe d'accès."""
        if not self.api:
            self.groups_status.configure(text="Connectez-vous d'abord.", text_color=T_ALERTE)
            return
        if not self.site_urls:
            self.groups_status.configure(
                text="Aucun site chargé (requis pour créer un groupe).", text_color=T_ALERTE)
            return
        win = ctk.CTkToplevel(self)
        win.title("Nouveau groupe manuel")
        win.geometry("420x230")
        _focus_toplevel(win, self)
        ctk.CTkLabel(win, text="Créer un groupe d'accès manuel",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(padx=16, pady=(16, 6), anchor="w")
        ctk.CTkLabel(win, text="Nom du groupe (le préfixe « grp_ » est ajouté\n"
                               "automatiquement s'il manque) :",
                     font=ctk.CTkFont(size=11), text_color=T_SECONDAIRE,
                     justify="left").pack(padx=16, anchor="w")
        name_entry = ctk.CTkEntry(win, width=360, placeholder_text="ex. eformation  →  grp_eformation")
        name_entry.pack(padx=16, pady=8)
        msg = ctk.CTkLabel(win, text="", font=ctk.CTkFont(size=11), text_color=T_ERREUR)
        msg.pack(padx=16, anchor="w")

        def _do_create():
            """Valide le formulaire et crée l'élément."""
            raw = name_entry.get().strip()
            if not raw:
                msg.configure(text="Indiquez un nom."); return
            code = raw if raw.lower().startswith("grp_") else f"grp_{raw}"
            # Refuser les doublons
            if any(g.get("code_name") == code for g in self.access_groups):
                msg.configure(text=f"« {code} » existe déjà."); return
            win.destroy()
            self.groups_status.configure(text=f"⏳  Création de {code}…", text_color=T_SECONDAIRE)
            self._run(self._do_groups_create, code)

        ctk.CTkButton(win, text="Créer", fg_color=C_SUCCES, hover_color=C_SUCCES_SURV,
                      command=_do_create).pack(padx=16, pady=10, anchor="w")

    def _do_groups_create(self, code):
        """(Thread) Crée le groupe puis recharge la liste."""
        try:
            self.api.create_access_group(code, self.site_urls, display_name=code)
            self._ui(self.groups_status.configure,
                     text=f"✅  Groupe « {code} » créé.", text_color=T_SUCCES)
            self._ui(self._log, f"Groupe d'accès créé : {code}")
            self._do_groups_reload()
        except Exception as e:
            self._ui(self.groups_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Création groupe {code} : {e}")

    # — Suppression d'un groupe manuel —
    def _groups_delete(self, g):
        """Supprime un groupe d'accès (après confirmation)."""
        code = g.get("code_name")
        if not messagebox.askyesno(
                "Supprimer le groupe",
                f"Supprimer définitivement le groupe « {code} » ?\n\n"
                "Les vidéos restreintes à ce groupe perdront cette restriction."):
            return
        self._run(self._do_groups_delete, g)

    def _do_groups_delete(self, g):
        """(Thread) DELETE du groupe puis rechargement."""
        try:
            self.api.delete_access_group(g.get("url"))
            self._ui(self.groups_status.configure,
                     text=f"✅  Groupe « {g.get('code_name')} » supprimé.", text_color=T_SUCCES)
            self._ui(self._log, f"Groupe d'accès supprimé : {g.get('code_name')}")
            self._do_groups_reload()
        except Exception as e:
            self._ui(self.groups_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Suppression groupe : {e}")

    # — Gestion des membres d'un groupe manuel —
    def _groups_manage_members(self, g):
        """Ouvre un sélecteur de comptes pré-coché sur les membres actuels.
        À la validation, convertit les comptes choisis en URLs /owners/ et
        remplace la liste des membres du groupe."""
        if not self.all_users:
            self.groups_status.configure(text="⏳  Chargement des comptes…", text_color=T_SECONDAIRE)
            self._run(lambda: (self._reload_users_for_admin(),
                               self._ui(lambda: self._groups_open_member_picker(g))))
            return
        self._groups_open_member_picker(g)

    def _groups_open_member_picker(self, g):
        # Pré-sélection : convertir les membres actuels (URLs /owners/) en URLs
        # de comptes /users/ pour les retrouver dans le sélecteur.
        # IMPORTANT : OwnerPicker compare les URLs *brutes* (avec slash final),
        # donc les clés de `pre` doivent être l'URL brute du compte, pas sa
        # version normalisée — sinon rien n'apparaît coché.
        owner_to_norm_user = {str(ov).rstrip("/"): str(uu).rstrip("/")
                              for uu, ov in self.groups_owners_map.items()}
        # URL brute + libellé indexés par URL de compte normalisée
        raw_by_norm = {str(u.get("url", "")).rstrip("/"): u.get("url", "")
                       for u in (self.all_users or [])}
        label_by_norm = {str(u.get("url", "")).rstrip("/"): self._user_label(u)
                         for u in (self.all_users or [])}
        cur_owner_urls = [str(x.get("url") if isinstance(x, dict) else x).rstrip("/")
                          for x in (g.get("users") or [])]
        pre = {}
        for ourl in cur_owner_urls:
            unorm = owner_to_norm_user.get(ourl)
            if not unorm:
                continue
            raw = raw_by_norm.get(unorm)
            if raw:                                   # clé = URL brute (avec slash)
                pre[raw] = label_by_norm.get(unorm, raw)
        OwnerPicker(self,
                    on_done=lambda urls, labels: self._groups_apply_members(g, urls),
                    title=f"Membres de « {g.get('code_name')} »",
                    preselected=pre)

    def _groups_apply_members(self, g, user_urls):
        """Convertit les comptes choisis en URLs /owners/ puis applique."""
        owner_urls = []
        missing = []
        for uurl in user_urls:
            ourl = self.groups_owners_map.get(str(uurl).rstrip("/"))
            if ourl:
                owner_urls.append(ourl)
            else:
                missing.append(uurl)
        if missing:
            self._log(f"⚠️ {len(missing)} compte(s) sans owner correspondant (ignoré·s).")
        self._run(self._do_groups_apply_members, g, owner_urls)

    def _do_groups_apply_members(self, g, owner_urls):
        """(Thread) PATCH des membres du groupe puis rechargement."""
        try:
            self.api.set_access_group_members(g.get("url"), owner_urls)
            self._ui(self.groups_status.configure,
                     text=f"✅  « {g.get('code_name')} » : {len(owner_urls)} membre·s.",
                     text_color=T_SUCCES)
            self._ui(self._log,
                     f"Groupe « {g.get('code_name')} » : {len(owner_urls)} membre·s définis.")
            self._do_groups_reload()
        except Exception as e:
            self._ui(self.groups_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Membres groupe « {g.get('code_name')} » : {e}")

    # ═════════════════════════════════════════════════════════════════════
    #  ONGLET CONFIGURATION
    # ═════════════════════════════════════════════════════════════════════

    def _build_tab_config(self):
        """Construit l'onglet Configuration (connexion API + réglages de l'instance).

        Cet onglet est DÉFILABLE, contrairement aux autres : son contenu est long
        (connexion, compte véhicule, agent déposant, aide au jeton, réglages de
        l'instance) et, en fenêtre réduite, les dernières sections étaient
        purement et simplement invisibles — sans même une barre de défilement
        pour le laisser deviner."""
        frame = ctk.CTkScrollableFrame(self.content, fg_color="transparent")
        self.tabs["config"] = frame

        ctk.CTkLabel(frame, text="⚙️  Configuration",
                     font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=(0, 10))

        # — Connexion API —
        api_box = ctk.CTkFrame(frame, fg_color=S_CARTE)
        api_box.pack(fill="x")
        ctk.CTkLabel(api_box, text="Connexion à l'instance Pod (compte superutilisateur)",
                     font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=3,
                                                           padx=12, pady=(12, 4), sticky="w")

        ctk.CTkLabel(api_box, text="URL :", width=110, anchor="e").grid(row=1, column=0, padx=8, pady=8)
        self.url_entry = ctk.CTkEntry(api_box, width=430)
        self.url_entry.insert(0, self.config_data.get("url", ""))
        self.url_entry.grid(row=1, column=1, padx=8, pady=8, sticky="ew")

        ctk.CTkLabel(api_box, text="Token :", width=110, anchor="e").grid(row=2, column=0, padx=8, pady=8)
        self.token_entry = ctk.CTkEntry(api_box, width=430, show="*")
        if self.token:
            self.token_entry.insert(0, self.token)
        self.token_entry.grid(row=2, column=1, padx=8, pady=8, sticky="ew")

        self.show_token = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(api_box, text="Afficher", variable=self.show_token,
                        command=lambda: self.token_entry.configure(
                            show="" if self.show_token.get() else "*")).grid(row=2, column=2, padx=4)

        # Où le token est-il rangé ? Normalement dans le coffre-fort de l'OS
        # (chiffré). En cas d'indisponibilité, l'application bascule sur un
        # fichier EN CLAIR : ce libellé prévient alors l'utilisateur, car le
        # token porte des droits d'administration sur toute l'instance.
        self.token_storage_lbl = ctk.CTkLabel(
            api_box, text="", font=ctk.CTkFont(size=11),
            text_color=T_ALERTE, anchor="w", wraplength=620, justify="left")
        self.token_storage_lbl.grid(row=3, column=1, columnspan=2, sticky="w", padx=8)

        # — Compte VÉHICULE (local) pour le chunké des gros fichiers —
        ctk.CTkLabel(api_box,
                     text=f"Compte véhicule (local) — utilisé pour les gros fichiers (> "
                          f"{cfg.CHUNK_THRESHOLD_BYTES // 1024 // 1024} Mo), envoyés par morceaux "
                          "puis réattribués au propriétaire choisi.\n"
                          "FACULTATIF : un compte intégré est déjà utilisé par défaut. "
                          "Ne remplissez ces champs que pour employer un autre compte.",
                     text_color=T_SECONDAIRE, font=ctk.CTkFont(size=11)).grid(
            row=3, column=0, columnspan=3, padx=12, pady=(6, 0), sticky="w")

        ctk.CTkLabel(api_box, text="Identifiant :", width=110, anchor="e").grid(row=4, column=0, padx=8, pady=8)
        self.user_entry = ctk.CTkEntry(api_box, width=430,
                                       placeholder_text="identifiant local du compte véhicule (optionnel)")
        if self.vehicle_username:
            self.user_entry.insert(0, self.vehicle_username)
        self.user_entry.grid(row=4, column=1, padx=8, pady=8, sticky="ew")

        ctk.CTkLabel(api_box, text="Mot de passe :", width=110, anchor="e").grid(row=5, column=0, padx=8, pady=8)
        self.pass_entry = ctk.CTkEntry(api_box, width=430, show="*",
                                       placeholder_text="mot de passe (collage Ctrl+V possible)")
        if self.vehicle_password:
            self.pass_entry.insert(0, self.vehicle_password)
        self.pass_entry.grid(row=5, column=1, padx=8, pady=8, sticky="ew")
        self.show_pass = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(api_box, text="Afficher", variable=self.show_pass,
                        command=lambda: self.pass_entry.configure(
                            show="" if self.show_pass.get() else "*")).grid(row=5, column=2, padx=4)

        btn_row = ctk.CTkFrame(api_box, fg_color="transparent")
        btn_row.grid(row=6, column=1, columnspan=2, padx=8, pady=10, sticky="w")
        ctk.CTkButton(btn_row, text="🔌  Tester & se connecter", fg_color=C_SUCCES,
                      hover_color=C_SUCCES_SURV, command=self._connect).pack(side="left")
        ctk.CTkButton(btn_row, text="🚪  Oublier le token / Se déconnecter", width=260,
                      fg_color=C_NEUTRE, hover_color="#7f1d1d",
                      command=self._forget_token, text_color=T_SUR_NEUTRE).pack(side="left", padx=10)
        api_box.columnconfigure(1, weight=1)

        self.config_msg = ctk.CTkLabel(frame, text="", font=ctk.CTkFont(size=12))
        self.config_msg.pack(anchor="w", pady=4)

        ctk.CTkFrame(frame, height=1, fg_color=S_FILET).pack(fill="x", pady=8)

        # — Agent déposant —
        agent_box = ctk.CTkFrame(frame, fg_color=S_CARTE)
        agent_box.pack(fill="x")
        ctk.CTkLabel(agent_box, text="Agent déposant (propriétaire des vidéos)",
                     font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=3,
                                                           padx=12, pady=(12, 2), sticky="w")
        ctk.CTkLabel(agent_box, text="Les vidéos déposées appartiendront à ce compte Pod.",
                     text_color=T_SECONDAIRE, font=ctk.CTkFont(size=11)).grid(
            row=1, column=0, columnspan=3, padx=12, pady=(0, 6), sticky="w")

        self.agent_filter = ctk.CTkEntry(agent_box, width=300,
                                         placeholder_text="🔍 nom / identifiant…")
        self.agent_filter.grid(row=2, column=0, columnspan=2, padx=8, pady=8, sticky="ew")
        self.agent_filter.bind("<KeyRelease>",
                               lambda e: self._debounce("users", self._render_users))
        ctk.CTkButton(agent_box, text="🔄  Recharger", width=130,
                      fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=lambda: self._run(self._load_all_users), text_color=T_SUR_NEUTRE).grid(row=2, column=2, padx=8, pady=8)

        self.users_count_lbl = ctk.CTkLabel(agent_box, text="", text_color=T_SECONDAIRE,
                                            font=ctk.CTkFont(size=11))
        self.users_count_lbl.grid(row=3, column=0, columnspan=3, padx=12, sticky="w")

        self.agent_results = ctk.CTkScrollableFrame(agent_box, height=220, fg_color=S_CARTE)
        self.agent_results.grid(row=4, column=0, columnspan=3, padx=12, pady=(0, 10), sticky="ew")
        agent_box.columnconfigure(1, weight=1)

        # — Aide token —
        help_box = ctk.CTkFrame(frame, fg_color=S_CARTE, corner_radius=8)
        help_box.pack(fill="x", pady=8)
        ctk.CTkLabel(help_box, text="ℹ️  Créer le token de service",
                     font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=14, pady=(10, 2))
        ctk.CTkLabel(
            help_box,
            text="Connecté en administrateur, ouvrez  <URL>/admin/authtoken/  → "
                 "« Add token » → choisissez le compte de service.\n"
                 "⚠️ Le token hérite des droits de ce compte. Il est stocké chiffré "
                 "dans le coffre-fort de votre système (Keychain / Credential Manager).",
            justify="left", text_color=T_SECONDAIRE, wraplength=820).pack(anchor="w", padx=14, pady=(0, 12))

        # ── Réglages de l'instance (hors API) ────────────────────────────
        # Certains paramètres de Pod ne sont PAS exposés par l'API REST : la
        # page d'accueil, par exemple, relève de la configuration du serveur.
        # L'application ne peut donc pas les modifier — mais elle peut éviter
        # de chercher l'adresse, en ouvrant directement la bonne page de
        # l'administration dans le navigateur, où l'utilisateur est déjà
        # authentifié. Même principe que le bouton « 🔑 Token » de l'onglet
        # Comptes.
        ctk.CTkFrame(frame, height=1, fg_color=S_FILET).pack(fill="x", pady=8)

        inst_box = ctk.CTkFrame(frame, fg_color=S_CARTE)
        inst_box.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(inst_box, text="🛠  Réglages de l'instance",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(
            anchor="w", padx=14, pady=(10, 2))
        ctk.CTkLabel(
            inst_box,
            text="Ces réglages ne sont pas accessibles par l'API : ils se modifient "
                 "dans l'administration de Pod. Les boutons ci-dessous l'ouvrent "
                 "directement à la bonne page, dans votre navigateur.\n"
                 "La page d'accueil se règle à deux endroits : le TEXTE de "
                 "présentation (page statique) et les VIGNETTES affichées (blocs).",
            justify="left", text_color=T_SECONDAIRE, font=ctk.CTkFont(size=11),
            wraplength=820).pack(anchor="w", padx=14, pady=(0, 8))

        rangee = ctk.CTkFrame(inst_box, fg_color="transparent")
        rangee.pack(fill="x", padx=14, pady=(0, 12))
        # La page d'accueil se règle en DEUX endroits distincts de
        # l'administration (adresses vérifiées sur l'instance) :
        #   • le TEXTE de présentation est une « page statique » (flatpage) ;
        #   • les VIGNETTES et encarts affichés sont des « blocs ».
        # D'où deux boutons plutôt qu'un seul, pour éviter de chercher.
        ctk.CTkButton(rangee, text="🏠  Accueil : texte", width=170,
                      fg_color=C_ACCENT, hover_color=C_ACCENT_SURV,
                      command=lambda: self._ouvrir_admin(
                          "/admin/flatpages/flatpage/",
                          "texte de la page d'accueil")).pack(side="left")
        ctk.CTkButton(rangee, text="🧩  Accueil : blocs", width=170,
                      fg_color=C_ACCENT, hover_color=C_ACCENT_SURV,
                      command=lambda: self._ouvrir_admin(
                          "/admin/main/block/",
                          "blocs de la page d'accueil")).pack(side="left", padx=8)
        ctk.CTkButton(rangee, text="🔑  Jetons", width=120, fg_color=C_NEUTRE,
                      hover_color=C_NEUTRE_SURV,
                      command=lambda: self._ouvrir_admin(
                          "/admin/authtoken/tokenproxy/", "jetons"), text_color=T_SUR_NEUTRE).pack(side="left")
        ctk.CTkButton(rangee, text="⚙️  Administration", width=160, fg_color=C_NEUTRE,
                      hover_color=C_NEUTRE_SURV,
                      command=lambda: self._ouvrir_admin(
                          "/admin/", "administration"), text_color=T_SUR_NEUTRE).pack(side="left", padx=8)

        self.config_admin_msg = ctk.CTkLabel(
            inst_box, text="", font=ctk.CTkFont(size=11), text_color=T_SECONDAIRE,
            wraplength=820, justify="left", anchor="w")
        self.config_admin_msg.pack(fill="x", padx=14, pady=(0, 10))

    def _ouvrir_admin(self, chemin: str, libelle: str):
        """Ouvre une page de l'administration Pod dans le navigateur.

        L'application ne manipule aucun identifiant privilégié : c'est la
        session déjà ouverte dans le navigateur qui authentifie l'utilisateur.
        """
        base = (self.config_data.get("url") or "").rstrip("/")
        if not base:
            self.config_admin_msg.configure(
                text="Renseignez d'abord l'adresse de l'instance, plus haut.",
                text_color=T_ALERTE)
            return
        url = f"{base}{chemin}"
        try:
            import webbrowser
            webbrowser.open(url)
            self._log(f"🛠 Administration ouverte ({libelle}) : {url}")
            self.config_admin_msg.configure(
                text=f"Page « {libelle} » ouverte dans votre navigateur. "
                     "Connectez-vous en administrateur si ce n'est pas déjà fait.",
                text_color=T_SUCCES)
        except Exception as e:
            self.config_admin_msg.configure(text=f"❌ Ouverture impossible : {e}",
                                            text_color=T_ERREUR)
            self._log(f"❌ Ouverture de l'administration : {e}")

    def _forget_token(self):
        """Efface le token ET les identifiants véhicule de ce poste, et se déconnecte."""
        cfg.clear_token()
        cfg.clear_vehicle_credentials()
        self.token = ""
        # On efface le véhicule PERSONNALISÉ, mais on retombe sur le compte
        # embarqué : les gros fichiers restent gérés après une déconnexion.
        self.vehicle_username = getattr(cfg, "VEHICLE_USERNAME", "")
        self.vehicle_password = getattr(cfg, "VEHICLE_PASSWORD", "")
        self.vehicle_owner_url = ""
        self.api = None
        self.all_users = []
        if hasattr(self, "token_entry"):
            self.token_entry.delete(0, "end")
        if hasattr(self, "user_entry"):
            self.user_entry.delete(0, "end")
        if hasattr(self, "pass_entry"):
            self.pass_entry.delete(0, "end")
        if hasattr(self, "agent_results"):
            self._render_users()
        if hasattr(self, "users_count_lbl"):
            self.users_count_lbl.configure(text="")
        self._set_status(False)
        self.config_msg.configure(
            text="🚪  Token effacé de ce poste. Saisissez-le à nouveau pour vous reconnecter.",
            text_color=T_ALERTE)
        self._log("Token effacé du poste — déconnexion.")

    def _connect(self):
        """Lit URL + token (+ identifiants véhicule optionnels) et lance la connexion."""
        url = self.url_entry.get().strip()
        token = self.token_entry.get().strip()
        if not url or not token:
            self.config_msg.configure(text="URL et token requis.", text_color=T_ERREUR)
            return
        # Identifiants véhicule : OPTIONNELS (requis seulement pour les gros fichiers).
        v_user = self.user_entry.get().strip() if hasattr(self, "user_entry") else ""
        v_pass = self.pass_entry.get().rstrip("\r\n") if hasattr(self, "pass_entry") else ""
        self.config_msg.configure(text="⏳  Connexion…", text_color=T_SECONDAIRE)
        self._run(self._do_connect, url, token, v_user, v_pass)

    def _do_connect(self, url, token, v_user="", v_pass=""):
        """(Thread) Teste le token API, puis (si fournis) la session du compte véhicule."""
        try:
            api = PodAPI(url, token)
            count = api.test_connection()
        except Exception as e:
            self._ui(self.config_msg.configure, text=f"❌  Échec : {e}", text_color=T_ERREUR)
            self._ui(self._set_status, False)
            return
        # Test de la session véhicule seulement si des identifiants sont saisis.
        vehicle_ok = None
        if v_user and v_pass:
            try:
                chk = PodChunkedSession(url, v_user, v_pass)
                chk.login()
                chk.close()
                vehicle_ok = True
            except Exception as e:
                vehicle_ok = False
                self._ui(self._log, f"⚠️ Session véhicule en échec : {e}")
        self._ui(self._on_connected, api, url, token, v_user, v_pass, vehicle_ok, count)

    def _on_connected(self, api, url, token, v_user, v_pass, vehicle_ok, count):
        """Connexion réussie : mémorise le client, enregistre token + véhicule, charge types/comptes."""
        self.api = api
        self.token = token
        self.vehicle_username = v_user
        self.vehicle_password = v_pass
        self.config_data["url"] = url
        self._auto_loaded = set()      # nouvelle connexion → les onglets se rechargeront frais
        # `save_token` renvoie "keyring" (coffre-fort de l'OS, chiffré) ou
        # "file" (fichier de repli en clair dans le dossier personnel).
        # Le repli n'arrive que si le coffre-fort est indisponible — mais il
        # faut alors le DIRE : ce token porte des droits d'administration sur
        # toute l'instance, et le fichier reste lisible par tout programme
        # lancé sous la même session Windows.
        moyen_token = cfg.save_token(token)
        self._token_en_clair = (moyen_token != "keyring")
        if self._token_en_clair:
            self._log("⚠️ Le coffre-fort du système est indisponible : le token a été "
                      "enregistré dans un FICHIER de votre dossier personnel, en clair. "
                      "Sur un poste partagé, préférez « Oublier le token » après usage.")
            try:
                self.token_storage_lbl.configure(
                    text="⚠️  Coffre-fort indisponible : token stocké en clair dans un "
                         "fichier de votre dossier personnel.",
                    text_color=T_ALERTE)
            except Exception:
                pass
        cfg.save_vehicle_credentials(v_user, v_pass)
        cfg.save_config(self.config_data)
        self._set_status(True)
        # L'utilisateur a pu ouvrir un onglet à liste pendant la connexion :
        # on rejoue son chargement, sinon il resterait vide indéfiniment.
        self._on_connexion_etablie()
        # Message selon l'état de la session véhicule
        if vehicle_ok is True:
            extra = " Compte véhicule OK (gros fichiers activés)."
            color = "#22c55e"
        elif vehicle_ok is False:
            extra = " ⚠️ Le compte véhicule saisi est invalide (les gros fichiers échoueront)."
            color = "#f59e0b"
        else:
            extra = " (Gros fichiers : compte véhicule intégré.)"
            color = "#22c55e"
        self.config_msg.configure(text=f"✅  Connecté — {count} vidéo(s) accessibles.{extra}",
                                  text_color=color)
        self._run(self._load_types)
        self._run(self._load_all_users)
        # Même raison : l'onglet ouvert pendant la reconnexion doit se charger.
        self._on_connexion_etablie()
        if v_user:
            self._run(self._resolve_vehicle_owner)   # URL Pod du véhicule (pour la vérif post-504)

    def _resolve_vehicle_owner(self):
        """(Thread) Résout l'URL Pod du compte véhicule (correspondance EXACTE de
        l'identifiant), pour pouvoir, après un 504, reconnaître la vidéo qu'il
        vient de créer. Aucun repli sur un autre compte (jamais d'à-peu-près)."""
        uname = (self.vehicle_username or "").strip()
        if not (self.api and uname):
            return
        try:
            found = None
            for u in (self.api.search_users(uname) or []):
                if (u.get("username", "") or "").strip().lower() == uname.lower():
                    found = u
                    break
            if found and found.get("url"):
                self.vehicle_owner_url = found["url"]
                self._ui(self._log, f"Compte véhicule résolu : {found.get('username')}")
            else:
                self.vehicle_owner_url = ""
                self._ui(self._log,
                         f"⚠️ Compte véhicule « {uname} » non résolu (username exact introuvable). "
                         "La récupération après 504 sera moins précise.")
        except Exception as e:
            self._ui(self._log, f"⚠️ Résolution du compte véhicule impossible : {e}")

    def _verifier_maj(self):
        """Lance la vérification de mise à jour en ARRIÈRE-PLAN.

        Appelée peu après le démarrage. Tout se passe dans un thread : si le
        réseau est absent ou le serveur injoignable, l'application n'attend rien
        et l'utilisateur ne voit rien."""
        def travail():
            """(Thread) Interroge le fichier de version publié."""
            def tracer(message):
                """Consigne un échec de vérification dans le Journal.

                Sans cette trace, une panne était indétectable : la vérification
                échouait en silence et l'utilisateur ne voyait simplement jamais
                de bandeau, sans pouvoir en connaître la raison."""
                self._ui(self._log, f"ℹ Mise à jour — {message}")

            try:
                info = maj.etat_mise_a_jour(
                    APP_VERSION,
                    getattr(cfg, "UPDATE_URL", ""),
                    getattr(cfg, "UPDATE_TIMEOUT_S", 5),
                    journal=tracer)
            except Exception as e:
                info = None              # jamais bloquant
                self._ui(self._log, f"ℹ Mise à jour — vérification interrompue : {e}")
            if info:
                self._ui(self._afficher_bandeau_maj, info)
            else:
                # Cas normal le plus fréquent : on est à jour. On le note
                # discrètement pour confirmer que la vérification a bien eu lieu.
                self._ui(self._log,
                         f"ℹ Mise à jour — version {APP_VERSION} : aucune plus récente.")
        self._run(travail)

    def _afficher_bandeau_maj(self, info: dict):
        """Affiche le bandeau annonçant une nouvelle version.

        Volontairement NON bloquant, même quand la version installée est
        périmée : le ton se durcit (couleur, libellé), mais l'application reste
        pleinement utilisable. Empêcher quelqu'un de travailler à un mauvais
        moment coûterait plus cher que le retard de mise à jour."""
        urgent = bool(info.get("urgent"))
        couleur = "#b45309" if urgent else "#1d4ed8"
        titre = ("⚠️  Version obsolète" if urgent
                 else f"⬆️  Version {info['version']} disponible")

        # Un éventuel bandeau précédent est retiré avant d'en poser un nouveau.
        if self.maj_bandeau is not None:
            try:
                self.maj_bandeau.destroy()
            except Exception:
                pass
            self.maj_bandeau = None

        # Le bandeau est créé DIRECTEMENT dans la barre latérale, sans cadre
        # conteneur : c'est ce conteneur transparent qui apparaissait en carré
        # noir sur macOS.
        cadre = ctk.CTkFrame(self.sidebar, fg_color=couleur, corner_radius=6)
        cadre.pack(side="bottom", fill="x", padx=8, pady=(0, 2))
        self.maj_bandeau = cadre
        ctk.CTkLabel(cadre, text=titre, font=ctk.CTkFont(size=11, weight="bold"),
                     text_color="#ffffff", wraplength=190,
                     justify="left").pack(anchor="w", padx=8, pady=(6, 2))
        if urgent:
            ctk.CTkLabel(cadre,
                         text=f"La version {info['version']} corrige un point important. "
                              "Mettez à jour dès que possible.",
                         font=ctk.CTkFont(size=10), text_color="#ffffff",
                         wraplength=190, justify="left").pack(anchor="w", padx=8)
        elif info.get("notes"):
            ctk.CTkLabel(cadre, text=info["notes"], font=ctk.CTkFont(size=10),
                         text_color="#ffffff", wraplength=190,
                         justify="left").pack(anchor="w", padx=8)
        if info.get("url"):
            # Bouton clair sur fond coloré. Les couleurs sont données en
            # hexadécimal plutôt que par leur nom : les noms symboliques
            # (« white », « gray90 ») ne sont pas rendus de la même façon
            # partout, et macOS s'en accommode mal.
            ctk.CTkButton(cadre, text="Télécharger", height=26,
                          fg_color="#ffffff", text_color=couleur,
                          hover_color="#e5e7eb",
                          font=ctk.CTkFont(size=11, weight="bold"),
                          command=lambda u=info["url"]: self._ouvrir_lien_maj(u)
                          ).pack(fill="x", padx=8, pady=(6, 8))
        else:
            # Simple marge basse. On ajuste l'espacement du dernier libellé
            # plutôt que d'ajouter un widget vide, qui pouvait laisser une
            # trace visible sur certains systèmes.
            cadre.configure(height=0)      # laisse le contenu fixer la hauteur
        self._log(f"⬆️ Version {info['version']} disponible"
                  + (" (mise à jour recommandée sans délai)." if urgent else "."))

    def _ouvrir_lien_maj(self, url: str):
        """Ouvre la page de téléchargement de la nouvelle version."""
        try:
            import webbrowser
            webbrowser.open(url)
            self._log("Page de téléchargement ouverte.")
        except Exception as e:
            self._log(f"❌ Ouverture du lien de mise à jour : {e}")

    def _auto_connect(self):
        """(Thread) Reconnexion automatique au démarrage si un token est déjà enregistré."""
        try:
            api = PodAPI(self.config_data["url"], self.token)
            count = api.test_connection()
            self._ui(self._on_auto_ok, api, count)
        except Exception:
            self._ui(self._set_status, False)

    def _on_auto_ok(self, api, count):
        """Reconnexion auto réussie : active l'état connecté et charge types et comptes."""
        self.api = api
        self._set_status(True)
        u = self.config_data.get("agent_username", "")
        if u:
            self._definir_agent(f"Dépôt au nom de :\n{u}")
        self._run(self._load_types)
        self._run(self._load_all_users)
        if self.vehicle_username:
            self._run(self._resolve_vehicle_owner)

    def _set_status(self, ok: bool):
        """Met à jour l'indicateur de connexion (pastille + libellé) de la barre latérale."""
        self.status_dot.configure(text="🟢" if ok else "🔴")
        self.status_lbl.configure(text="Connecté" if ok else "Non connecté",
                                  text_color=T_SUCCES if ok else "#ef4444")

    def _load_types(self):
        """(Thread) Charge les types de vidéo et les sites (champ requis à l'upload)."""
        try:
            self.types = self.api.get_types()
            self.type_map = {t.get("title", f"type-{t.get('id')}"): t.get("url", "")
                             for t in self.types}
            titles = list(self.type_map.keys()) or ["(aucun type)"]
            self._ui(self.type_combo.configure, values=titles)
            self._ui(self.type_combo.set, titles[0])
            # Met aussi à jour les menus de type de l'onglet Vidéos
            self._ui(self._browse_refresh_type_menu)
        except Exception as e:
            self._ui(self._log, f"Impossible de charger les types : {e}")
        # Sites (champ requis à l'upload sur instance multi-établissements)
        try:
            sites = self.api.get_sites()
            self.site_urls = [s.get("url", "") for s in sites if s.get("url")]
            if self.site_urls:
                names = ", ".join(s.get("name", s.get("domain", "?")) for s in sites)
                self._ui(self._log, f"Site(s) détecté(s) : {names}")
            else:
                self._ui(self._log, "⚠️ Aucun site retourné par /rest/sites/ — l'upload pourrait échouer.")
        except Exception as e:
            self._ui(self._log, f"Impossible de charger les sites : {e}")
        # Groupes d'accès (pour la restriction de visibilité par groupe).
        # La reconstruction sonde /accessgroups/<id>/ → on le fait en tâche de
        # fond pour ne pas ralentir la connexion ; échec silencieux si absent.
        try:
            self.access_groups = self.api.get_access_groups()
            if self.access_groups:
                self._ui(self._log,
                         f"{len(self.access_groups)} groupe(s) d'accès détecté(s).")
        except Exception as e:
            self.access_groups = []
            self._ui(self._log, f"Groupes d'accès non chargés : {e}")

    def _load_all_users(self):
        """(Thread) Charge tous les comptes Pod (paginé) et rafraîchit les vues qui en dépendent."""
        if not self.api:
            self._ui(self.users_count_lbl.configure,
                     text="Connectez-vous d'abord.", text_color=T_ALERTE)
            return
        self._ui(self.users_count_lbl.configure,
                 text="⏳  Chargement de la liste des utilisateurs…", text_color=T_SECONDAIRE)
        self._ui(self._log, "Chargement des utilisateurs (/rest/users/)…")
        try:
            users = self.api.get_all_users()
            users.sort(key=lambda u: (u.get("username") or "").lower())
            self.all_users = users
            self._ui(self._render_users)
            self._ui(self._render_comptes)   # PodAdmin : rafraîchir l'onglet Comptes
            self._ui(self._refresh_reassign_pickers)  # … et les sélecteurs de réaffectation
            # Présélection : pré-remplir le filtre avec le propriétaire enregistré
            ag = self.config_data.get("agent_username", "")
            if ag:
                self._ui(self._preselect_agent, ag)
            if users:
                self._ui(self.users_count_lbl.configure,
                         text=f"✅  {len(users)} utilisateur(s) chargé(s). Filtrez puis cliquez pour choisir.",
                         text_color=T_SUCCES)
                self._ui(self._log, f"Utilisateurs chargés : {len(users)}.")
            else:
                self._ui(self.users_count_lbl.configure,
                         text="⚠️  Aucun utilisateur renvoyé. Le compte du token n'a peut-être "
                              "pas le droit de lister les utilisateurs (compte superutilisateur requis).",
                         text_color=T_ALERTE)
                self._ui(self._log, "⚠️ /rest/users/ a renvoyé 0 utilisateur — vérifiez les droits du token "
                                    "(ou lancez verifier.py).")
        except Exception as e:
            self._ui(self.users_count_lbl.configure, text=f"❌  Erreur : {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Erreur chargement utilisateurs : {e}")

    def _user_pk(self, u: dict) -> str:
        """Identifiant NUMÉRIQUE d'un compte (clé primaire Django).

        L'API renvoie tantôt un champ `id`, tantôt seulement une `url` du type
        .../rest/users/42/. On récupère l'un ou l'autre : c'est ce numéro
        qu'attend le formulaire d'administration pour pré-sélectionner
        l'utilisateur."""
        pk = u.get("id")
        if pk:
            return str(pk)
        url = str(u.get("url", "")).rstrip("/")
        return url.rsplit("/", 1)[-1] if url else ""

    def _compte_creer_token(self, u: dict):
        """Ouvre l'administration Pod sur les jetons, RECHERCHE PRÉ-REMPLIE au
        nom du compte.

        POURQUOI LA LISTE ET NON LE FORMULAIRE DE CRÉATION : dans Django REST
        Framework, la liaison jeton ↔ utilisateur est de type « un à un ». Un
        compte ne peut donc posséder QU'UN SEUL jeton, et tenter d'en créer un
        second échoue avec une erreur de validation.

        Ouvrir la LISTE filtrée permet de voir d'abord si la personne en a déjà
        un :
          • si oui, il suffit de le recopier — inutile (et risqué) d'en refaire
            un, car supprimer l'ancien casserait immédiatement son application ;
          • si non, la liste est vide et le bouton « Ajouter » de
            l'administration est à un clic.

        POURQUOI PASSER PAR LE NAVIGATEUR : l'API REST d'Esup-Pod n'expose pas
        les jetons — ils ne se créent que dans l'administration Django. Les
        créer depuis l'application supposerait d'y stocker des identifiants
        SUPERUTILISATEUR, ce qui serait un vrai recul de sécurité (le compte
        véhicule embarqué est, lui, délibérément sans privilège).
        """
        base = (self.config_data.get("url") or "").rstrip("/")
        if not base:
            self._log("❌ URL de l'instance non renseignée (onglet Configuration).")
            return
        username = (u.get("username") or "").strip()
        if not username:
            self._log("❌ Identifiant introuvable pour ce compte.")
            return
        # Django REST expose le modèle sous « tokenproxy » depuis DRF 3.14, et
        # sous « token » avant. `?q=` est la recherche standard de toute liste
        # d'administration Django : elle fonctionne dans les deux cas.
        url = f"{base}/admin/authtoken/tokenproxy/?q={urllib.parse.quote(username)}"
        try:
            import webbrowser
            webbrowser.open(url)
            self._log(f"🔑 Jetons de l'administration ouverts pour {username}.")
            self.comptes_msg.configure(
                text=f"Jetons ouverts pour « {username} ». S'il en a DÉJÀ un, copiez-le "
                     "(n'en recréez pas : supprimer l'ancien couperait son accès). "
                     "Sinon, cliquez sur « Ajouter » dans l'administration.",
                text_color=T_SUCCES)
        except Exception as e:
            self._log(f"❌ Ouverture du navigateur : {e}")

    def _compte_mail_token(self, u: dict):
        """Ouvre le client de messagerie avec une réponse prête à l'emploi.

        Le jeton n'est PAS inséré automatiquement (l'application ne le connaît
        pas) : un emplacement clairement repérable est laissé dans le message,
        à remplacer par le jeton copié depuis l'administration."""
        email = (u.get("email") or "").strip()
        if not email:
            self._log(f"❌ Aucune adresse connue pour {u.get('username','?')}.")
            return
        # Texte officiel du service (modèle fourni par le support Pod).
        # Le jeton n'est PAS inséré : l'application ne le connaît pas. Le
        # marqueur « >>> COLLER ICI LA CLÉ <<< », encadré de lignes de
        # tirets, reste impossible à manquer. Un lien « mailto: » ne transporte
        # que du TEXTE BRUT : ni gras ni couleur ne sont possibles, d'où ce
        # repère typographique.
        # Texte validé par le service (version simplifiée) : les explications
        # d'installation ne sont plus dans le message, elles sont sur Moodle.
        #
        # VOCABULAIRE : on parle de « clé d'activation » vers les enseignants,
        # et non de « token ». Le terme technique reste employé côté
        # administration (PodAdmin, administration Django), mais il n'évoque
        # rien pour un utilisateur.
        objet = "Votre accès au dépôt de vidéos (videos.utoulouse.fr)"
        corps = (
            "Bonjour,\n\n"
            "Vous pouvez, maintenant, déposer des vidéos sur le serveur vidéo "
            "institutionnel de l'Université de Toulouse : "
            "https://videos.utoulouse.fr (POD).\n\n"
            "Pour publier des vidéos, connectez-vous à la plateforme : "
            "https://videos.utoulouse.fr avec vos identifiants institutionnels.\n\n"
            "Une autre option consiste à utiliser notre application dédiée, "
            "PODTéléverseur, qui est compatible avec les systèmes d'exploitation "
            "Windows et macOS.\n\n"
            "L'application vous permettra de facilement accéder à toutes les "
            "fonctionnalités de téléversement vidéo. Nous vous recommandons "
            "d'utiliser PODTéléverseur, qui offre notamment la possibilité de "
            "téléverser des lots de vidéos, une fonctionnalité qui n'est pas "
            "disponible sur la version web.\n\n"
            "Le tutoriel vous guidera dans l'installation de l'application :\n"
            f"{MOODLE_URL}\n\n"
            "Au premier lancement, l'application vous demandera une clé "
            "d'activation. Voici la vôtre :\n\n"
            "----------------------------------------------------------\n"
            "  >>>  COLLER ICI LA CLÉ  <<<\n"
            "----------------------------------------------------------\n\n"
            "Cette clé est personnelle et vaut accès à votre compte : "
            "NE PAS LA TRANSMETTRE. Vous n'en avez besoin que pour l'application ; "
            "l'accès par navigateur se fait avec vos identifiants habituels.\n\n"
            "N'utilisez plus PRISMES pour déposer des vidéos, vos nouvelles vidéos "
            "DOIVENT IMPÉRATIVEMENT être déposées sur videos.utoulouse.fr.\n\n"
            "Nous migrons actuellement les vidéos de PRISMES vers "
            "videos.utoulouse.fr. Vos liens Moodle resteront valides. Nous vous "
            "contacterons par email lors de la migration de votre compte.\n\n"
            "Pour toute aide ou accompagnement spécifique, n'hésitez pas à contacter "
            f"le Support POD à {SUPPORT_MAIL}.\n\n"
            "Bonne Journée\n\n"
            "—\n"
            "Support videos.utoulouse.fr (POD)\n"
            f"{SUPPORT_MAIL}")

        try:
            import urllib.parse
            import webbrowser
            # Copie cachée au support : trace de la délivrance du jeton, sans
            # exposer l'adresse interne au destinataire.
            lien = (f"mailto:{urllib.parse.quote(email)}"
                    f"?subject={urllib.parse.quote(objet)}"
                    f"&bcc={urllib.parse.quote(SUPPORT_MAIL)}"
                    f"&body={urllib.parse.quote(corps)}")
            webbrowser.open(lien)
            self._log(f"✉️ Brouillon de réponse ouvert pour {email}.")
            self.comptes_msg.configure(
                text=f"Message préparé pour {email} — collez le jeton à l'emplacement prévu.",
                text_color=T_SUCCES)
        except Exception as e:
            self._log(f"❌ Ouverture du client de messagerie : {e}")

    def _user_label(self, u: dict) -> str:
        """Libellé lisible d'un compte : « identifiant — Prénom Nom »."""
        return f"{u.get('username','?')} — {u.get('first_name','')} {u.get('last_name','')}".strip()

    def _preselect_agent(self, username: str):
        """Présélection : pré-remplit le filtre avec le propriétaire enregistré.
        La liste reste entièrement utilisable : effacer le filtre permet de
        choisir un autre compte (le support dépose sur différents comptes)."""
        if hasattr(self, "agent_filter") and not self.agent_filter.get().strip():
            self.agent_filter.insert(0, username)
            self._render_users()

    def _render_users(self):
        """Affiche la liste filtrée des comptes pour choisir l'agent déposant."""
        flt = self.agent_filter.get().strip().lower() if hasattr(self, "agent_filter") else ""
        for w in self.agent_results.winfo_children():
            w.destroy()

        if not self.all_users:
            ctk.CTkLabel(self.agent_results,
                         text="Liste non chargée. Cliquez sur « Recharger ».",
                         text_color=T_SECONDAIRE).pack(pady=10)
            return

        matches = [u for u in self.all_users if not flt or flt in self._user_label(u).lower()]
        CAP = 300  # éviter de créer des milliers de boutons (Tk gèlerait)
        current_username = self.config_data.get("agent_username", "")

        for u in matches[:CAP]:
            is_current = (u.get("username", "") == current_username)
            label = ("✅  " if is_current else "      ") + self._user_label(u)
            ctk.CTkButton(self.agent_results, text=label, anchor="w",
                          fg_color=S_SELECTION if is_current else "transparent",
                          text_color=("gray10", "gray90"), hover_color=("gray75", "gray28"),
                          height=28, font=ctk.CTkFont(size=12),
                          command=lambda uu=u: self._pick_agent(uu)).pack(fill="x", pady=1)

        if len(matches) > CAP:
            ctk.CTkLabel(self.agent_results,
                         text=f"… +{len(matches) - CAP} autres. Affinez le filtre.",
                         text_color=T_SECONDAIRE).pack(pady=4)
        elif not matches:
            ctk.CTkLabel(self.agent_results,
                         text="Aucun résultat ne correspond au filtre.",
                         text_color=T_SECONDAIRE).pack(pady=8)

    def _pick_agent(self, user: dict):
        """Enregistre le compte choisi comme propriétaire par défaut des dépôts."""
        self.config_data["agent_username"] = user.get("username", "")
        self.config_data["agent_owner_url"] = user.get("url", "")
        self.config_data["agent_owner_label"] = self._user_label(user)
        cfg.save_config(self.config_data)
        self._definir_agent(f"Dépôt au nom de :\n{user.get('username','')}")
        self.config_msg.configure(
            text=f"✅  Propriétaire des vidéos : {user.get('username','')}", text_color=T_SUCCES)
        if hasattr(self, "agent_results"):
            self._render_users()   # met à jour la coche ✅

    # ═════════════════════════════════════════════════════════════════════
    #  ONGLET ENCODAGE — supervision du transcodage + relance
    # ═════════════════════════════════════════════════════════════════════
    #
    #  L'API n'expose pas de file d'encodage dédiée (vérifié au sondage).
    #  On se base donc sur les champs par vidéo : `encoded`,
    #  `encoding_in_progress`, `get_encoding_step`, `is_draft`. On classe
    #  chaque vidéo par état, et on peut relancer l'encodage (launch_encoding)
    #  à l'unité ou en masse sur les vidéos « à problème ».
    #
    #  États retenus :
    #    ✅ encodée            : encoded == True
    #    ⏳ en cours           : encoding_in_progress == True
    #    📝 non lancée         : brouillon non encodé (is_draft, pas encore lancé)
    #    ❌ à problème         : ni encodée, ni en cours, ni brouillon (suspecte)

    @staticmethod
    def _encode_state(v: dict) -> str:
        """Classe une vidéo dans un état d'encodage (clé interne)."""
        if v.get("encoded"):
            return "ok"
        if v.get("encoding_in_progress"):
            return "running"
        if v.get("is_draft"):
            return "draft"
        return "failed"

    # Libellés lisibles + pastille pour chaque état
    _ENCODE_LABELS = {
        "ok":      "✅ Encodée",
        "running": "⏳ En cours",
        "failed":  "❌ À problème",
        "draft":   "📝 Non lancée",
    }

    def _build_tab_encode(self):
        """Construit l'onglet Encodage (supervision du transcodage + relance)."""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tabs["encode"] = frame

        ctk.CTkLabel(frame, text="🎬  Encodage",
                     font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(
            frame,
            text="Supervisez le transcodage : voyez les encodages en cours, terminés ou en "
                 "échec, et relancez l'encodage des vidéos qui posent problème (à l'unité ou "
                 "en masse).",
            text_color=T_SECONDAIRE, font=ctk.CTkFont(size=12),
            justify="left", wraplength=860).pack(anchor="w", pady=(0, 8))

        # — Ligne : scan + état —
        top = ctk.CTkFrame(frame, fg_color="transparent")
        top.pack(fill="x")
        ctk.CTkButton(top, text="🔄  Rafraîchir", fg_color=C_NEUTRE,
                      hover_color=C_NEUTRE_SURV, command=self._encode_scan, text_color=T_SUR_NEUTRE).pack(side="left")
        self.encode_status = ctk.CTkLabel(top, text="(aucun scan)", text_color=T_SECONDAIRE,
                                          font=ctk.CTkFont(size=11))
        self.encode_status.pack(side="left", padx=10)

        # — Bandeau de compteurs —
        self.encode_counters = ctk.CTkLabel(frame, text="", justify="left", anchor="w",
                                            font=ctk.CTkFont(size=13))
        self.encode_counters.pack(anchor="w", pady=(8, 4))

        # — Ligne : filtre par état + relance en masse —
        bar = ctk.CTkFrame(frame, fg_color="transparent")
        bar.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(bar, text="État :").pack(side="left", padx=(0, 4))
        self.encode_filter = ctk.CTkOptionMenu(
            bar, width=150,
            values=["Tous", "En cours", "À problème", "Non lancées", "Encodées"],
            command=lambda _c: self._render_encode(), **STYLE_CHAMP)
        self.encode_filter.set("À problème")
        self.encode_filter.pack(side="left")
        # Relance en masse de tout ce qui est affiché
        self.encode_relaunch_btn = ctk.CTkButton(
            bar, text="🔁  Relancer l'encodage des vidéos affichées",
            fg_color=C_SUCCES, hover_color=C_SUCCES_SURV, command=self._encode_relaunch_shown)
        self.encode_relaunch_btn.pack(side="left", padx=10)
        self.encode_progress = ctk.CTkLabel(bar, text="", text_color=T_SECONDAIRE,
                                            font=ctk.CTkFont(size=11))
        self.encode_progress.pack(side="left", padx=6)

        # — Liste —
        self.encode_list = ctk.CTkScrollableFrame(frame, label_text="Vidéos", fg_color=S_CARTE, label_anchor="w",
                                                  label_font=ctk.CTkFont(size=12, weight="bold"))
        self.encode_list.pack(fill="both", expand=True, pady=(4, 0))

        # — Données —
        self.encode_videos = []      # scan complet (cache)
        self.encode_filtered = []    # sous-ensemble affiché

    # ── Scan ────────────────────────────────────────────────────────────────

    def _encode_scan(self):
        """Déclenche le scan complet des vidéos (en arrière-plan)."""
        if not self.api:
            self.encode_status.configure(text="Connectez-vous d'abord.", text_color=T_ALERTE)
            return
        self.encode_status.configure(text="⏳  Scan…", text_color=T_SECONDAIRE)
        self._run(self._do_encode_scan)

    def _do_encode_scan(self):
        """(Thread) Récupère toutes les vidéos puis met à jour compteurs + liste."""
        try:
            def prog(n):
                """Callback de progression (avancement du scan)."""
                self._ui(self.encode_status.configure,
                         text=f"⏳  {n} vidéos lues…", text_color=T_SECONDAIRE)
            # MAGASIN PARTAGÉ : plus de scan propre à cet onglet. Auparavant il
            # relisait toute l'instance et gardait sa copie, que rien ne
            # rafraîchissait — il proposait donc de relancer l'encodage de
            # vidéos supprimées depuis.
            videos = self.ensure_videos_sync(progress_cb=prog)
            self.encode_videos = videos
            self._ui(self._render_encode)
            skipped = getattr(self.api, "last_scan_skipped", 0)
            if skipped:
                self._ui(self.encode_status.configure,
                         text=f"✅  {len(videos)} vidéos analysées "
                              f"(⚠️ {skipped} ignorée·s : illisibles côté serveur)  ·  "
                              f"{self._loaded_stamp()}",
                         text_color=T_ALERTE)
                self._ui(self._log,
                         f"Encodage : {len(videos)} vidéos scannées, {skipped} ignorée·s "
                         f"(erreur serveur sur ces vidéos — à corriger côté Pod/DSI).")
            else:
                self._ui(self.encode_status.configure,
                         text=f"✅  {len(videos)} vidéos analysées  ·  "
                              f"chargé à {datetime.now().strftime('%H:%M')}",
                         text_color=T_SUCCES)
                self._ui(self._log, f"Encodage : {len(videos)} vidéos scannées.")
        except Exception as e:
            self._ui(self.encode_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Scan encodage : {e}")

    # ── Rendu (compteurs + liste filtrée) ───────────────────────────────────

    def _render_encode(self, *_):
        """Met à jour les compteurs par état et la liste filtrée."""
        if not hasattr(self, "encode_list"):
            return
        for w in self.encode_list.winfo_children():
            w.destroy()

        if not self.encode_videos:
            self.encode_counters.configure(text="")
            ctk.CTkLabel(self.encode_list, text="Cliquez sur « Scanner ».",
                         text_color=T_SECONDAIRE).pack(pady=10)
            return

        # Comptage par état
        counts = {"ok": 0, "running": 0, "failed": 0, "draft": 0}
        for v in self.encode_videos:
            counts[self._encode_state(v)] += 1
        self.encode_counters.configure(
            text=(f"✅ {counts['ok']} encodées      "
                  f"⏳ {counts['running']} en cours      "
                  f"❌ {counts['failed']} à problème      "
                  f"📝 {counts['draft']} non lancées"))

        # Filtre par état
        sel = self.encode_filter.get()
        wanted = {"En cours": "running", "À problème": "failed",
                  "Non lancées": "draft", "Encodées": "ok"}.get(sel)
        if wanted:
            vids = [v for v in self.encode_videos if self._encode_state(v) == wanted]
        else:
            vids = list(self.encode_videos)
        self.encode_filtered = vids

        # Le bouton de relance en masse n'a de sens que pour les états relançables
        relaunchable = sel in ("À problème", "Non lancées", "En cours", "Tous")
        self.encode_relaunch_btn.configure(state="normal" if (vids and relaunchable) else "disabled")

        if not vids:
            ctk.CTkLabel(self.encode_list, text="Aucune vidéo dans cet état.",
                         text_color=T_SECONDAIRE).pack(pady=10)
            return

        # Une ligne par vidéo : [pastille état] titre · slug · étape  [Relancer]
        CAP = 400
        for v in vids[:CAP]:
            st = self._encode_state(v)
            row = ctk.CTkFrame(self.encode_list, fg_color=S_LIGNE,
                               corner_radius=6)
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=self._ENCODE_LABELS[st], width=110, anchor="w",
                         font=ctk.CTkFont(size=12)).pack(side="left", padx=(10, 4), pady=6)
            title = (v.get("title") or "(sans titre)")[:46]
            step = v.get("get_encoding_step", "")
            ctk.CTkLabel(row, text=f"{title}   ·   {v.get('slug','?')}   ·   {step}",
                         anchor="w", font=ctk.CTkFont(size=12)).pack(
                side="left", padx=4, pady=6, fill="x", expand=True)
            # Bouton de relance individuelle (sauf si déjà encodée — relance possible quand même,
            # mais on la réserve aux états non terminés pour éviter les clics inutiles)
            if st != "ok":
                ctk.CTkButton(row, text="🔁 Relancer", width=100, height=26, fg_color=C_NEUTRE,
                              command=lambda vv=v: self._encode_relaunch_one(vv), text_color=T_SUR_NEUTRE).pack(
                    side="right", padx=8)
        if len(vids) > CAP:
            ctk.CTkLabel(self.encode_list,
                         text=f"… +{len(vids) - CAP} autres. Affinez le filtre.",
                         text_color=T_SECONDAIRE).pack(pady=4)

    # ── Relance de l'encodage ────────────────────────────────────────────────

    def _encode_relaunch_one(self, v):
        """Relance l'encodage d'une vidéo après confirmation."""
        if not messagebox.askyesno(
                "Relancer l'encodage",
                f"Relancer l'encodage de :\n\n{v.get('title')}  ({v.get('slug')}) ?"):
            return
        self._run(self._do_encode_relaunch, [v])

    def _encode_relaunch_shown(self):
        """Relance l'encodage de toutes les vidéos actuellement affichées."""
        vids = list(self.encode_filtered)
        if not vids:
            return
        if not messagebox.askyesno(
                "Relancer en masse",
                f"Relancer l'encodage de {len(vids)} vidéo(s) ?\n\n"
                "Chaque vidéo sera renvoyée dans la file d'encodage du serveur."):
            return
        self._run(self._do_encode_relaunch, vids)

    def _do_encode_relaunch(self, vids):
        """(Thread) Appelle launch_encoding(slug) pour chaque vidéo, avec bilan."""
        ok = fail = 0
        for i, v in enumerate(vids, 1):
            slug = v.get("slug", "")
            try:
                self.api.launch_encoding(slug)          # GET /rest/launch_encode_view/?slug=
                ok += 1
                self._ui(self._log, f"🔁 Encodage relancé : {slug}")
            except Exception as e:
                fail += 1
                self._ui(self._log, f"❌ Relance {slug} : {e}")
            self._ui(self.encode_progress.configure,
                     text=f"⏳  {i}/{len(vids)}…", text_color=T_SECONDAIRE)
        self._ui(self.encode_progress.configure,
                 text=f"Terminé : {ok} relancée(s), {fail} échec(s).",
                 text_color=T_SUCCES if not fail else "#f59e0b")
        # Re-scan pour refléter les nouveaux états (en cours d'encodage)
        self._do_encode_scan()

    # ═════════════════════════════════════════════════════════════════════
    #  ONGLET COMPTES — statut « équipe » (is_staff)
    # ═════════════════════════════════════════════════════════════════════

    def _build_tab_comptes(self):
        """Construit l'onglet Comptes (statut équipe is_staff)."""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tabs["comptes"] = frame

        ctk.CTkLabel(frame, text="👤  Comptes — statut « équipe »",
                     font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(
            frame,
            text="Le statut « équipe » (is_staff) autorise un compte à ajouter et gérer "
                 "des vidéos sur Pod. Activez l'interrupteur pour l'accorder, désactivez-le "
                 "pour le retirer. Chaque changement est confirmé puis appliqué immédiatement.",
            text_color=T_SECONDAIRE, font=ctk.CTkFont(size=12),
            justify="left", wraplength=860).pack(anchor="w", pady=(0, 10))

        bar = ctk.CTkFrame(frame, fg_color="transparent")
        bar.pack(fill="x")
        self.comptes_filter = ctk.CTkEntry(
            bar, placeholder_text="🔍 nom / prénom / identifiant…")
        self.comptes_filter.pack(side="left", fill="x", expand=True)
        self.comptes_filter.bind("<KeyRelease>",
                                 lambda e: self._debounce("comptes", self._render_comptes))
        # Filtre par statut équipe (is_staff)
        self.comptes_statut = ctk.CTkOptionMenu(
            bar, width=150,
            values=["Tous", "Équipe", "Sans statut"],
            command=lambda _c: self._render_comptes(), **STYLE_CHAMP)
        self.comptes_statut.set("Tous")
        self.comptes_statut.pack(side="left", padx=6)
        # Regrouper : trie pour rassembler les comptes « équipe » en tête
        self.comptes_group = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(bar, text="Regrouper", variable=self.comptes_group,
                        command=self._render_comptes, width=90).pack(side="left", padx=6)
        ctk.CTkButton(bar, text="🔄  Recharger", width=130,
                      fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=lambda: self._run(self._reload_users_for_admin), text_color=T_SUR_NEUTRE).pack(side="left", padx=8)

        self.comptes_count_lbl = ctk.CTkLabel(frame, text="", text_color=T_SECONDAIRE,
                                              font=ctk.CTkFont(size=11))
        self.comptes_count_lbl.pack(anchor="w", pady=(6, 2))

        # Retour visuel des actions « 🔑 Token » et « ✉️ » de chaque ligne.
        self.comptes_msg = ctk.CTkLabel(frame, text="", text_color=T_SECONDAIRE,
                                        font=ctk.CTkFont(size=11),
                                        anchor="w", wraplength=900, justify="left")
        self.comptes_msg.pack(anchor="w", fill="x", pady=(0, 4))

        self.comptes_results = ctk.CTkScrollableFrame(frame, fg_color=S_CARTE)
        self.comptes_results.pack(fill="both", expand=True, pady=(0, 4))

        self._render_comptes()

    def _reload_users_for_admin(self):
        """Recharge la liste complète des comptes puis rafraîchit les vues."""
        if not self.api:
            self._ui(self.comptes_count_lbl.configure,
                     text="Connectez-vous d'abord (onglet Configuration).",
                     text_color=T_ALERTE)
            return
        self._ui(self.comptes_count_lbl.configure,
                 text="⏳  Chargement des comptes…", text_color=T_SECONDAIRE)
        try:
            users = self.api.get_all_users()
            users.sort(key=lambda u: (u.get("username") or "").lower())
            self.all_users = users
            self._ui(self._render_comptes)
            if hasattr(self, "agent_results"):
                self._ui(self._render_users)
            self._ui(self._refresh_reassign_pickers)
            self._ui(self._log, f"Comptes rechargés : {len(users)}.")
        except Exception as e:
            self._ui(self.comptes_count_lbl.configure,
                     text=f"❌  Erreur : {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Erreur chargement comptes : {e}")

    def _render_comptes(self):
        """Affiche la liste filtrée des comptes dans l'onglet Comptes."""
        if not hasattr(self, "comptes_results"):
            return
        flt = self.comptes_filter.get().strip().lower() if hasattr(self, "comptes_filter") else ""
        for w in self.comptes_results.winfo_children():
            w.destroy()

        if not self.all_users:
            ctk.CTkLabel(self.comptes_results,
                         text="Liste non chargée. Connectez-vous puis cliquez sur « Recharger ».",
                         text_color=T_SECONDAIRE).pack(pady=10)
            self.comptes_count_lbl.configure(text="")
            return

        matches = [u for u in self.all_users
                   if not flt or flt in self._user_label(u).lower()]

        # Filtre par statut équipe (menu déroulant)
        statut = self.comptes_statut.get() if hasattr(self, "comptes_statut") else "Tous"
        if statut == "Équipe":
            matches = [u for u in matches if u.get("is_staff")]
        elif statut == "Sans statut":
            matches = [u for u in matches if not u.get("is_staff")]

        # Regroupement : comptes « équipe » d'abord, puis tri alphabétique
        if hasattr(self, "comptes_group") and self.comptes_group.get():
            matches.sort(key=lambda u: (not u.get("is_staff"),
                                        (u.get("username") or "").lower()))

        staff_n = sum(1 for u in self.all_users if u.get("is_staff"))
        self.comptes_count_lbl.configure(
            text=f"{len(self.all_users)} compte(s) — {staff_n} avec statut équipe. "
                 f"{len(matches)} affiché(s).")

        CAP = 300
        for u in matches[:CAP]:
            row = ctk.CTkFrame(self.comptes_results, fg_color=S_LIGNE,
                               corner_radius=6)
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=self._user_label(u), anchor="w",
                         font=ctk.CTkFont(size=12)).pack(
                side="left", padx=12, pady=6, fill="x", expand=True)
            # Bouton « mail » : prépare la réponse au demandeur (grisé sans adresse).
            a_un_mail = bool((u.get("email") or "").strip())
            ctk.CTkButton(row, text="✉️", width=34, height=26,
                          fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                          state="normal" if a_un_mail else "disabled",
                          command=lambda uu=u: self._compte_mail_token(uu), text_color=T_SUR_NEUTRE).pack(
                side="right", padx=(0, 8))
            # Bouton « token » : ouvre le formulaire d'admin Django, utilisateur
            # pré-sélectionné. La création reste faite dans le navigateur, sous
            # l'identité d'administrateur de l'utilisateur — l'application ne
            # manipule aucun identifiant privilégié.
            ctk.CTkButton(row, text="🔑 Token", width=80, height=26,
                          fg_color=C_ACCENT, hover_color=C_ACCENT_SURV,
                          command=lambda uu=u: self._compte_creer_token(uu)).pack(
                side="right", padx=(0, 6))

            var = ctk.BooleanVar(value=bool(u.get("is_staff")))
            sw = ctk.CTkSwitch(row, text="Équipe", variable=var, width=80,
                               command=lambda uu=u, vv=var: self._on_staff_toggle(uu, vv))
            sw.pack(side="right", padx=12, pady=4)

        if len(matches) > CAP:
            ctk.CTkLabel(self.comptes_results,
                         text=f"… +{len(matches) - CAP} autres. Affinez le filtre.",
                         text_color=T_SECONDAIRE).pack(pady=4)
        elif not matches:
            ctk.CTkLabel(self.comptes_results,
                         text="Aucun résultat ne correspond au filtre.",
                         text_color=T_SECONDAIRE).pack(pady=8)

    def _on_staff_toggle(self, user: dict, var):
        """Confirme puis applique le changement de statut équipe."""
        want = bool(var.get())
        verb = "DONNER" if want else "RETIRER"
        consequence = ("Ce compte pourra ajouter et gérer des vidéos sur Pod."
                       if want else
                       "Ce compte ne pourra plus ajouter de vidéos (accès équipe retiré).")
        ok = messagebox.askyesno(
            "Confirmer le changement de statut",
            f"{verb} le statut « équipe » à :\n\n"
            f"    {self._user_label(user)}\n\n{consequence}\n\nContinuer ?")
        if not ok:
            var.set(not want)          # annulation → revenir à l'état précédent
            return
        self._run(self._do_staff, user, want, var)

    def _do_staff(self, user: dict, want: bool, var):
        """(Thread) Donne ou retire le statut « équipe » (staff) à un compte."""
        uname = user.get("username", "?")
        try:
            self.api.set_user_staff(user.get("url", ""), want)
            user["is_staff"] = want    # met à jour le cache local
            self._ui(self._log,
                     f"{'✅ Statut équipe ACCORDÉ à' if want else '⛔ Statut équipe RETIRÉ à'} "
                     f"{uname}.")
            self._ui(self._render_comptes)   # rafraîchit compteur + état
        except Exception as e:
            self._ui(var.set, not want)      # échec → revenir à l'état précédent
            self._ui(self.comptes_count_lbl.configure,
                     text=f"❌  Échec pour {uname} : {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Échec MAJ statut de {uname} : {e}")

    # ═════════════════════════════════════════════════════════════════════
    #  ONGLET VIDÉOS — explorateur + édition d'UNE vidéo
    # ═════════════════════════════════════════════════════════════════════
    #
    #  Rechercher/filtrer une vidéo dans une liste (gauche), puis l'éditer dans
    #  un panneau de détail (droite) : renommer, changer le statut
    #  (brouillon/public, restreinte), gérer les co-propriétaires et les
    #  chaînes, supprimer. Les actions portent sur UNE vidéo à la fois.

    def _build_tab_browse(self):
        """Construit l'onglet Vidéos (recherche, détail, actions)."""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tabs["browse"] = frame

        ctk.CTkLabel(frame, text="🎞️  Vidéos",
                     font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(
            frame,
            text="Recherchez une vidéo, sélectionnez-la dans la liste, puis éditez-la dans "
                 "le panneau de droite (titre, statut, co-propriétaires, chaînes, suppression).",
            text_color=T_SECONDAIRE, font=ctk.CTkFont(size=12),
            justify="left", wraplength=860).pack(anchor="w", pady=(0, 8))

        # — Ligne : charger + statut —
        top = ctk.CTkFrame(frame, fg_color="transparent")
        top.pack(fill="x")
        ctk.CTkButton(top, text="🔄  Rafraîchir", fg_color=C_NEUTRE,
                      hover_color=C_NEUTRE_SURV,
                      command=lambda: self._browse_load(force=True), text_color=T_SUR_NEUTRE).pack(side="left")
        self.browse_status = ctk.CTkLabel(top, text="(en attente de connexion…)", text_color=T_SECONDAIRE,
                                          font=ctk.CTkFont(size=11))
        self.browse_status.pack(side="left", padx=10)

        # — Ligne : filtres —
        # BARRE DE FILTRES SUR DEUX RANGÉES.
        # Sur une seule ligne, les huit contrôles réclamaient 1237 px alors que
        # la zone utile en fenêtre minimale (1000 px) n'en offre que 752 :
        # 485 px étaient purement TRONQUÉS à droite. Le tri et les filtres de
        # détection, placés en fin de ligne, étaient donc invisibles — présents
        # dans le code, inutilisables à l'écran.
        filt = ctk.CTkFrame(frame, fg_color="transparent")
        filt.pack(fill="x", pady=(8, 4))
        bloc_texte = bloc_filtre(filt, "Rechercher")
        bloc_texte.pack(side="left", fill="x", expand=True)
        self.browse_text = ctk.CTkEntry(bloc_texte,
                                        placeholder_text="🔍 titre / slug / propriétaire…")
        self.browse_text.pack(fill="x")
        self.browse_text.bind("<KeyRelease>", lambda e: self._browse_apply_filter())

        bloc_statut = bloc_filtre(filt, "Statut")
        bloc_statut.pack(side="left", padx=6)
        self.browse_statut = ctk.CTkOptionMenu(
            bloc_statut, width=130, values=["Tous", "Brouillon", "Public", "Restreinte"],
            command=lambda _c: self._browse_apply_filter(), **STYLE_CHAMP)
        self.browse_statut.set("Tous")
        self.browse_statut.pack()

        bloc_encode = bloc_filtre(filt, "Encodage")
        bloc_encode.pack(side="left", padx=6)
        self.browse_encode = ctk.CTkOptionMenu(
            bloc_encode, width=150, values=["Tout", "Encodées", "Non-encodées"],
            command=lambda _c: self._browse_apply_filter(), **STYLE_CHAMP)
        self.browse_encode.set("Tout")
        self.browse_encode.pack()
        bloc_chan = bloc_filtre(filt, "Chaîne")
        bloc_chan.pack(side="left", padx=6)
        self.browse_chan = ctk.CTkOptionMenu(bloc_chan, width=150, values=["Toutes"],
                                             command=lambda _c: self._browse_apply_filter(), **STYLE_CHAMP)
        self.browse_chan.set("Toutes")
        self.browse_chan.pack()

        bloc_type = bloc_filtre(filt, "Type")
        bloc_type.pack(side="left", padx=6)
        self.browse_type = ctk.CTkOptionMenu(bloc_type, width=132, values=["Tous"],
                                             command=lambda _c: self._browse_apply_filter(), **STYLE_CHAMP)
        self.browse_type.set("Tous")
        self.browse_type.pack()

        # ── Seconde rangée : tri et détection ────────────────────────────
        filt2 = ctk.CTkFrame(frame, fg_color="transparent")
        filt2.pack(fill="x", pady=(4, 6))


        # Ordre d'affichage. Par défaut « Plus récentes » : c'est l'ordre
        # renvoyé par l'API, celui auquel on est habitué. Le tri alphabétique
        # sert à retrouver une vidéo dont on connaît le titre.
        bloc_tri = bloc_filtre(filt2, "Trier")
        bloc_tri.pack(side="left", padx=(0, 14))
        self.browse_tri = ctk.CTkOptionMenu(
            bloc_tri, width=160,
            values=["Plus récentes", "A → Z", "Z → A"],
            command=lambda _c: self._browse_apply_filter(), **STYLE_CHAMP)
        self.browse_tri.set("Plus récentes")
        self.browse_tri.pack()

        # Filtres de DIAGNOSTIC, repris de l'Explorateur : ils ne cherchent pas
        # une vidéo précise mais révèlent des anomalies dans le fonds.
        bloc_detect = bloc_filtre(filt2, "Détecter")
        bloc_detect.pack(side="left", padx=(0, 6))
        self.browse_detect = ctk.CTkOptionMenu(
            bloc_detect, width=175,
            values=["Aucune", "Doublons de titre", "Vieux brouillons"],
            command=lambda _c: self._browse_apply_filter(), **STYLE_CHAMP)
        self.browse_detect.set("Aucune")
        self.browse_detect.pack()

        # Ancienneté retenue pour « Vieux brouillons ».
        ctk.CTkLabel(filt2, text="plus de", font=ctk.CTkFont(size=11),
                     text_color=T_SECONDAIRE).pack(side="left", padx=(4, 2))
        self.browse_months = ctk.CTkEntry(filt2, width=44)
        self.browse_months.insert(0, "6")
        self.browse_months.pack(side="left", padx=2)
        self.browse_months.bind("<KeyRelease>", lambda e: self._browse_apply_filter())
        ctk.CTkLabel(filt2, text="mois", font=ctk.CTkFont(size=11),
                     text_color=T_SECONDAIRE).pack(side="left")

        # — Action « en masse » : modifier le type des vidéos AFFICHÉES —
        # On la détache nettement des filtres ci-dessus (séparateur + cadre
        # encadré + libellé d'action) pour qu'on ne la confonde pas avec un filtre.
        ctk.CTkFrame(frame, height=1, fg_color=S_FILET).pack(fill="x", pady=(6, 0))
        massbar = ctk.CTkFrame(frame, fg_color=S_CARTE,
                               corner_radius=8, border_width=1, border_color=S_FILET)
        massbar.pack(fill="x", pady=(4, 4))
        ctk.CTkLabel(massbar, text="✏️  Modifier en masse — appliquer ce type aux vidéos affichées :",
                     font=ctk.CTkFont(size=11), text_color=T_SECONDAIRE
                     ).pack(side="left", padx=(10, 8), pady=6)
        self.browse_mass_type = ctk.CTkOptionMenu(massbar, width=170, values=["(aucun type)"], **STYLE_CHAMP)
        self.browse_mass_type.pack(side="left", pady=6)
        # LE COMPTE EST DANS LE BOUTON.
        #
        # « Appliquer » nu, à côté d'un libellé disant « aux vidéos affichées »
        # sans jamais dire combien, c'était le seul élément saturé de l'écran —
        # et l'action la plus lourde de conséquences. Porter la cardinalité
        # dans le bouton est la meilleure protection contre le clic de masse
        # par inadvertance : on ne peut plus cliquer sans avoir lu le nombre.
        #
        # Teinte d'ALERTE et non d'action : ce n'est pas l'opération courante
        # de l'écran, c'est une opération de masse irréversible.
        self.browse_mass_btn = ctk.CTkButton(
            massbar, text="Appliquer", width=170,
            fg_color=C_ALERTE, hover_color=C_ALERTE_SURV,
            command=self._browse_mass_set_type)
        self.browse_mass_btn.pack(side="left", padx=8, pady=6)

        # — Corps : liste (gauche) + détail (droite) —
        body = ctk.CTkFrame(frame, fg_color="transparent")
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=2)
        body.columnconfigure(1, weight=3)
        body.rowconfigure(1, weight=1)

        # Ligne d'en-tête : compteur à gauche, sélection globale à droite.
        entete = ctk.CTkFrame(body, fg_color="transparent")
        entete.grid(row=0, column=0, sticky="ew", pady=(0, 2))
        entete.columnconfigure(0, weight=1)

        self.browse_count_lbl = ctk.CTkLabel(entete, text="", text_color=T_SECONDAIRE,
                                             font=ctk.CTkFont(size=11), anchor="w")
        self.browse_count_lbl.grid(row=0, column=0, sticky="w")

        # « Tout sélectionner » porte sur TOUTES les vidéos filtrées, y compris
        # celles qui ne sont pas affichées : l'affichage est plafonné à 300
        # lignes pour rester fluide, mais la sélection ne l'est pas.
        ctk.CTkButton(entete, text="☑ Tout sélectionner", width=140, height=24,
                      font=ctk.CTkFont(size=11), fg_color=C_NEUTRE,
                      hover_color=C_NEUTRE_SURV,
                      command=self._browse_tout_selectionner, text_color=T_SUR_NEUTRE).grid(row=0, column=1, padx=4)
        self.browse_list = ctk.CTkScrollableFrame(body, label_text="Résultats", fg_color=S_CARTE, label_anchor="w",
                                                  label_font=ctk.CTkFont(size=12, weight="bold"))
        self.browse_list.grid(row=1, column=0, sticky="nsew", padx=(0, 6))
        self.browse_detail = ctk.CTkScrollableFrame(body, label_text="Détail / actions", fg_color=S_CARTE, label_anchor="w",
                                                  label_font=ctk.CTkFont(size=12, weight="bold"))
        self.browse_detail.grid(row=1, column=1, sticky="nsew", padx=(6, 0))

        # — Données —
        # `browse_videos` n'est plus un cache propre à l'onglet : c'est un ALIAS
        # du magasin partagé `self.videos` (même objet liste). Conservé le temps
        # de la migration pour ne rien casser si un code résiduel y accède.
        self.browse_videos = self.videos
        self.browse_channels = []       # chaînes (pour filtre + sélecteur)
        self.browse_chan_by_url = {}    # URL chaîne → titre
        self.browse_filtered = []       # sous-ensemble affiché
        self.browse_selected = None     # vidéo en cours d'édition
        # SÉLECTION MULTIPLE (Ctrl+clic / Maj+clic) : ensemble des slugs retenus.
        # Vide = mode normal, une seule vidéo affichée dans le panneau de détail.
        # Non vide = le panneau bascule sur les actions groupées.
        self.browse_multi: set = set()
        # Demande d'interruption d'un traitement par lot. Depuis « Tout
        # sélectionner », un lot peut porter sur plusieurs centaines de vidéos :
        # sans ce drapeau, le seul recours serait de tuer l'application, en
        # laissant le traitement à moitié fait sans savoir où il s'est arrêté.
        self.lot_interrompu = threading.Event()
        self.browse_ancre = None        # dernière ligne cliquée, pour Maj+clic

        self._browse_render_detail()    # affiche le message d'invite

    # ── Chargement (vidéos + chaînes) ──────────────────────────────────────

    def _browse_load(self, force: bool = False):
        """Charge (ou rafraîchit) la liste des vidéos de l'onglet Vidéos.

        `force=True` correspond au bouton « Rafraîchir » : on relit réellement
        le serveur. À la simple OUVERTURE de l'onglet, `force` reste False et le
        magasin partagé répond depuis le cache s'il est déjà chargé — c'est tout
        l'intérêt du cache partagé (aucun rechargement inutile)."""
        if not self.api:
            self.browse_status.configure(text="Connectez-vous d'abord.", text_color=T_ALERTE)
            return
        self.browse_status.configure(text="⏳  Chargement…", text_color=T_SECONDAIRE)
        self._browse_force_reload = bool(force)
        self._run(self._do_browse_load)

    def _do_browse_load(self):
        """(Thread) Récupère les chaînes, puis demande les vidéos au MAGASIN PARTAGÉ.

        Les vidéos ne sont plus chargées ici : elles proviennent de
        `ensure_videos()`, source unique partagée avec l'Explorateur et les
        Chaînes. Si un autre onglet les a déjà chargées, l'affichage est
        immédiat (aucun appel réseau)."""
        try:
            def prog(n):
                """Callback de progression (avancement du scan)."""
                self._ui(self.browse_status.configure,
                         text=f"⏳  {n} vidéos lues…", text_color=T_SECONDAIRE)
            try:
                channels = self.api.get_channels()
            except Exception:
                channels = []
            self.browse_channels = channels
            self.browse_chan_by_url = {str(c.get("url", "")).rstrip("/"): c.get("title", "?")
                                       for c in channels}
            self._ui(self._browse_refresh_channel_menu)
            # Demande au magasin partagé : `force` reflète le clic sur
            # « Rafraîchir » (voir _browse_load).
            self._ui(self.ensure_videos,
                     on_ready=self._browse_after_videos,
                     force=getattr(self, "_browse_force_reload", False),
                     progress_cb=prog)
        except Exception as e:
            self._ui(self.browse_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Chargement explorateur : {e}")

    def _browse_after_videos(self):
        """Appelé quand le magasin partagé est prêt : rafraîchit l'onglet Vidéos."""
        self._browse_force_reload = False      # le rechargement forcé est consommé
        # La sélection en cours pointe vers un objet de l'ANCIENNE liste après un
        # rechargement : on la ré-associe par slug, sinon le panneau de détail
        # afficherait un objet qui n'est plus dans le magasin.
        if self.browse_selected:
            slug = self.browse_selected.get("slug")
            self.browse_selected = next(
                (v for v in self.videos if v.get("slug") == slug), None)
        self._browse_apply_filter()
        alerte = self.scan_truncated_warning()
        self.browse_status.configure(
            text=(alerte if alerte else
                  f"✅  {len(self.videos)} vidéos, {len(self.browse_channels)} chaîne(s)  ·  "
                  f"chargé à {datetime.now().strftime('%H:%M')}"),
            text_color=T_ERREUR if alerte else "#22c55e")
        if alerte:
            self._log(alerte)
        self._log(f"Onglet Vidéos : {len(self.videos)} vidéos (magasin partagé).")

    def _browse_refresh_channel_menu(self):
        """Remplit le filtre par chaîne avec les titres chargés."""
        vals = ["Toutes"] + sorted(self.browse_chan_by_url.values(), key=str.lower)
        self.browse_chan.configure(values=vals)
        self.browse_chan.set("Toutes")
        self._browse_refresh_type_menu()

    def _browse_refresh_type_menu(self):
        """Remplit le filtre par type et le menu « en masse » avec les types chargés.
        Sans danger si appelé avant que les types soient chargés."""
        titles = sorted((self.type_map or {}).keys(), key=str.lower)
        if hasattr(self, "browse_type"):
            self.browse_type.configure(values=["Tous"] + titles)
            if self.browse_type.get() not in (["Tous"] + titles):
                self.browse_type.set("Tous")
        if hasattr(self, "browse_mass_type"):
            self.browse_mass_type.configure(values=titles or ["(aucun type)"])
            if titles and self.browse_mass_type.get() not in titles:
                self.browse_mass_type.set(titles[0])

    # ── Filtrage ───────────────────────────────────────────────────────────

    def _browse_owner_label(self, v) -> str:
        """Nom lisible du propriétaire d'une vidéo (via la liste des comptes)."""
        oid = str(self._video_owner_id(v)).rstrip("/")
        for u in (self.all_users or []):
            if str(u.get("url", "")).rstrip("/") == oid or u.get("username") == oid:
                return u.get("username", oid)
        return oid or "—"

    def _browse_apply_filter(self, *_):
        """Demande un filtrage de la liste — avec TEMPORISATION.

        Le filtre est branché sur chaque frappe clavier. Sans temporisation,
        taper « conference » reconstruisait dix fois la liste entière (une par
        touche), soit plusieurs secondes de blocage pour neuf rendus jetés
        aussitôt. On attend donc une courte pause dans la frappe avant de
        recalculer : une seule reconstruction au lieu de dix."""
        job = getattr(self, "_browse_filter_job", None)
        if job:
            try:
                self.after_cancel(job)      # annule le rendu encore en attente
            except Exception:
                pass
        self._browse_filter_job = self.after(FILTER_DELAY_MS, self._browse_do_filter)

    def _browse_do_filter(self):
        """Filtre réellement la liste des vidéos côté client (appelé après la
        temporisation de `_browse_apply_filter`)."""
        self._browse_filter_job = None
        if not self.videos:
            self.browse_count_lbl.configure(text="Cliquez sur « Charger les vidéos ».")
            for w in self.browse_list.winfo_children():
                w.destroy()
            return

        vids = self.videos              # source unique partagée
        # Filtre statut
        st = self.browse_statut.get()
        if st == "Brouillon":
            vids = [v for v in vids if v.get("is_draft")]
        elif st == "Public":
            vids = [v for v in vids if not v.get("is_draft")]
        elif st == "Restreinte":
            vids = [v for v in vids if v.get("is_restricted")]
        # Filtre encodage
        enc = self.browse_encode.get()
        if enc == "Encodées":
            vids = [v for v in vids if v.get("encoded")]
        elif enc == "Non-encodées":
            vids = [v for v in vids if PodAPI.is_unencoded(v)]
        # Filtre chaîne
        ch = self.browse_chan.get()
        if ch and ch != "Toutes":
            # On retrouve l'URL de la chaîne à partir de son titre
            wanted = [u for u, t in self.browse_chan_by_url.items() if t == ch]
            def in_chan(v):
                """Teste si une vidéo appartient à la chaîne filtrée."""
                cs = v.get("channel") or []
                if isinstance(cs, str):
                    cs = [cs]
                cs = [str(c).rstrip("/") for c in cs]
                return any(w in cs for w in wanted)
            vids = [v for v in vids if in_chan(v)]
        # Filtre type (valeur unique : on compare l'URL du type)
        ty = self.browse_type.get() if hasattr(self, "browse_type") else "Tous"
        if ty and ty != "Tous":
            turl = str((self.type_map or {}).get(ty, "")).rstrip("/")
            def has_type(v):
                """Teste si une vidéo est du type filtré."""
                vt = v.get("type")
                vt = vt.get("url") if isinstance(vt, dict) else vt
                return str(vt).rstrip("/") == turl
            vids = [v for v in vids if has_type(v)]
        # Filtre texte (titre / slug / propriétaire)
        txt = self.browse_text.get().strip().lower()
        if txt:
            def hay(v):
                """Concatène les champs d'une vidéo pour la recherche plein-texte."""
                return f"{v.get('title','')} {v.get('slug','')} {self._browse_owner_label(v)}".lower()
            vids = [v for v in vids if txt in hay(v)]

        # Détection d'anomalies (doublons, brouillons oubliés). Ces filtres
        # s'appliquent APRÈS les filtres ordinaires : on cherche les doublons
        # parmi les vidéos retenues, pas dans tout le fonds.
        detect = self.browse_detect.get() if hasattr(self, "browse_detect") else ""
        if detect == "Doublons de titre":
            vids = self._duplicate_title_videos(vids)
        elif detect == "Vieux brouillons":
            cutoff = self._months_ago_iso(self._browse_months_value())
            vids = [v for v in vids if PodAPI.is_stale_draft(v, cutoff)]

        # Tri demandé. Le tri alphabétique ignore la casse et les espaces de
        # début, sans quoi «  atelier » passerait avant « Anatomie ».
        tri = self.browse_tri.get() if hasattr(self, "browse_tri") else "Plus récentes"
        if tri in ("A → Z", "Z → A"):
            vids = sorted(vids,
                          key=lambda v: (v.get("title") or "").strip().lower(),
                          reverse=(tri == "Z → A"))
        self.browse_filtered = vids
        self._render_browse_list()

    def _render_browse_list(self):
        """Affiche la liste filtrée des vidéos (plafonnée pour ne pas figer l'interface)."""
        for w in self.browse_list.winfo_children():
            w.destroy()
        self.browse_count_lbl.configure(text=f"{len(self.browse_filtered)} vidéo(s) trouvée(s).")
        self._maj_bouton_masse()

        if not self.browse_filtered:
            ctk.CTkLabel(self.browse_list, text="Aucune vidéo ne correspond.",
                         text_color=T_SECONDAIRE).pack(pady=10)
            return

        CAP = 300
        sel_slug = self.browse_selected.get("slug") if self.browse_selected else None
        # On garde une référence sur chaque bouton : cela permet, lors d'un clic,
        # de ne recolorer QUE les deux lignes concernées au lieu de reconstruire
        # toute la liste (voir _browse_select).
        self.browse_rowbtns = {}
        police = ctk.CTkFont(size=12)      # police partagée par toutes les lignes
        for v in self.browse_filtered[:CAP]:
            slug = v.get("slug", "?")
            is_sel = slug == sel_slug
            title = (v.get("title") or "(sans titre)")[:48]
            tag = "📝" if v.get("is_draft") else "🌐"        # brouillon / public
            btn = ctk.CTkButton(
                self.browse_list, text=f"{tag}  {title}", anchor="w", height=28,
                fg_color=S_SELECTION if is_sel else "transparent",
                text_color=("gray10", "gray90"), hover_color=("gray75", "gray28"),
                font=police,
                command=lambda vv=v: self._browse_select(vv))
            btn.pack(fill="x", pady=1)
            # Sélection multiple à la manière d'un explorateur de fichiers :
            # Ctrl+clic ajoute ou retire une ligne, Maj+clic prend une plage.
            # `add="+"` conserve l'action normale du bouton (clic simple).
            btn.bind("<Control-Button-1>", lambda e, vv=v: self._browse_toggle_multi(vv))
            btn.bind("<Shift-Button-1>", lambda e, vv=v: self._browse_plage_multi(vv))
            self.browse_rowbtns[slug] = btn
            # Teinte particulière pour les lignes retenues en sélection multiple.
            if slug in self.browse_multi:
                btn.configure(fg_color=("#93c5fd", "#1e3a8a"))
        if len(self.browse_filtered) > CAP:
            ctk.CTkLabel(self.browse_list,
                         text=f"… +{len(self.browse_filtered) - CAP} autres. Affinez le filtre.",
                         text_color=T_SECONDAIRE).pack(pady=4)

    # ── Sélection multiple (Ctrl+clic / Maj+clic) ─────────────────────────

    def _browse_toggle_multi(self, v):
        """Ctrl+clic : ajoute ou retire une vidéo de la sélection multiple."""
        slug = v.get("slug")
        if slug in self.browse_multi:
            self.browse_multi.discard(slug)
        else:
            self.browse_multi.add(slug)
        self.browse_ancre = slug
        self._browse_refresh_multi()
        return "break"          # empêche le clic simple de réinitialiser

    def _browse_plage_multi(self, v):
        """Maj+clic : sélectionne toute la plage depuis la dernière ligne cliquée."""
        slugs = [x.get("slug") for x in self.browse_filtered[:300]]
        cible = v.get("slug")
        depart = self.browse_ancre or (
            self.browse_selected.get("slug") if self.browse_selected else None)
        if depart not in slugs or cible not in slugs:
            return self._browse_toggle_multi(v)
        i, j = sorted((slugs.index(depart), slugs.index(cible)))
        self.browse_multi.update(slugs[i:j + 1])
        self._browse_refresh_multi()
        return "break"

    def _browse_months_value(self) -> int:
        """Lit le champ « mois » de l'onglet Vidéos (entier ≥ 0, défaut 6)."""
        try:
            return max(0, int(self.browse_months.get().strip()))
        except Exception:
            return 6

    def _lot_interrompre(self):
        """Demande l'arrêt du traitement par lot en cours.

        L'arrêt est PROPRE : la vidéo en cours est menée à son terme, et le
        traitement s'arrête avant la suivante. On ne coupe jamais une opération
        au milieu, ce qui laisserait un état incohérent côté serveur."""
        self.lot_interrompu.set()
        try:
            self.browse_stop_btn.configure(state="disabled",
                                           text="⏳  Arrêt en cours…")
            self.browse_multi_msg.configure(
                text="Arrêt demandé : la vidéo en cours est terminée…",
                text_color=T_ALERTE)
        except Exception:
            pass

    def _lot_debut(self):
        """Active le bouton d'interruption au démarrage d'un lot."""
        try:
            self.browse_stop_btn.configure(state="normal",
                                           text="🛑  Interrompre le traitement")
        except Exception:
            pass

    def _lot_fin(self):
        """Désactive le bouton d'interruption à la fin d'un lot."""
        try:
            self.browse_stop_btn.configure(state="disabled",
                                           text="🛑  Interrompre le traitement")
        except Exception:
            pass

    def _browse_tout_selectionner(self):
        """Sélectionne TOUTES les vidéos filtrées, affichées ou non.

        L'affichage est plafonné à 300 lignes pour que l'interface reste
        fluide, mais ce plafond ne doit pas limiter les actions : sans ce
        bouton, il serait impossible de traiter plus de 300 vidéos d'un coup
        depuis cet onglet."""
        if not self.browse_filtered:
            return
        self.browse_multi = {v.get("slug") for v in self.browse_filtered}
        self.browse_ancre = None
        self._browse_refresh_multi()

    def _browse_multi_delete(self):
        """Supprime définitivement les vidéos sélectionnées.

        DOUBLE CONFIRMATION : un avertissement, puis la saisie du NOMBRE de
        vidéos concernées. Recopier un chiffre oblige à regarder combien on
        s'apprête à détruire — un simple « Oui » se clique sans lire.

        Pod n'a pas de corbeille : l'opération est irréversible.
        """
        videos = self._browse_videos_multi()
        n = len(videos)
        if not n:
            return

        if not messagebox.askyesno(
                "Suppression définitive",
                f"Supprimer définitivement {n} vidéo(s) ?\n\n"
                "Cette action est IRRÉVERSIBLE : Pod n'a pas de corbeille, les "
                "fichiers et leurs métadonnées seront perdus.\n\n"
                "Astuce : pour masquer des vidéos sans les perdre, utilisez "
                "plutôt « Mettre en brouillon »."):
            return

        saisie = simpledialog.askstring(
            "Confirmation",
            f"Pour confirmer, saisissez le nombre de vidéos à supprimer :\n\n"
            f"       {n}\n",
            parent=self)
        if saisie is None:
            return                                   # annulé
        if saisie.strip() != str(n):
            self.browse_multi_msg.configure(
                text="Suppression annulée : le nombre saisi ne correspond pas.",
                text_color=T_ALERTE)
            return

        self.browse_multi_msg.configure(text="⏳ Suppression en cours…",
                                        text_color=T_SECONDAIRE)
        self._lot_debut()
        self._run(self._do_browse_multi_delete, videos)

    def _do_browse_multi_delete(self, videos):
        """(Thread) Supprime les vidéos une par une, puis rafraîchit."""
        ok = fail = 0
        self.lot_interrompu.clear()
        interrompu = False
        for v in videos:
            if self.lot_interrompu.is_set():
                interrompu = True
                break
            slug = v.get("slug", "?")
            try:
                self.api.delete_video(v)
                with self._videos_lock:
                    if v in self.videos:
                        self.videos.remove(v)
                self._sync_video_caches(slug, removed=True)
                ok += 1
            except Exception as e:
                fail += 1
                self._ui(self._log, f"❌ Suppression {slug} : {e}")
            self._ui(self.browse_multi_msg.configure,
                     text=f"⏳ {ok + fail}/{len(videos)} traitée(s)…", text_color=T_SECONDAIRE)

        reste = len(videos) - ok - fail
        mention = f" — INTERROMPU, {reste} non supprimée(s)" if interrompu else ""
        self._ui(self._log,
                 f"🗑 Suppression multiple : {ok} supprimée(s), {fail} échec(s){mention}.")
        self._ui(self._lot_fin)
        # La sélection porte sur des vidéos qui n'existent plus : on la vide.
        self._ui(self._browse_vider_multi)
        self._ui(self._browse_set_msg,
                 f"🗑 {ok} vidéo(s) supprimée(s)" + (f", {fail} échec(s)." if fail else "."),
                 "#22c55e" if not fail else "#f59e0b")
        self._ui(self._refresh_video_views)

    def _browse_vider_multi(self):
        """Annule la sélection multiple et revient au détail d'une vidéo."""
        self.browse_multi.clear()
        self.browse_ancre = None
        self._browse_refresh_multi()

    def _browse_refresh_multi(self):
        """Met à jour la teinte des lignes puis rebascule le panneau de droite."""
        for slug, btn in getattr(self, "browse_rowbtns", {}).items():
            try:
                if not btn.winfo_exists():
                    continue
                if slug in self.browse_multi:
                    btn.configure(fg_color=("#93c5fd", "#1e3a8a"))
                elif self.browse_selected and slug == self.browse_selected.get("slug"):
                    btn.configure(fg_color=S_SELECTION)
                else:
                    btn.configure(fg_color="transparent")
            except Exception:
                pass
        self._browse_render_detail()

    def _browse_videos_multi(self) -> list:
        """Renvoie les objets vidéo correspondant à la sélection multiple.

        On parcourt TOUTES les vidéos filtrées, pas seulement les 300 affichées :
        « Tout sélectionner » peut en retenir davantage."""
        return [v for v in self.browse_filtered if v.get("slug") in self.browse_multi]

    def _browse_render_multi_panel(self):
        """Panneau d'actions groupées, affiché quand plusieurs vidéos sont retenues.

        Il REMPLACE le détail d'une vidéo : afficher les deux serait ambigu
        (sur laquelle porterait l'action ?). Un bouton permet de revenir au mode
        normal.

        Les actions réutilisent les mêmes méthodes d'API que l'Explorateur :
        aucune logique n'est dupliquée, seule la présentation change.
        """
        for w in self.browse_detail.winfo_children():
            w.destroy()
        videos = self._browse_videos_multi()
        n = len(videos)

        ctk.CTkLabel(self.browse_detail, text=f"☑  {n} vidéo(s) sélectionnée(s)",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(
            anchor="w", padx=6, pady=(6, 2))
        # Une sélection peut dépasser les lignes visibles : on le dit clairement,
        # pour que personne n'agisse sur plus de vidéos qu'il ne croit.
        caches = n - len(getattr(self, "browse_rowbtns", {}).keys() & self.browse_multi)
        if caches > 0:
            ctk.CTkLabel(self.browse_detail,
                         text=f"⚠  dont {caches} hors affichage (liste limitée à 300 lignes)",
                         font=ctk.CTkFont(size=11), text_color=T_ALERTE,
                         wraplength=360, justify="left").pack(anchor="w", padx=6)
        ctk.CTkLabel(self.browse_detail,
                     text="Ctrl+clic pour ajouter ou retirer une vidéo, "
                          "Maj+clic pour une plage.",
                     font=ctk.CTkFont(size=11), text_color=T_DISCRET,
                     wraplength=360, justify="left").pack(anchor="w", padx=6)

        # Aperçu des titres retenus, pour éviter d'agir à l'aveugle.
        apercu = ctk.CTkScrollableFrame(self.browse_detail, height=110,
                                        label_text="Vidéos concernées", fg_color=S_CARTE, label_anchor="w",
                                                  label_font=ctk.CTkFont(size=12, weight="bold"))
        apercu.pack(fill="x", padx=4, pady=8)
        for v in videos[:60]:
            ctk.CTkLabel(apercu, text=f"• {(v.get('title') or '(sans titre)')[:46]}",
                         font=ctk.CTkFont(size=11), anchor="w").pack(anchor="w")
        if n > 60:
            ctk.CTkLabel(apercu, text=f"… et {n - 60} autre(s)",
                         font=ctk.CTkFont(size=11), text_color=T_SECONDAIRE).pack(anchor="w")

        ctk.CTkLabel(self.browse_detail, text="Appliquer à toute la sélection",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(
            anchor="w", padx=6, pady=(6, 2))

        for libelle, action, couleur in (
                ("📝  Mettre en brouillon", "draft", "gray35"),
                ("🌐  Rendre public", "public", "#16a34a"),
                ("🔒  Rendre restreint", "restricted", "#b45309"),
                ("🔐  Restreindre au groupe…", "groups", "#7c3aed"),
                ("📺  Affecter à une chaîne…", "channels", "#2563eb")):
            ctk.CTkButton(self.browse_detail, text=libelle, anchor="w",
                          fg_color=couleur,
                          command=lambda a=action: self._browse_multi_action(a)).pack(
                fill="x", padx=6, pady=2)

        self.browse_multi_msg = ctk.CTkLabel(
            self.browse_detail, text="", font=ctk.CTkFont(size=11),
            wraplength=360, justify="left", anchor="w")
        self.browse_multi_msg.pack(fill="x", padx=6, pady=(6, 2))

        # Interruption d'un traitement en cours. Désactivé au repos : il ne
        # sert que pendant un lot, et un bouton toujours cliquable laisserait
        # croire qu'il fait quelque chose.
        self.browse_stop_btn = ctk.CTkButton(
            self.browse_detail, text="🛑  Interrompre le traitement",
            fg_color=C_ALERTE, hover_color=C_ALERTE_SURV, state="disabled",
            command=self._lot_interrompre)
        self.browse_stop_btn.pack(fill="x", padx=6, pady=(0, 2))

        # ── Zone sensible ────────────────────────────────────────────────
        # La suppression est SÉPARÉE des autres actions, et non alignée avec
        # elles : dans un onglet de travail quotidien, elle ne doit pas être à
        # portée de clic distrait. Elle est protégée par une double
        # confirmation dont la seconde exige de saisir le nombre de vidéos.
        ctk.CTkFrame(self.browse_detail, height=1,
                     fg_color=S_FILET).pack(fill="x", padx=6, pady=(10, 6))
        ctk.CTkLabel(self.browse_detail, text="Zone sensible",
                     font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=T_ERREUR).pack(anchor="w", padx=6)
        ctk.CTkButton(self.browse_detail,
                      text=f"🗑  Supprimer définitivement ces {n} vidéo(s)",
                      anchor="w", fg_color=C_DESTRUCTIF, hover_color=C_DESTR_SURV,
                      command=self._browse_multi_delete).pack(fill="x", padx=6, pady=4)

        ctk.CTkButton(self.browse_detail, text="✖  Annuler la sélection",
                      fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=self._browse_vider_multi, text_color=T_SUR_NEUTRE).pack(fill="x", padx=6, pady=(8, 6))

    def _browse_select(self, v):
        """Sélectionne une vidéo et affiche son panneau de détail.

        La surbrillance est déplacée en recolorant UNIQUEMENT l'ancienne et la
        nouvelle ligne. Auparavant, la liste entière était détruite puis
        recréée à chaque clic (jusqu'à 300 boutons) juste pour ce changement de
        couleur, ce qui rendait la sélection très lente sur les grandes listes."""
        ancien = self.browse_selected.get("slug") if self.browse_selected else None
        self.browse_selected = v
        nouveau = v.get("slug")
        boutons = getattr(self, "browse_rowbtns", {})
        if boutons:
            b = boutons.get(ancien)
            if b is not None and b.winfo_exists():
                b.configure(fg_color="transparent")            # désélection
            b = boutons.get(nouveau)
            if b is not None and b.winfo_exists():
                b.configure(fg_color=S_SELECTION)     # sélection
        else:
            self._render_browse_list()   # repli : liste pas encore construite
        self._browse_render_detail()

    # ── Panneau de détail / actions ────────────────────────────────────────

    def _browse_multi_action(self, action: str):
        """Applique une action à toutes les vidéos sélectionnées.

        Les groupes et les chaînes passent par les MÊMES fenêtres de choix que
        l'Explorateur : un seul comportement à connaître, un seul à maintenir.
        """
        videos = self._browse_videos_multi()
        if not videos:
            return

        payload, groupes, chaines = None, None, None

        if action == "draft":
            payload = {"is_draft": True, "is_restricted": False}
            libelle = "mise en brouillon"
        elif action == "public":
            payload = {"is_draft": False, "is_restricted": False}
            libelle = "passage en public"
        elif action == "restricted":
            payload = {"is_draft": False, "is_restricted": True}
            libelle = "passage en restreint"
        elif action == "groups":
            if not self.access_groups:
                self.browse_multi_msg.configure(
                    text="Aucun groupe d'accès chargé (voir l'onglet Groupes d'accès).",
                    text_color=T_ALERTE)
                return
            groupes = self._pick_groups()      # fenêtre partagée
            if groupes is None:
                return                                # annulé
            libelle = f"restriction à {len(groupes)} groupe(s)"
        else:                                         # channels
            if not self.browse_channels:
                self.browse_multi_msg.configure(
                    text="Aucune chaîne chargée. Cliquez sur « Rafraîchir ».",
                    text_color=T_ALERTE)
                return
            # Le sélecteur de chaînes rend la main par CALLBACK : la suite du
            # traitement se poursuit donc dans _browse_multi_channels.
            ChannelPicker(self, self.browse_channels,
                          on_done=lambda urls, labels: self._browse_multi_channels(
                              videos, list(urls)),
                          title=f"Chaînes pour {len(videos)} vidéo(s)")
            return

        if not messagebox.askyesno(
                "Confirmer l'action",
                f"Appliquer « {libelle} » à {len(videos)} vidéo(s) ?"):
            return
        self.browse_multi_msg.configure(text="⏳ Application en cours…",
                                        text_color=T_SECONDAIRE)
        self._lot_debut()
        self._run(self._do_browse_multi_action, videos, payload, groupes,
                  chaines, libelle)

    def _demander_mode_chaines(self, nb_videos: int, nb_chaines: int):
        """Demande s'il faut AJOUTER aux chaînes existantes ou les REMPLACER.

        Une vidéo peut légitimement appartenir à plusieurs chaînes : remplacer
        systématiquement ferait perdre des affectations sans prévenir. On laisse
        donc le choix, en proposant l'ajout par défaut — l'option la moins
        destructrice.

        Renvoie "ajouter", "remplacer", ou None si l'utilisateur annule.
        """
        fen = ctk.CTkToplevel(self)
        fen.title("Affectation aux chaînes")
        fen.geometry("470x300")
        fen.resizable(False, False)
        _focus_toplevel(fen, self)
        choix = {"mode": None}

        ctk.CTkLabel(fen, text="Comment appliquer ces chaînes ?",
                     font=ctk.CTkFont(size=15, weight="bold")).pack(
            anchor="w", padx=20, pady=(18, 4))
        ctk.CTkLabel(fen,
                     text=f"{nb_chaines} chaîne(s) choisie(s) pour {nb_videos} vidéo(s).",
                     text_color=T_SECONDAIRE, font=ctk.CTkFont(size=12)).pack(
            anchor="w", padx=20, pady=(0, 12))

        def retenir(mode):
            """Mémorise le mode choisi et referme la fenêtre."""
            choix["mode"] = mode
            fen.destroy()

        ctk.CTkButton(fen, text="➕  Ajouter aux chaînes existantes", height=40,
                      fg_color=C_SUCCES, hover_color=C_SUCCES_SURV,
                      command=lambda: retenir("ajouter")).pack(fill="x", padx=20, pady=4)
        ctk.CTkLabel(fen,
                     text="Les vidéos restent dans leurs chaînes actuelles ; "
                          "les nouvelles s'y ajoutent.",
                     font=ctk.CTkFont(size=11), text_color=T_DISCRET,
                     wraplength=420, justify="left").pack(anchor="w", padx=24)

        ctk.CTkButton(fen, text="🔄  Remplacer les chaînes actuelles", height=40,
                      fg_color=C_ALERTE, hover_color=C_ALERTE_SURV,
                      command=lambda: retenir("remplacer")).pack(fill="x", padx=20,
                                                                 pady=(12, 4))
        ctk.CTkLabel(fen,
                     text="Les affectations existantes seront PERDUES et remplacées "
                          "par la nouvelle sélection.",
                     font=ctk.CTkFont(size=11), text_color=T_DISCRET,
                     wraplength=420, justify="left").pack(anchor="w", padx=24)

        ctk.CTkButton(fen, text="Annuler", width=110, fg_color=C_NEUTRE,
                      hover_color=C_NEUTRE_SURV, command=fen.destroy, text_color=T_SUR_NEUTRE).pack(pady=(14, 10))

        self.wait_window(fen)
        return choix["mode"]

    def _browse_multi_channels(self, videos, chaines):
        """Suite du traitement après le choix des chaînes (appelé en callback)."""
        mode = self._demander_mode_chaines(len(videos), len(chaines))
        if mode is None:
            return                                   # annulé
        libelle = ("ajout à" if mode == "ajouter" else "remplacement par")
        self.browse_multi_msg.configure(text="⏳ Application en cours…",
                                        text_color=T_SECONDAIRE)
        self._run(self._do_browse_multi_action, videos, None, None,
                  (chaines, mode), f"{libelle} {len(chaines)} chaîne(s)")

    def _do_browse_multi_action(self, videos, payload, groupes, chaines, libelle):
        """(Thread) Applique l'action à chaque vidéo, une par une.

        Traitement séquentiel : plus lent que des envois simultanés, mais
        prévisible, et une erreur reste facile à situer."""
        ok = fail = 0
        self.lot_interrompu.clear()
        interrompu = False
        for v in videos:
            # On s'arrête ENTRE deux vidéos, jamais au milieu d'une opération :
            # la vidéo en cours est toujours menée à son terme, pour ne pas
            # laisser un état incohérent côté serveur.
            if self.lot_interrompu.is_set():
                interrompu = True
                break
            slug = v.get("slug", "?")
            try:
                if payload is not None:
                    self.api.patch_video(v, payload)
                    self._sync_video_caches(slug, payload)
                elif groupes is not None:
                    self.api.set_video_groups(v, groupes)
                    maj = {"restrict_access_to_groups": list(groupes),
                           "is_restricted": bool(groupes), "is_draft": False}
                    self._sync_video_caches(slug, maj)
                else:
                    # `chaines` est un couple (liste d'URLs, mode) : « ajouter »
                    # conserve les chaînes actuelles de CHAQUE vidéo — elles
                    # diffèrent d'une vidéo à l'autre — tandis que « remplacer »
                    # impose la même liste à toutes.
                    urls, mode = chaines
                    if mode == "ajouter":
                        # `_rel_urls` : le champ `channel` peut contenir des URLs
                        # ou des objets imbriqués selon le sérialiseur.
                        actuelles = self._rel_urls(v.get("channel"), normalise=False)
                        finales = list(dict.fromkeys(list(actuelles) + list(urls)))
                    else:
                        finales = list(urls)
                    self.api.assign_video_to_channels(v, finales)
                    self._sync_video_caches(slug, {"channel": finales})
                ok += 1
            except Exception as e:
                fail += 1
                self._ui(self._log, f"❌ {slug} : {e}")
            self._ui(self.browse_multi_msg.configure,
                     text=f"⏳ {ok + fail}/{len(videos)} traitée(s)…",
                     text_color=T_SECONDAIRE)

        reste = len(videos) - ok - fail
        mention = f" — INTERROMPU, {reste} vidéo(s) non traitée(s)" if interrompu else ""
        self._ui(self._log,
                 f"Sélection multiple — {libelle} : {ok} OK, {fail} échec(s){mention}.")
        self._ui(self.browse_multi_msg.configure,
                 text=(f"⏹ Interrompu : {ok} traitée(s), {reste} non traitée(s)."
                       if interrompu
                       else f"✅ Terminé : {ok} réussie(s), {fail} échec(s)."),
                 text_color=T_ALERTE if (interrompu or fail) else "#22c55e")
        self._ui(self._lot_fin)
        # La sélection est conservée : on peut enchaîner une autre action sur
        # les mêmes vidéos (restreindre puis affecter à une chaîne, par exemple).
        self._ui(self._refresh_video_views)

    def _browse_render_detail(self):
        """Affiche le panneau de détail de la vidéo sélectionnée."""
        # Si des vidéos sont retenues en sélection multiple, on affiche les
        # actions groupées à la place du détail : agir sur « la » vidéo alors
        # que plusieurs sont cochées serait ambigu.
        if self.browse_multi:
            self._browse_render_multi_panel()
            return
        for w in self.browse_detail.winfo_children():
            w.destroy()
        v = self.browse_selected
        if not v:
            ctk.CTkLabel(self.browse_detail,
                         text="Sélectionnez une vidéo dans la liste pour l'éditer.",
                         text_color=T_SECONDAIRE).pack(pady=14)
            return

        slug = v.get("slug", "?")

        # — Informations —
        ctk.CTkLabel(self.browse_detail, text=v.get("title", "(sans titre)"),
                     font=ctk.CTkFont(size=15, weight="bold"),
                     wraplength=420, justify="left").pack(anchor="w", padx=4, pady=(4, 0))
        chans = v.get("channel") or []
        if isinstance(chans, str):
            chans = [chans]
        chan_names = ", ".join(self.browse_chan_by_url.get(str(c).rstrip("/"), "?")
                               for c in chans) or "(aucune)"
        info = (f"slug : {slug}\n"
                f"propriétaire : {self._browse_owner_label(v)}\n"
                f"durée : {v.get('duration_in_time', '—')}     "
                f"encodée : {'oui' if v.get('encoded') else 'non'}\n"
                f"chaînes : {chan_names}")
        ctk.CTkLabel(self.browse_detail, text=info, justify="left", anchor="w",
                     text_color=T_SECONDAIRE, font=ctk.CTkFont(size=12)).pack(
            anchor="w", padx=4, pady=(2, 4))

        # Prévisualisation : PodAdmin (Tkinter) n'a pas de lecteur vidéo intégré,
        # on délègue la lecture au navigateur (page publique Pod de la vidéo).
        ctk.CTkButton(self.browse_detail, text="▶  Ouvrir dans le navigateur",
                      width=210, height=28, fg_color="#0f766e", hover_color="#115e59",
                      command=lambda s=slug: self._open_video_in_browser(s)).pack(
            anchor="w", padx=4, pady=(0, 10))

        # — Renommer —
        ren = ctk.CTkFrame(self.browse_detail, fg_color="transparent")
        ren.pack(fill="x", padx=4)
        self.browse_title_entry = ctk.CTkEntry(ren)
        self.browse_title_entry.insert(0, v.get("title", ""))
        self.browse_title_entry.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(ren, text="Renommer", width=90, fg_color=C_NEUTRE,
                      command=lambda: self._browse_rename(v), text_color=T_SUR_NEUTRE).pack(side="left", padx=6)

        # — Statut (interrupteurs) —
        # Important : on met à jour le cache local AVANT de lancer le thread,
        # pour que _browse_render_detail() reconstruise le panneau avec la
        # bonne valeur même si la réponse réseau tarde.
        ctk.CTkLabel(self.browse_detail, text="Statut", anchor="w",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=4, pady=(10, 2))
        # Les trois statuts sont EXCLUSIFS (et non des cases indépendantes) :
        #   Brouillon  = is_draft True
        #   Public     = is_draft False ET is_restricted False
        #   Restreint  = is_draft False ET is_restricted True
        # Un bouton segmenté garantit qu'on ne peut en choisir qu'un, et chaque
        # choix envoie les DEUX booléens cohérents d'un coup (corrige le bug où
        # l'ancien statut restait actif).
        def _status_of(vid):
            """Renvoie le libellé de statut d'une vidéo (Brouillon/Public/Restreint)."""
            if vid.get("is_draft"):
                return "Brouillon"
            return "Restreint" if vid.get("is_restricted") else "Public"
        status_seg = ctk.CTkSegmentedButton(
            self.browse_detail, values=["Brouillon", "Public", "Restreint"])
        status_seg.set(_status_of(v))
        status_seg.pack(fill="x", padx=4, pady=(0, 2))
        ctk.CTkLabel(self.browse_detail,
                     text="Restreint = visible mais connexion requise.",
                     font=ctk.CTkFont(size=10), text_color=T_DISCRET).pack(anchor="w", padx=6)
        def _apply_status(choice):
            # Calcule les deux booléens à partir du statut choisi
            payload = {"Brouillon": {"is_draft": True, "is_restricted": False},
                       "Public":    {"is_draft": False, "is_restricted": False},
                       "Restreint": {"is_draft": False, "is_restricted": True}}[choice]
            v.update(payload)                       # MAJ cache local immédiate
            self._browse_patch(v, payload, f"statut → {choice.lower()}")
        status_seg.configure(command=_apply_status)

        # — Vidéo 360° (panoramique / immersive) —
        # Case à cocher indépendante du statut. Le champ API est centralisé dans
        # la constante FIELD_360 (voir en tête de fichier) : si une instance
        # nommait ce champ autrement, un seul point est à changer.
        # Diagnostic du nom exact : lancer verifier_champ_360.py.
        ctk.CTkLabel(self.browse_detail, text="Format", anchor="w",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=4, pady=(12, 2))
        is360_var = ctk.BooleanVar(value=bool(v.get(FIELD_360)))

        def _apply_360():
            """Active/désactive le format 360° pour la vidéo affichée."""
            val = bool(is360_var.get())
            v[FIELD_360] = val                       # MAJ cache local immédiate
            self._browse_patch(v, {FIELD_360: val},
                               f"vidéo 360 → {'oui' if val else 'non'}")

        ctk.CTkCheckBox(self.browse_detail, text="Vidéo 360° (panoramique / immersive)",
                        variable=is360_var, command=_apply_360,
                        font=ctk.CTkFont(size=12)).pack(anchor="w", padx=6, pady=(0, 2))
        ctk.CTkLabel(self.browse_detail,
                     text="À cocher pour une vidéo filmée à 360°, afin que Pod utilise le "
                          "lecteur immersif.",
                     font=ctk.CTkFont(size=10), text_color=T_DISCRET,
                     wraplength=360, justify="left").pack(anchor="w", padx=6)

        # — Restreindre à des groupes d'accès —
        # Cocher au moins un groupe force le statut « Restreint » (couplage voulu).
        # Tout décocher retire la restriction par groupe (et repasse en public).
        ctk.CTkLabel(self.browse_detail, text="Restreindre à des groupes", anchor="w",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=4, pady=(12, 2))
        if not self.access_groups:
            ctk.CTkLabel(self.browse_detail,
                         text="(aucun groupe d'accès chargé)",
                         font=ctk.CTkFont(size=11), text_color=T_DISCRET).pack(anchor="w", padx=6)
        else:
            # URLs des groupes actuellement appliqués à la vidéo (normalisées)
            cur = v.get("restrict_access_to_groups") or []
            if isinstance(cur, str):
                cur = [cur]
            cur_urls = {str(g.get("url") if isinstance(g, dict) else g).rstrip("/")
                        for g in cur}
            grp_box = ctk.CTkFrame(self.browse_detail, fg_color="transparent")
            grp_box.pack(fill="x", padx=4)
            grp_vars = {}    # url → BooleanVar
            for g in self.access_groups:
                gurl = g.get("url", "")
                var = ctk.BooleanVar(value=str(gurl).rstrip("/") in cur_urls)
                grp_vars[gurl] = var
                ctk.CTkCheckBox(grp_box, text=g.get("code_name", "?"), variable=var,
                                font=ctk.CTkFont(size=11)).pack(anchor="w", pady=1)

            def _apply_groups():
                # Liste des URLs cochées
                urls = [u for u, var in grp_vars.items() if var.get()]
                v["restrict_access_to_groups"] = urls
                # Couplage statut : restreint si au moins un groupe, sinon public
                v["is_restricted"] = bool(urls)
                payload = {"restrict_access_to_groups": urls, "is_restricted": bool(urls)}
                if urls:
                    v["is_draft"] = False
                    payload["is_draft"] = False     # publier (sinon brouillon = invisible)
                    status_seg.set("Restreint")
                self._browse_patch(v, payload, f"groupes → {len(urls)} groupe(s)")
            ctk.CTkButton(self.browse_detail, text="Appliquer les groupes", width=180,
                          height=26, fg_color=C_ACTION, hover_color=C_ACTION_SURV, command=_apply_groups).pack(anchor="w", padx=4, pady=(4, 0))


        ctk.CTkLabel(self.browse_detail, text="Type", anchor="w",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=4, pady=(12, 2))
        # Titre du type courant de la vidéo (résolu depuis son URL)
        cur_url = v.get("type")
        cur_url = cur_url.get("url") if isinstance(cur_url, dict) else cur_url
        url_to_title = {str(u).rstrip("/"): t for t, u in (self.type_map or {}).items()}
        cur_title = url_to_title.get(str(cur_url).rstrip("/"), "(non défini)")
        titles = sorted((self.type_map or {}).keys(), key=str.lower) or ["(aucun type)"]
        type_menu = ctk.CTkOptionMenu(self.browse_detail, width=220, values=titles, **STYLE_CHAMP)
        type_menu.set(cur_title if cur_title in titles else titles[0])
        type_menu.pack(anchor="w", padx=4)
        def _apply_type(choice):
            # Change le type de la vidéo (PATCH valeur unique = URL du type)
            new_url = (self.type_map or {}).get(choice)
            if new_url and str(new_url).rstrip("/") != str(cur_url).rstrip("/"):
                v["type"] = new_url
                self._browse_patch(v, {"type": new_url}, f"type → {choice}")
        type_menu.configure(command=_apply_type)

        # — Co-propriétaires & chaînes —
        ctk.CTkLabel(self.browse_detail, text="Relations", anchor="w",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=4, pady=(12, 2))
        rel = ctk.CTkFrame(self.browse_detail, fg_color="transparent")
        rel.pack(fill="x", padx=4)
        ctk.CTkButton(rel, text="👥  Co-propriétaires…", fg_color=C_NEUTRE,
                      command=lambda: self._browse_edit_owners(v), text_color=T_SUR_NEUTRE).pack(side="left", padx=(0, 6))
        ctk.CTkButton(rel, text="🗂  Chaînes…", fg_color=C_NEUTRE,
                      command=lambda: self._browse_edit_channels(v), text_color=T_SUR_NEUTRE).pack(side="left")

        # — Sous-titres —
        ctk.CTkLabel(self.browse_detail, text="Sous-titres", anchor="w",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=4, pady=(12, 2))
        # Conteneur listant les pistes existantes (rempli en arrière-plan)
        self.browse_subs = ctk.CTkFrame(self.browse_detail, fg_color="transparent")
        self.browse_subs.pack(fill="x", padx=4)
        ctk.CTkLabel(self.browse_subs, text="Chargement…", text_color=T_SECONDAIRE,
                     font=ctk.CTkFont(size=11)).pack(anchor="w")
        # Bouton d'ajout d'un fichier .vtt / .srt
        ctk.CTkButton(self.browse_detail, text="➕  Ajouter un sous-titre (.vtt / .srt)",
                      fg_color=C_NEUTRE,
                      command=lambda: self._sub_add_dialog(v), text_color=T_SUR_NEUTRE).pack(anchor="w", padx=4, pady=(6, 0))
        # Chargement des pistes de cette vidéo en arrière-plan
        self._run(self._sub_load, v)

        # — Fichier source (remplacer + ré-encoder) —
        ctk.CTkLabel(self.browse_detail, text="Fichier source", anchor="w",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=4, pady=(14, 2))
        ctk.CTkLabel(self.browse_detail,
                     text="Remplace le fichier vidéo par un nouveau puis relance l'encodage. "
                          "La vidéo garde son titre, ses chaînes, ses droits… seul le média change.",
                     text_color=T_SECONDAIRE, font=ctk.CTkFont(size=11),
                     justify="left", wraplength=360).pack(anchor="w", padx=4)
        ctk.CTkButton(self.browse_detail, text="🎬  Remplacer le fichier & ré-encoder",
                      fg_color=C_ALERTE, hover_color=C_ALERTE_SURV,
                      command=lambda: self._browse_replace_source(v)).pack(anchor="w", padx=4, pady=(4, 0))

        # — Suppression —
        ctk.CTkLabel(self.browse_detail, text="Zone sensible", anchor="w",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=T_ERREUR).pack(anchor="w", padx=4, pady=(14, 2))
        ctk.CTkButton(self.browse_detail, text="🗑  Supprimer cette vidéo",
                      fg_color=C_DESTRUCTIF, hover_color=C_DESTR_SURV,
                      command=lambda: self._browse_delete(v)).pack(anchor="w", padx=4, pady=(0, 8))

        # Zone de message du panneau
        self.browse_msg = ctk.CTkLabel(self.browse_detail, text="", text_color=T_SECONDAIRE,
                                       font=ctk.CTkFont(size=11), wraplength=420, justify="left")
        self.browse_msg.pack(anchor="w", padx=4, pady=(4, 8))

    # ── Actions sur la vidéo sélectionnée ──────────────────────────────────

    # ── Sous-titres (tracks) ───────────────────────────────────────────────

    def _sub_load(self, v):
        """(Thread) Charge les pistes de sous-titres de la vidéo puis les affiche."""
        try:
            tracks = self.api.get_tracks(v)
            self._ui(self._sub_render, v, tracks)
        except Exception as e:
            self._ui(self._sub_render, v, None, str(e))

    def _sub_render(self, v, tracks, err=None):
        """Affiche la liste des pistes existantes (langue · type · 🗑)."""
        # Le panneau a pu être reconstruit entre-temps : on vérifie qu'il existe
        if not hasattr(self, "browse_subs") or not self.browse_subs.winfo_exists():
            return
        for w in self.browse_subs.winfo_children():
            w.destroy()
        if err:
            ctk.CTkLabel(self.browse_subs, text=f"❌ {err}", text_color=T_ERREUR,
                         font=ctk.CTkFont(size=11)).pack(anchor="w")
            return
        if not tracks:
            ctk.CTkLabel(self.browse_subs, text="Aucun sous-titre.", text_color=T_SECONDAIRE,
                         font=ctk.CTkFont(size=11)).pack(anchor="w")
            return
        # Dictionnaires code→libellé pour un affichage lisible
        langs = dict(SUBTITLE_LANGS)
        kinds = dict(SUBTITLE_KINDS)
        for t in tracks:
            row = ctk.CTkFrame(self.browse_subs, fg_color=S_LIGNE,
                               corner_radius=6)
            row.pack(fill="x", pady=2)
            lang = langs.get(t.get("lang"), t.get("lang"))
            kind = kinds.get(t.get("kind"), t.get("kind"))
            ctk.CTkLabel(row, text=f"{lang} · {kind}", anchor="w",
                         font=ctk.CTkFont(size=12)).pack(side="left", padx=10, pady=5,
                                                         fill="x", expand=True)
            # Déjà une icône seule, mais rouge en permanence : même motif
            # répété en liste que les chaînes et les thèmes.
            btn_s = ctk.CTkButton(row, text="🗑", width=34,
                                  fg_color=C_NEUTRE, hover_color=C_DESTRUCTIF,
                                  text_color=T_SUR_NEUTRE,
                                  command=lambda tt=t: self._sub_delete(v, tt))
            btn_s.pack(side="right", padx=6)
            ajouter_infobulle(btn_s, "Supprimer ce sous-titre")

    def _sub_add_dialog(self, v):
        """Fenêtre d'ajout : choix langue + type + fichier .vtt/.srt."""
        if not self.api:
            return
        win = ctk.CTkToplevel(self)
        win.title("Ajouter un sous-titre")
        win.geometry("440x300")
        _focus_toplevel(win, self)

        ctk.CTkLabel(win, text=f"Vidéo : {(v.get('title') or '')[:50]}",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(padx=16, pady=(16, 10), anchor="w")

        # Menu Langue
        ctk.CTkLabel(win, text="Langue :").pack(padx=16, anchor="w")
        lang_labels = [f"{lbl} ({code})" for code, lbl in SUBTITLE_LANGS]
        lang_menu = ctk.CTkOptionMenu(win, values=lang_labels, width=260, **STYLE_CHAMP)
        lang_menu.set("Français (fr)")
        lang_menu.pack(padx=16, pady=(0, 8), anchor="w")

        # Menu Type
        ctk.CTkLabel(win, text="Type :").pack(padx=16, anchor="w")
        kind_menu = ctk.CTkOptionMenu(
            win, values=[lbl for _c, lbl in SUBTITLE_KINDS], width=260, **STYLE_CHAMP)
        kind_menu.set("Sous-titres")
        kind_menu.pack(padx=16, pady=(0, 8), anchor="w")

        # Sélection du fichier
        path_var = {"p": None}
        path_lbl = ctk.CTkLabel(win, text="Aucun fichier choisi.", text_color=T_SECONDAIRE,
                                font=ctk.CTkFont(size=11), wraplength=400, justify="left")

        def choose():
            # Boîte de sélection limitée aux formats acceptés
            p = filedialog.askopenfilename(
                title="Choisir un fichier de sous-titres",
                filetypes=[("Sous-titres", "*.vtt *.srt"), ("Tous", "*.*")])
            if p:
                path_var["p"] = p
                path_lbl.configure(text=os.path.basename(p), text_color="#ffffff")

        ctk.CTkButton(win, text="📄  Choisir un fichier .vtt / .srt",
                      command=choose, fg_color=C_NEUTRE, text_color=T_SUR_NEUTRE).pack(padx=16, pady=(4, 2), anchor="w")
        path_lbl.pack(padx=16, anchor="w")

        def valider():
            # Résolution des codes à partir des libellés choisis
            lang_code = SUBTITLE_LANGS[lang_labels.index(lang_menu.get())][0]
            kind_code = next(c for c, lbl in SUBTITLE_KINDS
                             if lbl == kind_menu.get())
            path = path_var["p"]
            if not path:
                path_lbl.configure(text="⚠️ Choisissez d'abord un fichier.",
                                   text_color=T_ALERTE)
                return
            # Garde-fou d'extension (la conversion gère .srt, sinon .vtt attendu)
            if not path.lower().endswith((".vtt", ".srt")):
                path_lbl.configure(text="⚠️ Le fichier doit être .vtt ou .srt.",
                                   text_color=T_ALERTE)
                return
            win.destroy()
            self._run(self._sub_do_add, v, lang_code, kind_code, path)

        ctk.CTkButton(win, text="Ajouter", fg_color=C_SUCCES, hover_color=C_SUCCES_SURV,
                      command=valider).pack(pady=14)

    def _sub_do_add(self, v, lang, kind, path):
        """(Thread) Téléverse le fichier et crée la piste, puis rafraîchit la liste."""
        try:
            self.api.add_subtitle(v, lang, kind, path)   # conversion .srt incluse
            self._ui(self._log, f"➕ Sous-titre ajouté ({lang}/{kind}) à {v.get('slug')}")
            self._ui(self._browse_set_msg, "✅  Sous-titre ajouté.", "#22c55e")
            self._run(self._sub_load, v)                 # recharge la liste
        except Exception as e:
            self._ui(self._log, f"❌ Ajout sous-titre {v.get('slug')} : {e}")
            self._ui(self._browse_set_msg, f"❌  {e}", "#ef4444")

    def _sub_delete(self, v, track):
        """Supprime une piste après confirmation."""
        langs = dict(SUBTITLE_LANGS)
        lib = langs.get(track.get("lang"), track.get("lang"))
        if not messagebox.askyesno(
                "Supprimer le sous-titre",
                f"Supprimer la piste « {lib} » de cette vidéo ?"):
            return
        self._run(self._sub_do_delete, v, track)

    def _sub_do_delete(self, v, track):
        """(Thread) DELETE de la piste, puis rafraîchit la liste."""
        try:
            self.api.delete_track(track)
            self._ui(self._log, f"🗑 Sous-titre supprimé ({track.get('lang')}) de {v.get('slug')}")
            self._run(self._sub_load, v)
        except Exception as e:
            self._ui(self._log, f"❌ Suppression sous-titre : {e}")
            self._ui(self._browse_set_msg, f"❌  {e}", "#ef4444")

    def _maj_bouton_masse(self):
        """Inscrit le nombre de vidéos concernées dans le bouton de masse.

        Appelée à chaque rendu de la liste filtrée : le compte doit suivre le
        filtre, sinon il devient un mensonge — pire qu'une absence de compte.

        À zéro vidéo le bouton est DÉSACTIVÉ plutôt qu'affiché avec « 0 » : il
        n'y a rien à appliquer, et un bouton actif qui ne fait rien laisse
        croire à un échec."""
        bouton = getattr(self, "browse_mass_btn", None)
        if bouton is None:
            return
        n = len(getattr(self, "browse_filtered", []) or [])
        if n == 0:
            bouton.configure(text="Appliquer", state="disabled")
        elif n == 1:
            bouton.configure(text="Appliquer à 1 vidéo", state="normal")
        else:
            bouton.configure(text=f"Appliquer aux {n} vidéos", state="normal")

    def _browse_mass_set_type(self):
        """Affecte le type choisi à TOUTES les vidéos actuellement affichées
        (résultat du filtre courant). Double confirmation, puis exécution."""
        choice = self.browse_mass_type.get()
        new_url = (self.type_map or {}).get(choice)
        vids = list(self.browse_filtered)
        if not new_url or not vids:
            self._browse_set_msg("Rien à appliquer (aucun type ou aucune vidéo affichée).",
                                 "#f59e0b")
            return
        if not messagebox.askyesno(
                "Type en masse",
                f"Affecter le type « {choice} » à {len(vids)} vidéo(s) affichée(s) ?\n\n"
                "Cette action écrase le type actuel de chacune."):
            return
        self._run(self._do_browse_mass_set_type, vids, new_url, choice)

    def _do_browse_mass_set_type(self, vids, new_url, choice):
        """(Thread) Applique le type à chaque vidéo affichée, avec bilan."""
        new_n = str(new_url).rstrip("/")
        ok = fail = skip = 0
        for i, v in enumerate(vids, 1):
            cur = v.get("type")
            cur = cur.get("url") if isinstance(cur, dict) else cur
            if str(cur).rstrip("/") == new_n:
                skip += 1                              # déjà ce type : on n'appelle pas l'API
                continue
            try:
                self.api.patch_video(v, {"type": new_url})
                v["type"] = new_url                    # MAJ cache local
                # …et propagation aux caches des autres onglets, sinon
                # l'Explorateur et les Chaînes garderaient l'ancien type.
                self._sync_video_caches(v.get("slug"), {"type": new_url})
                ok += 1
            except Exception as e:
                fail += 1
                self._ui(self._log, f"❌ {v.get('slug')} : {e}")
            self._ui(self.browse_status.configure,
                     text=f"⏳  {i}/{len(vids)}…", text_color=T_SECONDAIRE)
        self._ui(self.browse_status.configure,
                 text=f"✅  Type « {choice} » : {ok} modifiée(s), {skip} déjà OK, {fail} échec(s).",
                 text_color=T_SUCCES if not fail else "#f59e0b")
        self._ui(self._log,
                 f"Type en masse « {choice} » : {ok} modifiée(s), {skip} inchangée(s), "
                 f"{fail} échec(s).")
        self._ui(self._browse_apply_filter)            # rafraîchit l'affichage

    def _browse_rename(self, v):
        """Renomme la vidéo sélectionnée (nouveau titre)."""
        new = self.browse_title_entry.get().strip()
        if new and new != v.get("title"):
            self._browse_patch(v, {"title": new}, f"titre → {new}")

    def _browse_patch(self, v, payload, msg):
        """Applique un PATCH sur la vidéo puis met à jour l'affichage."""
        self._run(self._do_browse_patch, v, payload, msg)

    def _do_browse_patch(self, v, payload, msg):
        """(Thread) Applique un PATCH à une vidéo et rafraîchit l'affichage."""
        slug = v.get("slug", "")
        try:
            self.api.patch_video(v, payload)
            v.update(payload)               # met à jour le cache local
            self._sync_video_caches(v.get("slug"), payload)   # … et les autres onglets
            self._ui(self._log, f"✏ {slug} : {msg}")
            self._ui(self._browse_set_msg, f"✅  {msg}", "#22c55e")
            self._ui(self._browse_render_detail)
            self._ui(self._render_browse_list)
        except Exception as e:
            self._ui(self._log, f"❌ {slug} : {e}")
            self._ui(self._browse_set_msg, f"❌  {e}", "#ef4444")
            self._ui(self._browse_render_detail)

    def _debounce(self, cle: str, fonction):
        """Diffère l'exécution de `fonction` jusqu'à une courte pause de frappe.

        Les filtres sont branchés sur l'événement « touche relâchée » : sans
        temporisation, chaque caractère saisi reconstruit toute une liste
        (opération coûteuse en widgets), et les rendus intermédiaires sont
        jetés aussitôt. On annule donc le rendu encore en attente et on en
        replanifie un seul.

        `cle` identifie le différé à remplacer (une clé par filtre)."""
        attrib = f"_job_{cle}"
        job = getattr(self, attrib, None)
        if job:
            try:
                self.after_cancel(job)
            except Exception:
                pass
        setattr(self, attrib, self.after(FILTER_DELAY_MS, fonction))

    @staticmethod
    def _rel_urls(value, normalise: bool = True) -> list:
        """Normalise une RELATION de l'API Pod en liste d'URLs (texte).

        L'API peut renvoyer une relation (`channel`, `theme`, `owner`,
        `restrict_access_to_groups`…) sous trois formes selon le sérialiseur :
          • une URL seule            → "https://…/channels/3/"
          • une liste d'URLs         → ["https://…/channels/3/", …]
          • une liste d'OBJETS       → [{"url": "https://…", "title": "…"}, …]

        Faire `str(x)` sans distinguer le cas « objet » produirait la
        représentation texte du dictionnaire au lieu de l'URL : la comparaison
        d'appartenance échouerait silencieusement et un PATCH enverrait une
        valeur invalide. Cette fonction est donc le passage OBLIGÉ pour lire
        une relation.

        `normalise` : retire la barre oblique finale (pour comparer des URLs).
        """
        if not value:
            return []
        if isinstance(value, (str, dict)):
            value = [value]
        out = []
        for x in value:
            url = x.get("url", "") if isinstance(x, dict) else x
            url = str(url)
            out.append(url.rstrip("/") if normalise else url)
        return out

    # ══════════════════════════════════════════════════════════════════════
    #  MAGASIN DE VIDÉOS PARTAGÉ
    # ══════════════════════════════════════════════════════════════════════

    def ensure_videos(self, on_ready=None, force: bool = False, progress_cb=None):
        """Garantit que `self.videos` est chargé, puis appelle `on_ready`.

        C'est le SEUL point d'entrée pour obtenir la liste des vidéos. Les
        onglets ne rappellent plus `get_all_videos()` chacun de leur côté :
          • si la liste est déjà chargée et `force=False` → `on_ready` est
            appelé immédiatement (aucun appel réseau) ;
          • sinon un scan est lancé en arrière-plan et `on_ready` est appelé à
            la fin, dans le thread principal.

        CONCURRENCE : si un scan est déjà en cours (l'utilisateur ouvre deux
        onglets coup sur coup), on n'en lance PAS un second — le `on_ready` est
        mis en attente et servi avec les autres à la fin du scan en cours.

        `force=True` : rechargement explicite (bouton « Actualiser »).
        `progress_cb` : callback d'avancement du scan (page par page).
        """
        # 1. Données déjà disponibles → service immédiat.
        if self.videos and not force:
            if on_ready:
                on_ready()
            return

        # 2. Un scan tourne déjà → on s'inscrit dans la file d'attente.
        if self.videos_loading:
            if on_ready:
                self._videos_waiters.append(on_ready)
            return

        # 3. Sinon, on lance LE scan.
        self.videos_loading = True
        if on_ready:
            self._videos_waiters.append(on_ready)
        self._run(self._do_load_videos, progress_cb)

    def _do_load_videos(self, progress_cb=None):
        """(Thread) Charge la liste complète des vidéos dans le magasin partagé,
        puis réveille tous les onglets qui l'attendaient."""
        from datetime import datetime
        try:
            vids = self.api.get_all_videos(progress_cb=progress_cb)
            with self._videos_lock:
                # Remplacement du CONTENU (et non de l'objet liste) : d'éventuelles
                # références conservées ailleurs restent ainsi valides.
                self.videos[:] = vids
                self.videos_loaded_at = datetime.now()
            self._ui(self._log, f"📚 {len(vids)} vidéo(s) chargée(s) (cache partagé).")
        except Exception as e:
            self._ui(self._log, f"❌ Chargement des vidéos : {e}")
        finally:
            # Quoi qu'il arrive : libérer le verrou et servir les demandeurs,
            # sinon les onglets resteraient bloqués en attente.
            self.videos_loading = False
            self._ui(self._flush_videos_waiters)

    def _flush_videos_waiters(self):
        """Appelle (thread principal) tous les callbacks en attente du scan."""
        waiters, self._videos_waiters = self._videos_waiters, []
        for cb in waiters:
            try:
                cb()
            except Exception as e:
                self._log(f"❌ Rafraîchissement d'un onglet : {e}")

    def ensure_videos_sync(self, progress_cb=None) -> list:
        """Version SYNCHRONE de `ensure_videos`, à appeler depuis un thread.

        L'onglet Chaînes travaille déjà dans un thread : il lui faut la liste
        immédiatement, pas un rappel différé. Si le magasin est vide, on le
        remplit ici même ; sinon on renvoie directement son contenu.

        Le VERROU garantit qu'un seul chargement remplit le magasin à la fois :
        si un scan asynchrone (`ensure_videos`) tourne déjà, on attend qu'il
        finisse et on réutilise son résultat au lieu de lancer un second scan
        complet qui écraserait le premier.

        Renvoie toujours `self.videos` (la liste partagée)."""
        from datetime import datetime
        # Si un scan ASYNCHRONE tourne déjà, inutile d'en lancer un second : on
        # patiente jusqu'à ce qu'il ait rempli le magasin. (On est dans un thread
        # de travail : attendre ici ne gèle pas l'interface.)
        attente = 0.0
        while self.videos_loading and not self.videos and attente < 120:
            time.sleep(0.2)
            attente += 0.2
        if self.videos:
            return self.videos

        with self._videos_lock:
            # Un autre thread a pu remplir le magasin pendant l'attente du
            # verrou : on re-teste avant de lancer un appel réseau inutile.
            if self.videos:
                return self.videos
            self.videos_loading = True
            try:
                vids = self.api.get_all_videos(progress_cb=progress_cb)
                self.videos[:] = vids
                self.videos_loaded_at = datetime.now()
                self._ui(self._log, f"📚 {len(vids)} vidéo(s) chargée(s) (cache partagé).")
            finally:
                # Toujours relâcher le drapeau, sinon `ensure_videos` croirait
                # qu'un scan tourne encore et n'en lancerait plus jamais.
                self.videos_loading = False
                self._ui(self._flush_videos_waiters)
        return self.videos

    def scan_truncated_warning(self, *endpoints) -> str:
        """Renvoie un avertissement si une lecture a été TRONQUÉE, sinon "".

        La pagination s'arrête sur une limite de sécurité (`max_pages`). Si
        cette limite est atteinte alors qu'il restait des pages, la liste est
        incomplète : les totaux de l'Inventaire sont faux, des vidéos manquent.
        Sans message, rien ne le laisse deviner — c'est le défaut le plus
        trompeur possible pour un outil d'inventaire, puisqu'il produit des
        chiffres crédibles mais faux.

        On interroge le registre PAR RESSOURCE (`api.troncatures`) et non le
        drapeau global : celui-ci était remis à zéro par chaque nouveau scan,
        si bien qu'une lecture terminée après coup pouvait EFFACER l'alerte
        d'une autre.

        `endpoints` : ressources à vérifier, par exemple "/videos/". Sans
        argument, on vérifie toutes celles lues depuis le démarrage.
        """
        api = self.api
        if not api:
            return ""
        try:
            tronque = api.est_tronque(*endpoints)
        except AttributeError:              # client plus ancien
            tronque = bool(getattr(api, "last_scan_truncated", False))
        if not tronque:
            return ""
        # Nommer la ressource concernée : « des comptes manquent » n'appelle pas
        # la même réaction que « des vidéos manquent ».
        noms = {"/videos/": "vidéos", "/users/": "comptes",
                "/owners/": "propriétaires", "/channels/": "chaînes"}
        concernees = [noms.get(e, e.strip("/"))
                      for e, t in getattr(api, "troncatures", {}).items() if t]
        detail = (" (" + ", ".join(concernees) + ")") if concernees else ""
        return (f"⚠️  LISTE INCOMPLÈTE{detail} : la limite de pagination a été "
                "atteinte. Les totaux affichés sont FAUX et des éléments "
                "manquent. Signalez-le au support (support-pod@utoulouse.fr).")

    def videos_stamp(self) -> str:
        """Libellé de fraîcheur du magasin partagé, affiché dans les onglets."""
        if not self.videos_loaded_at:
            return "(non chargé)"
        return (f"{len(self.videos)} vidéo(s) — chargées à "
                f"{self.videos_loaded_at.strftime('%H:%M')}")

    def _sync_video_caches(self, slug, payload=None, removed=False):
        """Applique une modification de vidéo au MAGASIN PARTAGÉ.

        Historique : cette méthode recopiait la modification dans les trois
        caches d'onglets, qui étaient des listes distinctes. Depuis la mise en
        place du magasin partagé `self.videos`, il n'existe plus qu'UN SEUL
        exemplaire de chaque vidéo : la propagation est donc automatique et
        cette méthode ne sert plus qu'à deux choses —
          • retirer une vidéo supprimée de la liste (`removed=True`) ;
          • garantir la mise à jour même si l'appelant a modifié une copie
            plutôt que l'objet du magasin (filet de sécurité).

        Elle est conservée pour rester compatible avec les appels existants et
        éviter tout oubli lors d'une future évolution.
        """
        with self._videos_lock:      # mutation du magasin : protégée
            if removed:
                self.videos[:] = [vv for vv in self.videos if vv.get("slug") != slug]
            elif payload:
                for vv in self.videos:
                    if vv.get("slug") == slug:
                        vv.update(payload)
        # Toute modification d'une vidéo doit se voir à l'écran. Comme TOUS les
        # points de mutation passent par ici, c'est le seul endroit à accrocher :
        # pas de risque d'en oublier un. Passage par _ui car on est souvent
        # appelé depuis un thread de travail.
        self._ui(self.schedule_refresh)

    def schedule_refresh(self, channels: bool = False):
        """Planifie un rafraîchissement des listes après un court délai.

        Appelée après TOUTE modification (suppression, statut, type,
        propriétaire, affectation à une chaîne ou un thème, groupes d'accès…).

        Deux raisons au délai :
          • laisser voir les ✔/✗ posés sur les lignes traitées ;
          • REGROUPER les appels : pendant un traitement par lot, chaque vidéo
            en déclenche un, mais seul le dernier survit (les précédents sont
            annulés) — donc un seul redessin, à la fin du lot.

        `channels=True` : les chaînes/thèmes ont changé, on rafraîchit aussi les
        menus de filtre qui en dépendent dans les autres onglets.
        """
        if channels:
            self._refresh_channels_pending = True
        job = getattr(self, "_refresh_job", None)
        if job:
            try:
                self.after_cancel(job)      # annule le rafraîchissement en attente
            except Exception:
                pass
        self._refresh_job = self.after(REFRESH_DELAY_MS, self._do_scheduled_refresh)

    def _do_scheduled_refresh(self):
        """Exécute le rafraîchissement planifié par `schedule_refresh`."""
        self._refresh_job = None
        if getattr(self, "_refresh_channels_pending", False):
            self._refresh_channels_pending = False
            self._refresh_channel_views()
        self._refresh_video_views()

    def _refresh_channel_views(self):
        """Met à jour les éléments d'interface qui dépendent des CHAÎNES.

        Une chaîne renommée, créée ou supprimée doit se refléter dans les menus
        de filtre « chaîne » de l'onglet Vidéos et de l'Explorateur, ainsi que
        dans les libellés de chaîne affichés sur chaque vidéo — sinon ces
        onglets continuent d'afficher un nom périmé, voire une chaîne disparue.
        """
        chans = getattr(self, "ct_channels", None) or getattr(self, "browse_channels", None)
        if not chans:
            return
        self.browse_channels = chans
        self.browse_chan_by_url = {str(c.get("url", "")).rstrip("/"): c.get("title", "?")
                                   for c in chans}
        try:
            if hasattr(self, "browse_chan"):
                self._browse_refresh_channel_menu()
        except Exception as e:
            self._log(f"Rafraîchissement du menu chaînes (Vidéos) : {e}")


    def _refresh_video_views(self):
        """Recalcule les listes AFFICHÉES des onglets qui montrent des vidéos.

        Le magasin `self.videos` est partagé, mais chaque onglet en garde une
        PROJECTION filtrée (`browse_filtered`) construite au
        moment du dernier filtrage. Supprimer une vidéo du magasin ne suffit
        donc pas : sans ce recalcul, la vidéo disparaît des données mais reste
        AFFICHÉE dans les listes déjà dessinées (vidéo « fantôme »).

        À appeler après toute suppression ou modification par lot. Sans effet
        visible si les onglets ne sont pas encore construits.
        """
        # La sélection de l'onglet Vidéos peut pointer sur une vidéo supprimée :
        # on la libère pour ne pas afficher le détail d'un objet disparu.
        if self.browse_selected and self.browse_selected not in self.videos:
            self.browse_selected = None
            try:
                self._browse_render_detail()
            except Exception:
                pass
        # Onglet Vidéos
        try:
            if hasattr(self, "browse_list"):
                self._browse_do_filter()
        except Exception as e:
            self._log(f"Rafraîchissement onglet Vidéos : {e}")

        # Les trois onglets ci-dessous gardent une PROJECTION du magasin
        # (`encode_videos`, `stats_videos`, `reassign_videos`). Comme
        # `browse_filtered`, ces listes ne se mettent pas à jour toutes seules :
        # sans ce recalcul, l'Encodage proposait de relancer une vidéo
        # supprimée, et l'Inventaire affichait un total périmé sous une mention
        # « chargé à … » qui inspirait confiance.
        #
        # On ne recharge PAS depuis le réseau : on relit le magasin, déjà à jour.

        # Encodage — la liste complète, filtrée à l'affichage.
        try:
            if hasattr(self, "encode_list") and self.encode_videos:
                self.encode_videos = list(self.videos)
                self._render_encode()
        except Exception as e:
            self._log(f"Rafraîchissement onglet Encodage : {e}")

        # Inventaire — les agrégats sont recalculés à partir du magasin.
        try:
            if getattr(self, "stats_data", None):
                self._recalculer_stats()
        except Exception as e:
            self._log(f"Rafraîchissement onglet Inventaire : {e}")

        # Réaffectation — on retire simplement les vidéos disparues de l'aperçu.
        # Le recalculer entièrement effacerait les cases cochées par
        # l'utilisateur, ce qui serait plus gênant qu'utile.
        try:
            if getattr(self, "reassign_videos", None):
                presents = {v.get("slug") for v in self.videos}
                avant = len(self.reassign_videos)
                self.reassign_videos = [v for v in self.reassign_videos
                                        if v.get("slug") in presents]
                if len(self.reassign_videos) != avant:
                    self._render_reassign_preview()
        except Exception as e:
            self._log(f"Rafraîchissement onglet Réaffectation : {e}")


    def _loaded_stamp(self) -> str:
        """Renvoie « chargé à HH:MM » pour indiquer la fraîcheur des données
        affichées dans un onglet (l'utilisateur sait ainsi quand rafraîchir,
        par exemple après une modification faite via le site web Pod)."""
        from datetime import datetime
        return f"chargé à {datetime.now().strftime('%H:%M')}"

    def _browse_set_msg(self, text, color):
        """Affiche un message d'état dans l'onglet Vidéos."""
        if hasattr(self, "browse_msg") and self.browse_msg.winfo_exists():
            self.browse_msg.configure(text=text, text_color=color)

    def _browse_edit_owners(self, v):
        """Ouvre le sélecteur d'utilisateurs (OwnerPicker) pour les co-propriétaires."""
        # Pré-sélection : co-propriétaires actuels (URL → libellé lisible)
        user_by_url = {str(u.get("url", "")).rstrip("/"): u for u in (self.all_users or [])}
        pre = {}
        for ourl in (v.get("additional_owners") or []):
            u = user_by_url.get(str(ourl).rstrip("/"))
            pre[ourl] = self._user_label(u) if u else ourl
        OwnerPicker(self, on_done=lambda urls, labels: self._browse_apply_owners(v, urls),
                    title="Co-propriétaires de la vidéo", preselected=pre)

    def _browse_apply_owners(self, v, urls):
        """Met à jour les propriétaires additionnels de la vidéo."""
        self._run(self._do_browse_patch, v, {"additional_owners": list(urls)},
                  f"{len(urls)} co-propriétaire(s)")

    def _browse_edit_channels(self, v):
        """Ouvre le sélecteur de chaînes pour cette vidéo."""
        cur = v.get("channel") or []
        if isinstance(cur, str):
            cur = [cur]
        pre = {c: self.browse_chan_by_url.get(str(c).rstrip("/"), str(c)) for c in cur}
        ChannelPicker(self, self.browse_channels,
                      on_done=lambda urls, labels: self._browse_apply_channels(v, urls),
                      title="Chaînes de la vidéo", preselected=pre)

    def _browse_apply_channels(self, v, urls):
        """Affecte la vidéo aux chaînes choisies."""
        self._run(self._do_browse_patch, v, {"channel": list(urls)},
                  f"{len(urls)} chaîne(s)")

    def _open_video_in_browser(self, slug):
        """Ouvre la page publique Pod de la vidéo dans le navigateur par défaut.
        (PodAdmin n'a pas de lecteur intégré : la lecture est déléguée au web.)"""
        import webbrowser
        base = str(self.config_data.get("url", "")).rstrip("/")
        if not base:
            base = "https://videos.utoulouse.fr"
        url = f"{base}/video/{slug}/"
        try:
            webbrowser.open(url)
            self._browse_set_msg(f"Ouverture de {slug} dans le navigateur…", "gray")
        except Exception as e:
            self._browse_set_msg(f"Impossible d'ouvrir le navigateur : {e}", "#f59e0b")

    def _browse_replace_source(self, v):
        """Remplace le fichier source d'une vidéo puis relance l'encodage.
        Demande le fichier + double confirmation (opération destructive)."""
        if not self.api:
            self._browse_set_msg("Connectez-vous d'abord.", "#f59e0b")
            return
        path = filedialog.askopenfilename(
            title="Choisir le nouveau fichier vidéo",
            filetypes=[("Vidéos", "*.mp4 *.mov *.avi *.mkv *.webm *.m4v *.wmv *.flv"),
                       ("Tous les fichiers", "*.*")])
        if not path:
            return
        import os as _os
        size_mo = _os.path.getsize(path) / 1024 / 1024
        # Au-delà du seuil, le remplacement bascule automatiquement sur la voie
        # CHUNKÉE, via la session web du compte véhicule. Celui-ci est embarqué
        # (DEPOT) : la bascule est donc TRANSPARENTE, sans configuration.
        gros = _os.path.getsize(path) > cfg.CHUNK_THRESHOLD_BYTES
        if not messagebox.askyesno(
                "⚠️  Remplacer le fichier source",
                f"Remplacer le fichier de « {v.get('title')} » par :\n"
                f"{_os.path.basename(path)}  ({size_mo:.0f} Mo)\n\n"
                "L'ancien fichier sera écrasé et la vidéo ré-encodée. "
                "Le titre, les chaînes et les droits sont conservés.\n\n"
                "Cette action ne peut pas être annulée. Continuer ?"):
            return
        self._browse_set_msg("⏳  Envoi du nouveau fichier…", "gray")
        # Fenêtre MODALE de progression : elle bloque toute autre manipulation
        # pendant l'envoi (une action concurrente couperait le téléversement)
        # et montre l'avancement.
        modal = ProgressModal(
            self,
            title="Remplacer le fichier & ré-encoder",
            subtitle=f"« {v.get('title', '')} »\n{_os.path.basename(path)} ({size_mo:.0f} Mo) — "
                     f"méthode {'par MORCEAUX (compte véhicule)' if gros else 'directe'}.",
            intro="Préparation de l'envoi…")
        self._run(self._do_browse_replace_source, v, path, modal, gros)

    def _do_browse_replace_source(self, v, path, modal=None, gros=False):
        """(Thread) PATCH du nouveau fichier (streamé + retry) puis encodage.

        `modal` : fenêtre ProgressModal ouverte par l'appelant. Elle est mise à
        jour à chaque étape et TOUJOURS clôturée en fin de traitement (y compris
        en cas d'erreur), pour ne jamais laisser l'interface bloquée."""
        slug = v.get("slug", "")
        import os as _os
        fname = _os.path.basename(path)

        def progress(sent, tot):
            """Callback de progression de l'envoi (met à jour la barre du fichier)."""
            self._ui(self._browse_set_msg,
                     f"⏳  Envoi {fname} — {sent/1024/1024:.0f}/{tot/1024/1024:.0f} Mo", "gray")
            if modal:
                self._ui(modal.set_progress, (sent / tot if tot else 0),
                         f"{sent/1024/1024:.0f} / {tot/1024/1024:.0f} Mo envoyés")

        def on_retry(attempt, total_try, err):
            """Callback de relance : trace la nouvelle tentative dans le Journal."""
            self._ui(self._log,
                     f"⟳ Nouvelle tentative {attempt}/{total_try} (remplacement {slug})…")
            self._ui(self._browse_set_msg,
                     f"⟳  Coupure réseau — nouvelle tentative {attempt}…", "#f59e0b")
            if modal:
                self._ui(modal.set_phase,
                         f"⟳  Coupure réseau — nouvelle tentative {attempt}/{total_try}…",
                         "#f59e0b")

        try:
            if gros:
                # ── Gros fichier : remplacement par MORCEAUX via le VÉHICULE ──
                # On finalise le chunké en passant le SLUG cible → Pod REMPLACE
                # le fichier de la vidéo existante (pas de nouvel enregistrement).
                self._ui(self._log,
                         f"🎬 Remplacement chunké (> {cfg.CHUNK_THRESHOLD_BYTES//1024//1024} Mo) "
                         f"pour {slug} ({fname})…")
                if modal:
                    self._ui(modal.set_phase, "Connexion au compte véhicule…")
                chunked = PodChunkedSession(self.config_data.get("url", ""),
                                            self.vehicle_username, self.vehicle_password)
                chunked.login()
                if modal:
                    self._ui(modal.set_phase, "Étape 1/3 — Envoi du fichier par morceaux…")
                try:
                    returned = chunked.upload_video_chunked(
                        path, chunk_size=cfg.CHUNK_SIZE_BYTES,
                        progress_cb=progress, retry_cb=on_retry,
                        target_slug=slug)          # ← slug cible = remplacement
                except PodChunkedError as ce:
                    # 502/503/504 à la finalisation : Pod termine côté serveur.
                    # On NE relance PAS l'encodage (le fichier n'est peut-être
                    # pas encore assemblé) : on informe et on s'arrête là.
                    if ce.status in (502, 503, 504):
                        self._ui(self._log,
                                 f"⏳ Remplacement {slug} : finalisation coupée par la passerelle "
                                 f"(HTTP {ce.status}) — Pod termine côté serveur.")
                        self._ui(self._browse_set_msg,
                                 "⏳  Remplacement en cours de finalisation côté serveur "
                                 "(jusqu'à ~10 min). Vérifiez côté web, puis relancez "
                                 "l'encodage.", "#f59e0b")
                        if modal:
                            self._ui(modal.finish, False,
                                     "Le fichier est envoyé, mais le serveur termine encore son "
                                     "assemblage (cela peut prendre ~10 min). Vérifiez la vidéo "
                                     "sur le site, puis relancez le ré-encodage si besoin.")
                        return
                    raise
                finally:
                    chunked.close()
                if returned and returned != slug:
                    # Sécurité : un slug différent = vidéo neuve créée au lieu
                    # d'un remplacement. On le signale clairement.
                    self._ui(self._log,
                             f"⚠️ Le remplacement a renvoyé un slug différent ({returned}) : "
                             "une vidéo neuve a peut-être été créée. À vérifier côté web.")
            else:
                # ── Fichier ≤ seuil : PATCH direct streamé (inchangé) ──
                if modal:
                    self._ui(modal.set_phase, "Étape 1/2 — Envoi du nouveau fichier…")
                self.api.replace_video_file(v, path, progress_cb=progress, retry_cb=on_retry)
            self._ui(self._log, f"🎬 Fichier remplacé pour {slug} ({fname}).")
            # Relancer l'encodage sur le nouveau fichier
            if modal:
                # L'avancement n'est plus mesurable ici → animation continue.
                self._ui(modal.set_phase,
                         f"Étape {'3/3' if gros else '2/2'} — Lancement du ré-encodage…")
                self._ui(modal.set_indeterminate, True)
            try:
                self.api.launch_encoding(slug)
                v["encoded"] = False
                v["encoding_in_progress"] = True
                self._sync_video_caches(slug, {"encoded": False, "encoding_in_progress": True})
                self._ui(self._log, f"⚙ Ré-encodage lancé pour {slug}.")
                self._ui(self._browse_set_msg,
                         "✅  Fichier remplacé, ré-encodage lancé.", "#22c55e")
                if modal:
                    self._ui(modal.finish, True,
                             "Fichier remplacé et ré-encodage lancé. La vidéo sera de nouveau "
                             "disponible à la fin de l'encodage.")
            except Exception as e:
                self._ui(self._browse_set_msg,
                         f"Fichier remplacé, mais encodage non lancé : {e}", "#f59e0b")
                self._ui(self._log, f"❌ Encodage non lancé ({slug}) : {e}")
                if modal:
                    self._ui(modal.finish, False,
                             f"Fichier remplacé, mais le ré-encodage n'a pas pu être lancé : {e}")
            self._ui(self._browse_render_detail)
        except Exception as e:
            self._ui(self._browse_set_msg, f"❌  {e}", "#ef4444")
            self._ui(self._log, f"❌ Remplacement {slug} : {e}")
            if modal:
                self._ui(modal.finish, False, f"Le remplacement a échoué : {e}")
        finally:
            # Quoi qu'il arrive, la modale doit être déverrouillée : une fenêtre
            # modale restée bloquée rendrait l'application inutilisable.
            if modal:
                self._ui(modal.ensure_unlocked)

    def _browse_delete(self, v):
        """Supprime la vidéo sélectionnée (après confirmation)."""
        if not messagebox.askyesno(
                "⚠️  Supprimer la vidéo",
                f"Supprimer DÉFINITIVEMENT « {v.get('title')} » ?\n\n"
                "Cette action est IRRÉVERSIBLE (pas de corbeille sur Pod)."):
            return
        if not messagebox.askyesno("Dernière confirmation",
                                   "Confirmez-vous la suppression de cette vidéo ?"):
            return
        self._run(self._do_browse_delete, v)

    def _do_browse_delete(self, v):
        """(Thread) Exécute la suppression de la vidéo côté serveur."""
        slug = v.get("slug", "")
        try:
            self.api.delete_video(v)
            with self._videos_lock:          # mutation du magasin : protégée
                if v in self.videos:
                    self.videos.remove(v)
            # Retirer aussi la vidéo des caches des AUTRES onglets (Explorateur,
            # Chaînes) : sans cela elle y resterait en « fantôme » et toute action
            # dessus échouerait (404), en faussant au passage les calculs de
            # groupes/appartenance de l'onglet Chaînes.
            self._sync_video_caches(slug, removed=True)
            self.browse_selected = None
            self._ui(self._log, f"🗑 Vidéo supprimée : {slug}")
            self._ui(self._browse_render_detail)
        except Exception as e:
            self._ui(self._log, f"❌ Suppression {slug} : {e}")
            self._ui(self._browse_set_msg, f"❌  {e}", "#ef4444")

    # ═════════════════════════════════════════════════════════════════════
    #  ONGLET RÉAFFECTATION — changer le propriétaire de vidéos par lot
    # ═════════════════════════════════════════════════════════════════════
    #
    #  Principe : on choisit un ANCIEN propriétaire (source) et un NOUVEAU
    #  (cible), on liste les vidéos de la source en mode « aperçu » (dry-run),
    #  l'utilisateur coche/décoche, confirme, puis on applique en PATCHant le
    #  champ `owner` de chaque vidéo. Aucune écriture n'a lieu avant le clic
    #  explicite sur « Appliquer ».

    def _build_tab_reassign(self):
        # Cadre racine de l'onglet (enregistré dans self.tabs)
        frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tabs["reassign"] = frame

        # Titre + courte explication du flux de travail
        ctk.CTkLabel(frame, text="🔄  Réaffectation de propriétaire",
                     font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(
            frame,
            text="Transférer les vidéos d'un compte vers un autre (ex. départ d'un agent). "
                 "Choisissez l'ancien et le nouveau propriétaire, lancez l'aperçu, "
                 "vérifiez la liste, puis appliquez.",
            text_color=T_SECONDAIRE, font=ctk.CTkFont(size=12),
            justify="left", wraplength=860).pack(anchor="w", pady=(0, 10))

        # — Deux sélecteurs de comptes côte à côte (source / cible) —
        pickers = ctk.CTkFrame(frame, fg_color="transparent")
        pickers.pack(fill="x")
        pickers.columnconfigure(0, weight=1)   # colonnes de largeur égale
        pickers.columnconfigure(1, weight=1)

        # Liste mémorisant les deux sélecteurs, pour les rafraîchir quand la
        # liste des comptes est (re)chargée. Initialisée AVANT _mini_user_picker.
        self._reassign_pickers = []

        # Sélecteur SOURCE (ancien propriétaire)
        src_box = self._mini_user_picker(
            pickers, "① Ancien propriétaire (source)", self._on_pick_source)
        src_box["frame"].grid(row=0, column=0, padx=(0, 6), sticky="nsew")
        # Sélecteur CIBLE (nouveau propriétaire)
        tgt_box = self._mini_user_picker(
            pickers, "② Nouveau propriétaire (cible)", self._on_pick_target)
        tgt_box["frame"].grid(row=0, column=1, padx=(6, 0), sticky="nsew")

        # Comptes sélectionnés (dicts utilisateur) ; None tant que rien n'est choisi
        self.reassign_source = None
        self.reassign_target = None

        # — Option : garder l'ancien propriétaire en co-propriétaire —
        opt = ctk.CTkFrame(frame, fg_color="transparent")
        opt.pack(fill="x", pady=(8, 2))
        # Si coché, l'ancien propriétaire est ajouté aux additional_owners :
        # il conserve les droits d'édition (mais pas de suppression).
        self.reassign_keep_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            opt, text="Conserver l'ancien propriétaire en co-propriétaire (additional_owners)",
            variable=self.reassign_keep_var).pack(side="left")

        # — Boutons d'action —
        actions = ctk.CTkFrame(frame, fg_color="transparent")
        actions.pack(fill="x", pady=(6, 4))
        # Aperçu = lecture seule, n'écrit RIEN
        ctk.CTkButton(actions, text="🔍  Aperçu (dry-run)", fg_color=C_ACTION,
                      hover_color=C_ACTION_SURV, command=self._reassign_preview).pack(side="left")
        # Appliquer = action en masse ; désactivé tant qu'aucun aperçu n'est fait
        self.reassign_apply_btn = ctk.CTkButton(
            actions, text="✅  Appliquer la réaffectation", fg_color=C_SUCCES,
            hover_color=C_SUCCES_SURV, state="disabled", command=self._reassign_confirm)
        self.reassign_apply_btn.pack(side="left", padx=10)
        # Libellé de progression / d'état
        self.reassign_progress = ctk.CTkLabel(actions, text="", text_color=T_SECONDAIRE,
                                              font=ctk.CTkFont(size=11))
        self.reassign_progress.pack(side="left", padx=8)

        # — Zone d'aperçu : la liste des vidéos concernées (cases à cocher) —
        self.reassign_results = ctk.CTkScrollableFrame(frame, label_text="Aperçu des vidéos", fg_color=S_CARTE, label_anchor="w",
                                                  label_font=ctk.CTkFont(size=12, weight="bold"))
        self.reassign_results.pack(fill="both", expand=True, pady=(4, 0))

        # Structures de données de l'aperçu :
        self.reassign_videos = []      # vidéos de la source (liste de dicts)
        self.reassign_rowvars = {}     # slug → BooleanVar (cochée = à réaffecter)
        self.reassign_rowlbls = {}     # slug → label de statut (✔/✗ après application)

    # ── Sélecteur de compte compact et réutilisable ───────────────────────

    def _mini_user_picker(self, parent, title, on_pick):
        """Construit un sélecteur de compte (titre + filtre + liste cliquable +
        libellé du choix). Renvoie un dict d'état réutilisé par _render_mini_picker.
        `on_pick(user)` est appelé quand l'utilisateur clique un compte."""
        box = ctk.CTkFrame(parent, fg_color=S_CARTE)
        ctk.CTkLabel(box, text=title, font=ctk.CTkFont(weight="bold")).pack(
            anchor="w", padx=10, pady=(8, 2))
        # Champ de filtre → filtrage CLIENT instantané (aucun appel serveur)
        fe = ctk.CTkEntry(box, placeholder_text="🔍 nom / identifiant…")
        fe.pack(fill="x", padx=10, pady=(0, 6))
        # Liste défilante des comptes correspondant au filtre
        res = ctk.CTkScrollableFrame(box, height=150, fg_color=S_CARTE)
        res.pack(fill="both", expand=True, padx=10, pady=(0, 6))
        # Rappel du compte actuellement sélectionné
        chosen = ctk.CTkLabel(box, text="Sélection : (aucune)", text_color=T_SECONDAIRE,
                              font=ctk.CTkFont(size=11), anchor="w")
        chosen.pack(fill="x", padx=10, pady=(0, 8))
        # État partagé entre le picker et son moteur de rendu
        state = {"frame": box, "filter": fe, "results": res,
                 "chosen": chosen, "on_pick": on_pick, "selected": None}
        # Re-render à chaque frappe dans le filtre
        fe.bind("<KeyRelease>",
                lambda e, s=state: self._debounce("mini", lambda: self._render_mini_picker(s)))
        self._reassign_pickers.append(state)
        self._render_mini_picker(state)
        return state

    def _render_mini_picker(self, state):
        """(Ré)affiche la liste filtrée d'un sélecteur de compte."""
        flt = state["filter"].get().strip().lower()
        # Vider la liste précédente
        for w in state["results"].winfo_children():
            w.destroy()
        # Aucun compte encore chargé
        if not self.all_users:
            ctk.CTkLabel(state["results"], text="Connectez-vous puis rechargez les comptes.",
                         text_color=T_SECONDAIRE).pack(pady=8)
            return
        # Filtrage client + plafond d'affichage (Tk gèle au-delà de quelques centaines)
        matches = [u for u in self.all_users
                   if not flt or flt in self._user_label(u).lower()]
        CAP = 200
        sel = state["selected"]
        for u in matches[:CAP]:
            # Marque ✅ le compte sélectionné dans ce picker
            is_sel = sel is not None and u.get("username") == sel.get("username")
            label = ("✅  " if is_sel else "      ") + self._user_label(u)
            ctk.CTkButton(
                state["results"], text=label, anchor="w", height=26,
                fg_color=S_SELECTION if is_sel else "transparent",
                text_color=("gray10", "gray90"), hover_color=("gray75", "gray28"),
                font=ctk.CTkFont(size=12),
                command=lambda uu=u, s=state: self._mini_pick(s, uu)).pack(fill="x", pady=1)
        # Indications de fin de liste
        if len(matches) > CAP:
            ctk.CTkLabel(state["results"],
                         text=f"… +{len(matches) - CAP} autres. Affinez le filtre.",
                         text_color=T_SECONDAIRE).pack(pady=4)
        elif not matches:
            ctk.CTkLabel(state["results"], text="Aucun compte.", text_color=T_SECONDAIRE).pack(pady=6)

    def _mini_pick(self, state, user):
        """Enregistre le compte choisi dans un sélecteur et notifie l'appelant."""
        state["selected"] = user
        state["chosen"].configure(text=f"Sélection : {self._user_label(user)}",
                                  text_color=T_SUCCES)
        self._render_mini_picker(state)     # met à jour la coche ✅
        state["on_pick"](user)              # callback spécifique (source ou cible)

    def _on_pick_source(self, user):
        # Mémorise l'ancien propriétaire ; tout aperçu précédent devient caduc
        self.reassign_source = user
        self.reassign_apply_btn.configure(state="disabled")

    def _on_pick_target(self, user):
        # Mémorise le nouveau propriétaire
        self.reassign_target = user

    def _refresh_reassign_pickers(self):
        """Rafraîchit les deux sélecteurs (appelé quand les comptes sont chargés)."""
        for st in getattr(self, "_reassign_pickers", []):
            self._render_mini_picker(st)

    # ── Appariement vidéo ↔ propriétaire (robuste au format du champ owner) ─

    def _video_owner_id(self, video):
        """Renvoie l'identifiant du propriétaire d'une vidéo, quel que soit le
        format renvoyé par l'API (URL, dict imbriqué, username ou id)."""
        o = video.get("owner")
        if isinstance(o, dict):                       # owner imbriqué {url, username…}
            return o.get("url") or o.get("username") or ""
        return o if o is not None else ""             # URL (str), username ou id

    def _video_belongs_to(self, video, user):
        """Vrai si la vidéo appartient au compte `user`. On compare l'identifiant
        propriétaire à l'URL, au username et à l'id du compte (couvre tous les cas)."""
        oid = str(self._video_owner_id(video))
        if not oid:
            return False
        candidates = {str(user.get("url", "")).rstrip("/"),
                      str(user.get("username", "")),
                      str(user.get("id", ""))}
        # Comparaison directe, ou URL se terminant par /users/<id>
        return (oid.rstrip("/") in candidates
                or oid.rstrip("/").endswith(f"/users/{user.get('id')}"))

    # ── Aperçu (dry-run) : lister les vidéos de la source ──────────────────

    def _reassign_preview(self):
        """Lance l'aperçu : vérifie les sélections puis scanne l'instance en
        arrière-plan. AUCUNE modification n'est effectuée ici."""
        if not self.api:
            self.reassign_progress.configure(text="Connectez-vous d'abord.",
                                             text_color=T_ALERTE)
            return
        if not self.reassign_source:
            self.reassign_progress.configure(text="Choisissez l'ancien propriétaire (source).",
                                             text_color=T_ALERTE)
            return
        self.reassign_apply_btn.configure(state="disabled")
        self.reassign_progress.configure(text="⏳  Analyse des vidéos…", text_color=T_SECONDAIRE)
        self._run(self._do_reassign_preview)

    def _do_reassign_preview(self):
        """(Thread) Scan complet de l'instance + filtrage par propriétaire."""
        try:
            # Callback de progression du scan paginé (mis à jour via le thread UI)
            def prog(n):
                self._ui(self.reassign_progress.configure,
                         text=f"⏳  {n} vidéos lues…", text_color=T_SECONDAIRE)
            # Magasin partagé (voir _do_encode_scan).
            videos = self.ensure_videos_sync(progress_cb=prog)
            # Ne conserver que les vidéos appartenant à la source
            mine = [v for v in videos if self._video_belongs_to(v, self.reassign_source)]
            self.reassign_videos = mine
            self._ui(self._render_reassign_preview)
            self._ui(self._log,
                     f"Aperçu réaffectation : {len(mine)} vidéo(s) pour "
                     f"{self.reassign_source.get('username')} (sur {len(videos)} au total).")
        except Exception as e:
            self._ui(self.reassign_progress.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Erreur aperçu réaffectation : {e}")

    def _render_reassign_preview(self):
        """Affiche la liste des vidéos concernées avec une case à cocher chacune."""
        # Vider l'aperçu précédent et réinitialiser les structures
        for w in self.reassign_results.winfo_children():
            w.destroy()
        self.reassign_rowvars = {}
        self.reassign_rowlbls = {}

        if not self.reassign_videos:
            ctk.CTkLabel(self.reassign_results,
                         text="Aucune vidéo trouvée pour ce compte.",
                         text_color=T_SECONDAIRE).pack(pady=10)
            self.reassign_progress.configure(text="0 vidéo.", text_color=T_SECONDAIRE)
            self.reassign_apply_btn.configure(state="disabled")
            return

        # En-tête : (dé)sélection globale
        head = ctk.CTkFrame(self.reassign_results, fg_color="transparent")
        head.pack(fill="x", pady=(0, 4))
        ctk.CTkButton(head, text="Tout cocher", width=100, height=24, fg_color=C_NEUTRE,
                      command=lambda: self._reassign_check_all(True), text_color=T_SUR_NEUTRE).pack(side="left", padx=2)
        ctk.CTkButton(head, text="Tout décocher", width=110, height=24, fg_color=C_NEUTRE,
                      command=lambda: self._reassign_check_all(False), text_color=T_SUR_NEUTRE).pack(side="left", padx=2)

        # Une ligne par vidéo : [case] titre · slug … [statut ✔/✗]
        for v in self.reassign_videos:
            slug = v.get("slug", "?")
            row = ctk.CTkFrame(self.reassign_results, fg_color=S_LIGNE,
                               corner_radius=6)
            row.pack(fill="x", pady=2)
            var = ctk.BooleanVar(value=True)        # cochée par défaut = sera réaffectée
            self.reassign_rowvars[slug] = var
            ctk.CTkCheckBox(row, text="", variable=var, width=24).pack(side="left", padx=(8, 0))
            title = (v.get("title") or "(sans titre)")[:70]
            ctk.CTkLabel(row, text=f"{title}   ·   {slug}", anchor="w",
                         font=ctk.CTkFont(size=12)).pack(
                side="left", padx=8, pady=6, fill="x", expand=True)
            stat = ctk.CTkLabel(row, text="", width=24, font=ctk.CTkFont(size=12))
            stat.pack(side="right", padx=8)
            self.reassign_rowlbls[slug] = stat

        self.reassign_progress.configure(
            text=f"{len(self.reassign_videos)} vidéo(s) prêtes à réaffecter.",
            text_color=T_SUCCES)
        # L'aperçu est prêt → on autorise l'application
        self.reassign_apply_btn.configure(state="normal")

    def _reassign_check_all(self, value: bool):
        """Coche ou décoche toutes les vidéos de l'aperçu."""
        for var in self.reassign_rowvars.values():
            var.set(value)

    # ── Application (action en masse, après confirmation explicite) ────────

    def _reassign_confirm(self):
        """Vérifie la cible, récapitule, puis demande confirmation avant d'écrire."""
        if not self.reassign_target:
            self.reassign_progress.configure(text="Choisissez le nouveau propriétaire (cible).",
                                             text_color=T_ALERTE)
            return
        # Les deux listes sont indépendantes : rien n'empêchait de choisir le
        # MÊME compte des deux côtés. On émettait alors un PATCH par vidéo pour
        # la réattribuer à son propriétaire actuel — et, case cochée, pour
        # l'ajouter comme son propre co-propriétaire. Sans effet utile, mais
        # coûteux et déroutant sur un lot important.
        src_url = str((self.reassign_source or {}).get("url", "")).rstrip("/")
        tgt_url = str((self.reassign_target or {}).get("url", "")).rstrip("/")
        if src_url and src_url == tgt_url:
            self.reassign_progress.configure(
                text="Le propriétaire actuel et le nouveau sont le même compte : "
                     "il n'y a rien à réaffecter.",
                text_color=T_ALERTE)
            return
        # Vidéos réellement cochées dans l'aperçu
        todo = [v for v in self.reassign_videos
                if self.reassign_rowvars.get(v.get("slug"))
                and self.reassign_rowvars[v.get("slug")].get()]
        if not todo:
            self.reassign_progress.configure(text="Aucune vidéo cochée.", text_color=T_ALERTE)
            return
        # Deux valeurs distinctes, qui étaient auparavant confondues sous le même
        # nom : `mention` est le TEXTE du récapitulatif, `keep` le BOOLÉEN passé
        # au traitement. Les mélanger fonctionnait par accident (une chaîne vide
        # est fausse) et masquait la relecture Tk hors thread principal.
        keep = bool(self.reassign_keep_var.get())
        mention = " (l'ancien propriétaire reste co-propriétaire)" if keep else ""
        # Dialogue de confirmation récapitulatif (dernier garde-fou avant écriture)
        ok = messagebox.askyesno(
            "Confirmer la réaffectation",
            f"Réaffecter {len(todo)} vidéo(s)\n\n"
            f"de :   {self._user_label(self.reassign_source)}\n"
            f"vers : {self._user_label(self.reassign_target)}{mention}\n\n"
            "Cette opération modifie le propriétaire de chaque vidéo. Continuer ?")
        if not ok:
            return
        # Désactiver le bouton pendant le traitement (évite les double-clics)
        self.reassign_apply_btn.configure(state="disabled")
        # Lecture du widget ICI (thread principal), puis passage en argument :
        # lire une variable Tk depuis un thread de travail n'est pas fiable.
        # `keep` est reçu en ARGUMENT (lu par l'appelant dans le thread principal).
        self._run(self._do_reassign_apply, todo, keep)

    def _do_reassign_apply(self, todo, keep: bool = False):
        """(Thread) Applique la réaffectation vidéo par vidéo via PATCH owner."""
        tgt = self.reassign_target
        # `keep` est reçu en ARGUMENT, lu par l'appelant dans le thread
        # principal. Ne PAS le relire ici : Tcl n'est pas thread-safe, et
        # c'est précisément la raison d'être de ce paramètre.
        ok = fail = 0
        for i, v in enumerate(todo, 1):
            slug = v.get("slug", "")
            try:
                add = None
                if keep:
                    # Conserver les co-propriétaires existants + ajouter l'ancien
                    existing = list(v.get("additional_owners") or [])
                    add = list(dict.fromkeys(existing + [self.reassign_source.get("url")]))
                # ÉCRITURE réelle : PATCH /rest/videos/<id>/ {owner: <url cible>}
                self.api.set_video_owner(v, tgt.get("url"), additional_owner_urls=add)
                v["owner"] = tgt.get("url")       # met à jour le cache local
                # …et les caches des autres onglets, sinon Vidéos / Explorateur /
                # Chaînes continueraient d'afficher l'ANCIEN propriétaire.
                sync_payload = {"owner": tgt.get("url")}
                if add is not None:
                    sync_payload["additional_owners"] = add
                self._sync_video_caches(slug, sync_payload)
                ok += 1
                self._ui(self._mark_reassign_row, slug, True)
            except Exception as e:
                fail += 1
                self._ui(self._mark_reassign_row, slug, False)
                self._ui(self._log, f"  ✗ {slug} : {e}")
            # Progression
            self._ui(self.reassign_progress.configure,
                     text=f"⏳  {i}/{len(todo)}…", text_color=T_SECONDAIRE)
        # Bilan final
        self._ui(self.reassign_progress.configure,
                 text=f"Terminé : {ok} réaffectée(s), {fail} échec(s).",
                 text_color=T_SUCCES if not fail else "#f59e0b")
        self._ui(self._log,
                 f"Réaffectation {self.reassign_source.get('username')} → "
                 f"{tgt.get('username')} : {ok} OK, {fail} échec(s).")
        # Réactiver le bouton pour une éventuelle nouvelle opération
        self._ui(self.reassign_apply_btn.configure, state="normal")

    def _mark_reassign_row(self, slug, success: bool):
        """Pose un ✔ (vert) ou ✗ (rouge) sur la ligne d'une vidéo traitée."""
        lbl = self.reassign_rowlbls.get(slug)
        if lbl:
            lbl.configure(text="✔" if success else "✗",
                          text_color=T_SUCCES if success else "#ef4444")

    # ═════════════════════════════════════════════════════════════════════
    #  FENÊTRES ET UTILITAIRES PARTAGÉS
    #  (L'onglet Explorateur a été fusionné dans l'onglet Vidéos ; ces fonctions
    #   lui survivent car les actions groupées s'en servent.)
    # ═════════════════════════════════════════════════════════════════════
    #  FENÊTRES PARTAGÉES PAR LES ACTIONS GROUPÉES
    # ═════════════════════════════════════════════════════════════════════

    def _duplicate_title_videos(self, vids: list) -> list:
        """Renvoie les vidéos dont le titre (normalisé) apparaît plus d'une fois,
        triées par titre pour regrouper visuellement les doublons."""
        from collections import Counter
        def norm(v):
            """Clé de tri normalisée (titre en minuscules)."""
            return (v.get("title") or "").strip().lower()
        counts = Counter(norm(v) for v in vids if norm(v))
        dups = [v for v in vids if norm(v) and counts[norm(v)] > 1]
        dups.sort(key=norm)
        return dups

    def _months_ago_iso(self, months: int) -> str:
        """Renvoie la date d'il y a `months` mois au format AAAA-MM-JJ."""
        import calendar
        today = datetime.now().date()
        m, y = today.month - months, today.year
        while m <= 0:          # report sur les années précédentes
            m += 12
            y -= 1
        # On borne le jour au dernier jour valide du mois cible
        d = min(today.day, calendar.monthrange(y, m)[1])
        return datetime(y, m, d).date().isoformat()

    def _pick_groups(self):
        """Fenêtre de choix des groupes d'accès (partagée par les actions groupées).

        Renvoie la liste des
        URLs sélectionnées, ou None si l'utilisateur annule (fermeture/Annuler)."""
        win = ctk.CTkToplevel(self)
        win.title("Restreindre à des groupes")
        win.geometry("380x440")
        win.transient(self)
        win.grab_set()                       # modale : bloque la fenêtre principale
        ctk.CTkLabel(win, text="Cochez le(s) groupe(s) autorisé(s) :",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(
            anchor="w", padx=16, pady=(16, 6))
        ctk.CTkLabel(win,
                     text="Les vidéos cochées passeront en « Restreint » et ne seront "
                          "visibles que par les membres de ces groupes (connexion requise).",
                     font=ctk.CTkFont(size=11), text_color=T_DISCRET,
                     wraplength=340, justify="left").pack(anchor="w", padx=16, pady=(0, 8))

        scroll = ctk.CTkScrollableFrame(win, height=250, fg_color=S_CARTE)
        scroll.pack(fill="both", expand=True, padx=16)
        vars_by_url = {}
        for g in self.access_groups:
            gurl = g.get("url", "")
            var = ctk.BooleanVar(value=False)
            vars_by_url[gurl] = var
            texte = g.get("display_name") or g.get("code_name", "?")
            ctk.CTkCheckBox(scroll, text=texte, variable=var,
                            font=ctk.CTkFont(size=12)).pack(anchor="w", pady=2)

        result = {"urls": None}              # None = annulé (par défaut)

        def valider():
            """Valide la sélection et ferme la fenêtre."""
            result["urls"] = [u for u, var in vars_by_url.items() if var.get()]
            win.destroy()

        def annuler():
            """Annule et ferme la fenêtre sans rien sélectionner."""
            result["urls"] = None
            win.destroy()

        btns = ctk.CTkFrame(win, fg_color="transparent")
        btns.pack(fill="x", padx=16, pady=14)
        ctk.CTkButton(btns, text="Valider", command=valider,
                      fg_color=C_SUCCES, hover_color=C_SUCCES_SURV).pack(side="left")
        ctk.CTkButton(btns, text="Annuler", command=annuler,
                      fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV, text_color=T_SUR_NEUTRE).pack(side="left", padx=10)

        win.wait_window()                    # bloque jusqu'à fermeture de la modale
        return result["urls"]

    # ═════════════════════════════════════════════════════════════════════
    #  ONGLET INVENTAIRE / STATISTIQUES (+ export Excel)
    # ═════════════════════════════════════════════════════════════════════
    #
    #  Lecture seule : on scanne toutes les vidéos (+ types + chaînes pour
    #  obtenir des libellés lisibles), on calcule des agrégats (totaux, par
    #  utilisateur / type / chaîne) et on peut exporter le tout en .xlsx.

    def _build_tab_stats(self):
        """Construit l'onglet Inventaire / statistiques."""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tabs["stats"] = frame

        ctk.CTkLabel(frame, text="📊  Inventaire / Statistiques",
                     font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(
            frame,
            text="Vue d'ensemble de l'instance : volumétrie, durées, répartition par "
                 "utilisateur, type et chaîne. Export possible vers un classeur Excel.",
            text_color=T_SECONDAIRE, font=ctk.CTkFont(size=12),
            justify="left", wraplength=860).pack(anchor="w", pady=(0, 10))

        # — Ligne d'actions : scan + export —
        top = ctk.CTkFrame(frame, fg_color="transparent")
        top.pack(fill="x")
        ctk.CTkButton(top, text="🔄  Rafraîchir", fg_color=C_NEUTRE,
                      hover_color=C_NEUTRE_SURV, command=self._stats_scan, text_color=T_SUR_NEUTRE).pack(side="left")
        self.stats_export_btn = ctk.CTkButton(
            top, text="📊  Exporter en Excel (.xlsx)", fg_color=C_SUCCES,
            hover_color=C_SUCCES_SURV, state="disabled", command=self._stats_export)
        self.stats_export_btn.pack(side="left", padx=10)
        self.stats_status = ctk.CTkLabel(top, text="(aucun scan)", text_color=T_SECONDAIRE,
                                         font=ctk.CTkFont(size=11))
        self.stats_status.pack(side="left", padx=8)

        # — Bandeau de chiffres clés —
        self.stats_summary = ctk.CTkLabel(frame, text="", justify="left", anchor="w",
                                          font=ctk.CTkFont(size=13))
        self.stats_summary.pack(anchor="w", pady=(10, 6))

        # — Sélecteur de dimension pour la répartition —
        dim = ctk.CTkFrame(frame, fg_color="transparent")
        dim.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(dim, text="Répartition par :").pack(side="left", padx=(0, 4))
        self.stats_dim = ctk.CTkOptionMenu(
            dim, width=210,
            values=["Utilisateur", "Type", "Chaîne",
                    "👁 Vues — vidéos", "👁 Vues — chaînes",
                    "👁 Vues — utilisateurs", "👁 Vues — par mois"],
            command=lambda _c: self._render_stats(), **STYLE_CHAMP)
        self.stats_dim.set("Utilisateur")
        self.stats_dim.pack(side="left")

        # — Tableau de répartition —
        self.stats_table = ctk.CTkScrollableFrame(frame, label_text="Répartition", fg_color=S_CARTE, label_anchor="w",
                                                  label_font=ctk.CTkFont(size=12, weight="bold"))
        self.stats_table.pack(fill="both", expand=True, pady=(4, 0))

        # Données calculées (remplies après un scan)
        self.stats_videos = []     # scan complet
        self.stats_data = None     # agrégats (dict) ; None tant qu'aucun scan

    # ── Scan + calcul ──────────────────────────────────────────────────────

    def _stats_scan(self):
        """(Thread) Parcourt les vidéos et calcule les statistiques."""
        if not self.api:
            self.stats_status.configure(text="Connectez-vous d'abord.", text_color=T_ALERTE)
            return
        self.stats_status.configure(text="⏳  Scan en cours…", text_color=T_SECONDAIRE)
        self.stats_export_btn.configure(state="disabled")
        self._run(self._do_stats_scan)

    def _do_stats_scan(self):
        """(Thread) Récupère vidéos + types + chaînes, puis calcule les agrégats."""
        try:
            def prog(n):
                """Callback de progression (avancement du scan)."""
                self._ui(self.stats_status.configure,
                         text=f"⏳  {n} vidéos lues…", text_color=T_SECONDAIRE)
            # Magasin partagé : sans cela l'Inventaire affichait des totaux
            # périmés sous une mention « chargé à … » qui inspirait confiance.
            videos = self.ensure_videos_sync(progress_cb=prog)

            # Cartes URL → libellé pour afficher des noms plutôt que des URLs
            try:
                type_by_url = {str(t.get("url", "")).rstrip("/"): t.get("title", "?")
                               for t in self.api.get_types()}
            except Exception:
                type_by_url = {}
            try:
                chan_by_url = {str(c.get("url", "")).rstrip("/"): c.get("title", "?")
                               for c in self.api.get_channels()}
            except Exception:
                chan_by_url = {}
            # Carte URL utilisateur → username (à partir des comptes déjà chargés)
            user_by_url = {str(u.get("url", "")).rstrip("/"): u.get("username", "?")
                           for u in (self.all_users or [])}

            self.stats_videos = videos
            # Consultations : détail jour par jour, par vidéo. Une erreur ici ne
            # doit pas priver l'utilisateur de tout l'inventaire — on continue
            # sans les vues plutôt que d'échouer.
            try:
                self._ui(self.stats_status.configure,
                         text="⏳  Lecture des consultations…", text_color=T_SECONDAIRE)
                vues = self.api.get_view_counts()
            except Exception as e:
                vues = []
                self._ui(self._log, f"Consultations indisponibles : {e}")

            # On CONSERVE les tables de correspondance et les consultations :
            # elles permettent de recalculer les agrégats après une suppression
            # ou une modification, sans relire l'instance.
            self._stats_contexte = (user_by_url, type_by_url, chan_by_url, vues)
            self.stats_data = self._compute_stats(videos, user_by_url,
                                                  type_by_url, chan_by_url, vues)
            self._ui(self._render_stats)
            # L'Inventaire est le plus exposé au scan tronqué : ses totaux
            # seraient faux sans que rien ne le signale.
            alerte = self.scan_truncated_warning()
            self._ui(self.stats_status.configure,
                     text=(alerte if alerte else f"✅  {len(videos)} vidéos analysées."),
                     text_color=T_ERREUR if alerte else "#22c55e")
            self._ui(self.stats_export_btn.configure, state="normal")
            self._ui(self._log,
                     alerte if alerte else f"Inventaire : {len(videos)} vidéos analysées.")
        except Exception as e:
            self._ui(self.stats_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Inventaire : {e}")

    @staticmethod
    def _fmt_duration(seconds) -> str:
        """Formate une durée en secondes → « 1h05m09s » ou « 5m09s »."""
        s = int(seconds or 0)
        h, m, sec = s // 3600, (s % 3600) // 60, s % 60
        return f"{h}h{m:02d}m{sec:02d}s" if h else f"{m}m{sec:02d}s"

    @staticmethod
    def _label_from_ref(ref, by_url, fallback="—"):
        """Libellé lisible d'une relation, qu'elle soit un dict, une URL ou autre."""
        if isinstance(ref, dict):
            return ref.get("title") or ref.get("username") or ref.get("url") or fallback
        if isinstance(ref, str):
            return by_url.get(ref.rstrip("/"), ref) if ref.startswith("http") else ref
        return fallback

    @staticmethod
    def _compute_views(videos, vues, user_by_url, chan_by_url) -> dict:
        """Agrège les consultations : classement, évolution, chaînes, propriétaires.

        Pod enregistre une entrée par vidéo ET PAR JOUR. On en tire :
          • `total`        — nombre total de vues ;
          • `top`          — les vidéos les plus vues, avec propriétaire ;
          • `par_mois`     — total par mois, pour suivre l'évolution ;
          • `par_chaine`   — total par chaîne ;
          • `par_proprio`  — total par propriétaire ;
          • `periode`      — première et dernière date connues.

        `periode` est important : il indique la PROFONDEUR d'historique réellement
        conservée par l'instance. Si elle se limite à quelques semaines, une
        courbe d'évolution longue n'a pas de sens.

        Fonction pure (sans réseau ni interface) : testable directement.
        """
        from collections import defaultdict

        # Index des vidéos par identifiant, pour rattacher chaque vue.
        def cle(valeur):
            """Extrait l'identifiant d'une vidéo depuis une URL ou un objet."""
            if isinstance(valeur, dict):
                valeur = valeur.get("url", "")
            return str(valeur or "").rstrip("/").rsplit("/", 1)[-1]

        infos = {}
        for v in videos:
            infos[cle(v.get("url") or v.get("id"))] = v

        total = 0
        par_video = defaultdict(int)
        par_mois = defaultdict(int)
        dates = []

        for e in (vues or []):
            n = int(e.get("count") or 0)
            if n <= 0:
                continue
            total += n
            par_video[cle(e.get("video"))] += n
            d = str(e.get("date") or "")
            if len(d) >= 7:
                par_mois[d[:7]] += n        # regroupement AAAA-MM
                dates.append(d[:10])

        # Classement des vidéos, enrichi du titre et du propriétaire.
        top = []
        for vid, n in sorted(par_video.items(), key=lambda x: -x[1]):
            v = infos.get(vid)
            if not v:
                continue                    # vidéo supprimée depuis
            proprio = v.get("owner")
            if isinstance(proprio, dict):
                proprio = proprio.get("url", "")
            top.append({
                "titre": (v.get("title") or "(sans titre)"),
                "vues": n,
                "proprio": user_by_url.get(str(proprio or "").rstrip("/"), "?"),
            })

        # Agrégats par chaîne et par propriétaire.
        par_chaine = defaultdict(int)
        par_proprio = defaultdict(int)
        for vid, n in par_video.items():
            v = infos.get(vid)
            if not v:
                continue
            proprio = v.get("owner")
            if isinstance(proprio, dict):
                proprio = proprio.get("url", "")
            par_proprio[user_by_url.get(str(proprio or "").rstrip("/"), "?")] += n
            chans = v.get("channel") or []
            if isinstance(chans, (str, dict)):
                chans = [chans]
            if not chans:
                par_chaine["(hors chaîne)"] += n
            for c in chans:
                if isinstance(c, dict):
                    c = c.get("url", "")
                par_chaine[chan_by_url.get(str(c or "").rstrip("/"), "?")] += n

        # NOTE : `total` compte TOUTES les vues enregistrées, y compris celles de
        # vidéos supprimées depuis. Le classement `top`, lui, n'affiche que les
        # vidéos encore présentes. La somme du classement peut donc être
        # inférieure au total — c'est voulu : le total reflète l'activité réelle
        # de la plateforme, le classement ne montre que ce qui est consultable.
        return {
            "total": total,
            "videos_vues": len(par_video),
            "top": top,
            "par_mois": dict(sorted(par_mois.items())),
            "par_chaine": dict(par_chaine),
            "par_proprio": dict(par_proprio),
            "periode": (min(dates), max(dates)) if dates else ("", ""),
        }

    def _compute_stats(self, videos, user_by_url, type_by_url, chan_by_url,
                       vues=None) -> dict:
        """Calcule tous les agrégats (logique pure, testable sans interface).

        `vues` : entrées de consultation {video, date, count}. Facultatif — sans
        elles, l'inventaire garde son contenu d'origine et les blocs de
        consultation restent simplement vides."""
        from collections import defaultdict
        total = len(videos)
        total_dur = sum(int(v.get("duration") or 0) for v in videos)
        drafts = sum(1 for v in videos if v.get("is_draft"))
        unencoded = sum(1 for v in videos if PodAPI.is_unencoded(v))
        restricted = sum(1 for v in videos if v.get("is_restricted"))

        by_owner = defaultdict(lambda: [0, 0])   # nom → [nombre, durée totale]
        by_type = defaultdict(int)               # type → nombre
        by_chan = defaultdict(int)               # chaîne → nombre
        rows = []                                # inventaire détaillé (pour l'export)

        for v in videos:
            owner = self._label_from_ref(v.get("owner"), user_by_url)
            dur = int(v.get("duration") or 0)
            by_owner[owner][0] += 1
            by_owner[owner][1] += dur

            tlabel = self._label_from_ref(v.get("type"), type_by_url)
            by_type[tlabel] += 1

            # channel = liste d'URLs (M2M) ; gérer aussi le cas chaîne unique / vide
            chans = v.get("channel") or []
            if isinstance(chans, str):
                chans = [chans]
            if not chans:
                by_chan["(aucune)"] += 1
            for c in chans:
                by_chan[self._label_from_ref(c, chan_by_url)] += 1

            rows.append({
                "Titre": v.get("title", ""),
                "Slug": v.get("slug", ""),
                "Propriétaire": owner,
                "Type": tlabel,
                "Brouillon": "oui" if v.get("is_draft") else "non",
                "Encodée": "oui" if v.get("encoded") else "non",
                "Restreinte": "oui" if v.get("is_restricted") else "non",
                "Durée (s)": dur,
                "Durée": self._fmt_duration(dur),
                "Ajoutée le": str(v.get("date_added", ""))[:10],
            })

        agregats_vues = self._compute_views(videos, vues or [], user_by_url, chan_by_url)

        return {
            "total": total, "total_dur": total_dur, "drafts": drafts,
            "unencoded": unencoded, "restricted": restricted,
            "by_owner": dict(by_owner), "by_type": dict(by_type),
            "by_chan": dict(by_chan), "rows": rows,
            "vues": agregats_vues,
        }

    # ── Rendu ──────────────────────────────────────────────────────────────

    def _recalculer_stats(self):
        """Recalcule les agrégats de l'Inventaire à partir du magasin partagé.

        Appelée après une suppression ou une modification : sans elle,
        l'Inventaire affichait des totaux périmés sous une mention
        « chargé à … » qui inspirait confiance — le pire cas pour un outil dont
        c'est précisément la fonction.

        Aucun appel réseau : on relit le magasin, déjà à jour, et les tables de
        correspondance mémorisées au dernier scan."""
        contexte = getattr(self, "_stats_contexte", None)
        if not contexte:
            return                        # aucun scan encore effectué
        user_by_url, type_by_url, chan_by_url, vues = contexte
        self.stats_videos = list(self.videos)
        self.stats_data = self._compute_stats(self.stats_videos, user_by_url,
                                              type_by_url, chan_by_url, vues)
        self._render_stats()

    def _render_stats(self):
        d = self.stats_data
        if not d:
            return
        # Chiffres clés
        vues = d.get("vues") or {}
        ligne_vues = ""
        if vues.get("total"):
            debut, fin = vues.get("periode", ("", ""))
            periode = f"  ·  du {debut} au {fin}" if debut else ""
            ligne_vues = (f"\n👁  {vues['total']} vue(s) sur "
                          f"{vues['videos_vues']} vidéo(s){periode}")
        self.stats_summary.configure(
            text=(f"📁  {d['total']} vidéos       "
                  f"⏱  {self._fmt_duration(d['total_dur'])} au total\n"
                  f"📝  {d['drafts']} brouillon(s)     "
                  f"⚙️  {d['unencoded']} non-encodée(s)     "
                  f"🔒  {d['restricted']} restreinte(s)"
                  + ligne_vues))

        # Tableau de répartition selon la dimension choisie
        for w in self.stats_table.winfo_children():
            w.destroy()

        dimension = self.stats_dim.get()

        # ── Dimensions de CONSULTATION ────────────────────────────────────
        if dimension.startswith("👁"):
            if not vues.get("total"):
                ctk.CTkLabel(self.stats_table,
                             text="Aucune consultation enregistrée pour l'instant.\n"
                                  "Les vues apparaîtront ici une fois la plateforme "
                                  "ouverte aux utilisateurs.",
                             text_color=T_SECONDAIRE, justify="left").pack(pady=20, padx=10)
                return
            if "vidéos" in dimension:
                self._stats_header(("Vidéo", "Vues", "Propriétaire"))
                for e in vues["top"][:200]:
                    self._stats_row((e["titre"][:60], str(e["vues"]), e["proprio"]))
            elif "chaînes" in dimension:
                self._stats_header(("Chaîne", "Vues", ""))
                for nom, n in sorted(vues["par_chaine"].items(), key=lambda kv: -kv[1]):
                    self._stats_row((nom, str(n), ""))
            elif "utilisateurs" in dimension:
                self._stats_header(("Utilisateur", "Vues", ""))
                for nom, n in sorted(vues["par_proprio"].items(), key=lambda kv: -kv[1]):
                    self._stats_row((nom, str(n), ""))
            else:      # par mois — dans l'ordre chronologique, pas par volume
                self._stats_header(("Mois", "Vues", "Évolution"))
                mois = vues["par_mois"]
                maxi = max(mois.values()) if mois else 1
                for m, n in mois.items():
                    # Petite barre en caractères : lisible sans bibliothèque
                    # graphique, et suffisante pour repérer une tendance.
                    barre = "█" * max(1, round(20 * n / maxi))
                    self._stats_row((m, str(n), barre))
            return

        if dimension == "Utilisateur":
            # [nombre, durée] → on trie par nombre décroissant
            items = sorted(d["by_owner"].items(), key=lambda kv: kv[1][0], reverse=True)
            self._stats_header(("Utilisateur", "Vidéos", "Durée"))
            for name, (cnt, dur) in items:
                self._stats_row((name, str(cnt), self._fmt_duration(dur)))
        else:
            src = d["by_type"] if dimension == "Type" else d["by_chan"]
            items = sorted(src.items(), key=lambda kv: kv[1], reverse=True)
            self._stats_header((dimension, "Vidéos", ""))
            for name, cnt in items:
                self._stats_row((name, str(cnt), ""))

    def _stats_header(self, cols):
        """Ligne d'en-tête du tableau (trois colonnes)."""
        row = ctk.CTkFrame(self.stats_table, fg_color="transparent")
        row.pack(fill="x", pady=(0, 2))
        widths = (360, 80, 120)
        for text, w in zip(cols, widths):
            ctk.CTkLabel(row, text=text, width=w, anchor="w",
                         font=ctk.CTkFont(size=12, weight="bold")).pack(side="left", padx=4)

    def _stats_row(self, cols):
        """Ligne de données du tableau."""
        row = ctk.CTkFrame(self.stats_table, fg_color=S_LIGNE, corner_radius=4)
        row.pack(fill="x", pady=1)
        widths = (360, 80, 120)
        for text, w in zip(cols, widths):
            ctk.CTkLabel(row, text=text, width=w, anchor="w",
                         font=ctk.CTkFont(size=12)).pack(side="left", padx=4, pady=3)

    # ── Export Excel ───────────────────────────────────────────────────────

    def _stats_export(self):
        """Demande où enregistrer puis lance l'export en arrière-plan."""
        if not self.stats_data:
            return
        path = filedialog.asksaveasfilename(
            title="Exporter l'inventaire",
            defaultextension=".xlsx",
            filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile=f"inventaire_pod_{datetime.now():%Y%m%d}.xlsx")
        if not path:        # annulé
            return
        self.stats_status.configure(text="⏳  Export…", text_color=T_SECONDAIRE)
        self._run(self._do_stats_export, path)

    def _do_stats_export(self, path):
        """(Thread) Construit le classeur Excel (plusieurs feuilles) et l'enregistre."""
        try:
            # openpyxl est importé ici pour ne pas bloquer le lancement de l'appli
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
            except ImportError:
                self._ui(self.stats_status.configure,
                         text="❌  Module « openpyxl » manquant (pip install openpyxl).",
                         text_color=T_ERREUR)
                self._ui(self._log, "❌ Export Excel impossible : openpyxl non installé.")
                return

            d = self.stats_data
            wb = Workbook()
            bold = Font(bold=True)

            # Feuille 1 : Résumé (couples libellé / valeur)
            ws = wb.active
            ws.title = "Résumé"
            ws.append(["Indicateur", "Valeur"])
            for c in ws[1]:
                c.font = bold
            ws.append(["Vidéos (total)", d["total"]])
            ws.append(["Durée totale (s)", d["total_dur"]])
            ws.append(["Durée totale", self._fmt_duration(d["total_dur"])])
            ws.append(["Brouillons", d["drafts"]])
            ws.append(["Non-encodées", d["unencoded"]])
            ws.append(["Restreintes", d["restricted"]])
            vues = d.get("vues") or {}
            if vues.get("total"):
                debut, fin = vues.get("periode", ("", ""))
                ws.append(["Vues (total)", vues["total"]])
                ws.append(["Vidéos consultées", vues["videos_vues"]])
                ws.append(["Période des consultations", f"{debut} → {fin}"])

            # Feuille 2 : Par utilisateur
            ws = wb.create_sheet("Par utilisateur")
            ws.append(["Utilisateur", "Vidéos", "Durée (s)", "Durée"])
            for c in ws[1]:
                c.font = bold
            for name, (cnt, dur) in sorted(d["by_owner"].items(),
                                           key=lambda kv: kv[1][0], reverse=True):
                ws.append([name, cnt, dur, self._fmt_duration(dur)])

            # Feuille 3 : Par type
            ws = wb.create_sheet("Par type")
            ws.append(["Type", "Vidéos"])
            for c in ws[1]:
                c.font = bold
            for name, cnt in sorted(d["by_type"].items(), key=lambda kv: kv[1], reverse=True):
                ws.append([name, cnt])

            # Feuille 4 : Par chaîne
            ws = wb.create_sheet("Par chaîne")
            ws.append(["Chaîne", "Vidéos"])
            for c in ws[1]:
                c.font = bold
            for name, cnt in sorted(d["by_chan"].items(), key=lambda kv: kv[1], reverse=True):
                ws.append([name, cnt])

            # Feuille 5 : Inventaire détaillé (toutes les vidéos)
            ws = wb.create_sheet("Inventaire")
            if d["rows"]:
                headers = list(d["rows"][0].keys())
                ws.append(headers)
                for c in ws[1]:
                    c.font = bold
                for r in d["rows"]:
                    ws.append([r[h] for h in headers])

            # ── Feuilles de CONSULTATION (si des vues existent) ──────────────

            vues = d.get("vues") or {}

            if vues.get("total"):

                ws = wb.create_sheet("Vues par vidéo")

                ws.append(["Vidéo", "Vues", "Propriétaire"])

                for c in ws[1]:

                    c.font = bold

                for e in vues["top"]:

                    ws.append([e["titre"], e["vues"], e["proprio"]])


                ws = wb.create_sheet("Vues par chaîne")

                ws.append(["Chaîne", "Vues"])

                for c in ws[1]:

                    c.font = bold

                for nom, n in sorted(vues["par_chaine"].items(), key=lambda kv: -kv[1]):

                    ws.append([nom, n])


                ws = wb.create_sheet("Vues par utilisateur")

                ws.append(["Utilisateur", "Vues"])

                for c in ws[1]:

                    c.font = bold

                for nom, n in sorted(vues["par_proprio"].items(), key=lambda kv: -kv[1]):

                    ws.append([nom, n])


                # Ordre CHRONOLOGIQUE ici (et non par volume) : c'est une évolution.

                ws = wb.create_sheet("Vues par mois")

                ws.append(["Mois", "Vues"])

                for c in ws[1]:

                    c.font = bold

                for mois, n in vues["par_mois"].items():

                    ws.append([mois, n])


            wb.save(path)
            self._ui(self.stats_status.configure,
                     text=f"✅  Exporté : {os.path.basename(path)}", text_color=T_SUCCES)
            self._ui(self._log, f"Inventaire exporté → {path}")
        except Exception as e:
            self._ui(self.stats_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Export inventaire : {e}")

    # ═════════════════════════════════════════════════════════════════════
    #  ONGLET CHAÎNES & THÈMES (création / modification / suppression)
    # ═════════════════════════════════════════════════════════════════════
    #
    #  Champs requis relevés au diagnostic :
    #    • Chaîne : title + themes (liste, vide acceptée)
    #    • Thème  : title + channel (URL)  ; hiérarchie possible via parentId
    #  Les thèmes sont regroupés sous leur chaîne grâce à leur champ `channel`.

    def _build_tab_ct(self):
        """Construit l'onglet Chaînes & thèmes."""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tabs["ct"] = frame

        ctk.CTkLabel(frame, text="📺  Chaînes & thèmes",
                     font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(
            frame,
            text="Gérer les chaînes et leurs thèmes : créer, renommer, basculer la visibilité, "
                 "supprimer. Les thèmes apparaissent sous leur chaîne.",
            text_color=T_SECONDAIRE, font=ctk.CTkFont(size=12),
            justify="left", wraplength=860).pack(anchor="w", pady=(0, 10))

        # — Ligne d'action : charger —
        top = ctk.CTkFrame(frame, fg_color="transparent")
        top.pack(fill="x")
        ctk.CTkButton(top, text="🔄  Rafraîchir", fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=self._ct_load, text_color=T_SUR_NEUTRE).pack(side="left")
        self.ct_status = ctk.CTkLabel(top, text="(en attente de connexion…)", text_color=T_SECONDAIRE,
                                      font=ctk.CTkFont(size=11))
        self.ct_status.pack(side="left", padx=10)

        # — Liste des chaînes + thèmes —
        self.ct_list = ctk.CTkScrollableFrame(frame, label_text="Chaînes et thèmes", fg_color=S_CARTE, label_anchor="w",
                                                  label_font=ctk.CTkFont(size=12, weight="bold"))
        self.ct_list.pack(fill="both", expand=True, pady=(8, 6))

        # — Formulaire : nouvelle chaîne —
        cform = ctk.CTkFrame(frame, fg_color=S_CARTE)
        cform.pack(fill="x", pady=(2, 4))
        ctk.CTkLabel(cform, text="Nouvelle chaîne :",
                     font=ctk.CTkFont(weight="bold")).pack(side="left", padx=(10, 6))
        self.ct_new_chan_title = ctk.CTkEntry(cform, placeholder_text="titre", width=180)
        self.ct_new_chan_title.pack(side="left", padx=4)
        self.ct_new_chan_desc = ctk.CTkEntry(cform, placeholder_text="description (option)", width=200)
        self.ct_new_chan_desc.pack(side="left", padx=4)
        self.ct_new_chan_visible = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(cform, text="visible", variable=self.ct_new_chan_visible).pack(side="left", padx=6)
        ctk.CTkButton(cform, text="＋ Créer", width=90, fg_color=C_SUCCES,
                      hover_color=C_SUCCES_SURV, command=self._ct_create_channel).pack(side="left", padx=6, pady=6)

        # — Formulaire : nouveau thème —
        tform = ctk.CTkFrame(frame, fg_color=S_CARTE)
        tform.pack(fill="x", pady=(2, 0))
        ctk.CTkLabel(tform, text="Nouveau thème :",
                     font=ctk.CTkFont(weight="bold")).pack(side="left", padx=(10, 6))
        self.ct_new_theme_title = ctk.CTkEntry(tform, placeholder_text="titre", width=180)
        self.ct_new_theme_title.pack(side="left", padx=4)
        ctk.CTkLabel(tform, text="dans la chaîne :").pack(side="left", padx=(6, 4))
        # Menu des chaînes cibles (rempli après chargement)
        self.ct_theme_channel = ctk.CTkOptionMenu(tform, width=200, values=["(charger d'abord)"], **STYLE_CHAMP)
        self.ct_theme_channel.pack(side="left", padx=4)
        ctk.CTkButton(tform, text="＋ Créer", width=90, fg_color=C_SUCCES,
                      hover_color=C_SUCCES_SURV, command=self._ct_create_theme).pack(side="left", padx=6, pady=6)

        # Données
        self.ct_channels = []           # liste de chaînes (dicts)
        self.ct_themes = []             # liste de thèmes (dicts)
        # ALIAS du magasin partagé `self.videos` (même objet liste), conservé le
        # temps de la migration pour ne rien casser si un code résiduel y accède.
        self.ct_videos = self.videos
        self.ct_channel_choices = {}    # titre de chaîne → URL (pour le menu thème)

    # ── Chargement ─────────────────────────────────────────────────────────

    def _ct_load(self):
        """(Thread) Charge chaînes et thèmes pour l'onglet Chaînes."""
        if not self.api:
            self.ct_status.configure(text="Connectez-vous d'abord.", text_color=T_ALERTE)
            return
        self.ct_status.configure(text="⏳  Chargement…", text_color=T_SECONDAIRE)
        self._run(self._do_ct_load)

    def _do_ct_load(self):
        """(Thread) Récupère chaînes et thèmes, puis rafraîchit l'affichage.
        Peut être appelé depuis un autre thread (création/édition) pour recharger."""
        try:
            chans = self.api.get_channels()
            themes = self.api.get_themes()
            # Tri alphabétique pour un affichage stable
            self.ct_channels = sorted(chans, key=lambda c: (c.get("title") or "").lower())
            self.ct_themes = themes
            # Carte titre → URL pour le menu de création de thème
            self.ct_channel_choices = {c.get("title", "?"): c.get("url")
                                       for c in self.ct_channels}
            self._ui(self._render_ct)
            self._ui(self._refresh_ct_channel_menu)
            # La liste des chaînes vient de changer (création, renommage,
            # suppression, rechargement) : les menus de filtre « chaîne » des
            # onglets Vidéos et Explorateur doivent suivre. Point d'accroche
            # unique, puisque toutes ces opérations repassent par ici.
            self._ui(self.schedule_refresh, channels=True)
            self._ui(self.ct_status.configure,
                     text=f"✅  {len(chans)} chaîne(s), {len(themes)} thème(s)  ·  "
                          f"{self._loaded_stamp()}",
                     text_color=T_SUCCES)
        except Exception as e:
            self._ui(self.ct_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Chargement chaînes/thèmes : {e}")

    def _refresh_ct_channel_menu(self):
        """Met à jour la liste déroulante des chaînes (création de thème)."""
        titles = list(self.ct_channel_choices.keys())
        if titles:
            self.ct_theme_channel.configure(values=titles)
            self.ct_theme_channel.set(titles[0])
        else:
            self.ct_theme_channel.configure(values=["(aucune chaîne)"])
            self.ct_theme_channel.set("(aucune chaîne)")

    # ── Rendu de la liste (chaînes + thèmes imbriqués) ─────────────────────

    def _render_ct(self):
        """Affiche la liste des chaînes et de leurs thèmes."""
        for w in self.ct_list.winfo_children():
            w.destroy()

        if not self.ct_channels:
            ctk.CTkLabel(self.ct_list,
                         text="Aucune chaîne. Cliquez « Charger » ou créez-en une ci-dessous.",
                         text_color=T_SECONDAIRE).pack(pady=10)
            return

        # Regrouper les thèmes par URL de chaîne (champ `channel` du thème)
        from collections import defaultdict
        themes_by_chan = defaultdict(list)
        for t in self.ct_themes:
            themes_by_chan[str(t.get("channel", "")).rstrip("/")].append(t)

        for ch in self.ct_channels:
            curl = str(ch.get("url", "")).rstrip("/")

            # — Ligne de la chaîne —
            crow = ctk.CTkFrame(self.ct_list, fg_color=S_LIGNE, corner_radius=6)
            crow.pack(fill="x", pady=(6, 0))
            vis = "👁" if ch.get("visible") else "🚫"
            ctk.CTkLabel(crow, text=f"{vis}  {ch.get('title', '(sans titre)')}",
                         anchor="w", font=ctk.CTkFont(size=13, weight="bold")).pack(
                side="left", padx=10, pady=6, fill="x", expand=True)
            # Actions de la chaîne
            ctk.CTkButton(crow, text="✏", width=34, fg_color=C_NEUTRE,
                          command=lambda c=ch: self._ct_rename_channel(c), text_color=T_SUR_NEUTRE).pack(side="left", padx=2)
            ctk.CTkButton(crow, text="🎬 Vidéos", width=80, fg_color=C_NEUTRE,
                          command=lambda c=ch: self._ct_manage_videos(c), text_color=T_SUR_NEUTRE).pack(side="left", padx=2)
            ctk.CTkButton(crow, text="👤 Admins", width=80, fg_color=C_NEUTRE,
                          command=lambda c=ch: self._ct_manage_owners(c), text_color=T_SUR_NEUTRE).pack(side="left", padx=2)
            ctk.CTkButton(crow, text="🔒 Restreindre", width=100, fg_color=C_NEUTRE,
                          command=lambda c=ch: self._ct_manage_groups(c), text_color=T_SUR_NEUTRE).pack(side="left", padx=2)
            ctk.CTkButton(crow, text="🎨 Habillage", width=94, fg_color=C_ACCENT,
                          hover_color=C_ACCENT_SURV,
                          command=lambda c=ch: self._ct_habillage(c, "channel")).pack(
                side="left", padx=2)
            ctk.CTkButton(crow, text="👁/🚫", width=54, fg_color=C_NEUTRE,
                          command=lambda c=ch: self._ct_toggle_visible(c), text_color=T_SUR_NEUTRE).pack(side="left", padx=2)
            # La suppression est IRRÉVERSIBLE et se trouvait collée au
            # basculement de visibilité — une action anodine, manipulée
            # souvent. On l'isole derrière un séparateur, en fin de rangée.
            #
            # Elle portait aussi un libellé, au motif qu'une icône seule
            # n'annonce pas la gravité. L'argument reste juste isolément, mais
            # répété sur chaque ligne le libellé rouge saturait l'écran. Le
            # séparateur, l'infobulle au survol, le rouge qui n'apparaît qu'au
            # survol et la confirmation obligatoire portent désormais cette
            # gravité à eux quatre.
            ctk.CTkFrame(crow, width=1, height=22,
                         fg_color=S_PUCE).pack(side="left", padx=8)
            # Icône seule, neutre, ROUGE AU SURVOL.
            #
            # Le libellé était là pour annoncer la gravité, mais répété sur
            # chaque ligne il produisait un mur rouge : sur vingt chaînes,
            # l'action la plus dangereuse devenait la plus visible, et l'œil
            # s'y habituait — l'inverse de l'effet recherché. Le mot revient
            # au survol, et la confirmation reste inchangée.
            btn = ctk.CTkButton(crow, text="🗑", width=34,
                                fg_color=C_NEUTRE, hover_color=C_DESTRUCTIF,
                                text_color=T_SUR_NEUTRE,
                                font=ctk.CTkFont(size=T_CORPS),
                                command=lambda c=ch: self._ct_delete_channel(c))
            btn.pack(side="left", padx=(0, 8))
            ajouter_infobulle(btn, "Supprimer cette chaîne")

            # — Thèmes de cette chaîne (indentés) —
            for t in themes_by_chan.get(curl, []):
                trow = ctk.CTkFrame(self.ct_list, fg_color="transparent")
                trow.pack(fill="x", padx=(28, 0))
                ctk.CTkLabel(trow, text=f"└  {t.get('title', '(sans titre)')}",
                             anchor="w", font=ctk.CTkFont(size=12)).pack(
                    side="left", padx=6, pady=2, fill="x", expand=True)
                ctk.CTkButton(trow, text="✏", width=34, height=24,
                              fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                              text_color=T_SUR_NEUTRE,
                              command=lambda th=t: self._ct_rename_theme(th)).pack(side="left", padx=2)
                ctk.CTkButton(trow, text="🎨", width=34, height=24, fg_color=C_ACCENT,
                              hover_color=C_ACCENT_SURV,
                              command=lambda th=t: self._ct_habillage(th, "theme")).pack(
                    side="left", padx=2)
                ctk.CTkFrame(trow, width=1, height=18,
                             fg_color=S_PUCE).pack(side="left", padx=6)
                btn_t = ctk.CTkButton(trow, text="🗑", width=32, height=24,
                                      font=ctk.CTkFont(size=T_PETIT),
                                      fg_color=C_NEUTRE, hover_color=C_DESTRUCTIF,
                                      text_color=T_SUR_NEUTRE,
                                      command=lambda th=t: self._ct_delete_theme(th))
                btn_t.pack(side="left", padx=(2, 8))
                ajouter_infobulle(btn_t, "Supprimer ce thème")

    # ── Création ───────────────────────────────────────────────────────────

    def _ct_create_channel(self):
        """Crée une nouvelle chaîne à partir du formulaire."""
        title = self.ct_new_chan_title.get().strip()
        if not title:
            self.ct_status.configure(text="Titre de chaîne requis.", text_color=T_ALERTE)
            return
        desc = self.ct_new_chan_desc.get().strip()
        visible = self.ct_new_chan_visible.get()
        self._run(self._do_ct_create_channel, title, desc, visible)

    def _do_ct_create_channel(self, title, desc, visible):
        try:
            # themes=[] : le champ est requis par l'API mais peut être vide
            self.api.create_channel(title, theme_urls=[], description=desc, visible=visible)
            self._ui(self._log, f"Chaîne créée : {title}")
            self._ui(self._ct_clear_new_channel)
            self._do_ct_load()      # recharge (on est déjà dans un thread)
        except Exception as e:
            self._ui(self.ct_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Création chaîne : {e}")

    def _ct_clear_new_channel(self):
        """Vide le formulaire de création de chaîne."""
        self.ct_new_chan_title.delete(0, "end")
        self.ct_new_chan_desc.delete(0, "end")

    def _ct_create_theme(self):
        """Crée un thème dans la chaîne sélectionnée."""
        title = self.ct_new_theme_title.get().strip()
        if not title:
            self.ct_status.configure(text="Titre de thème requis.", text_color=T_ALERTE)
            return
        channel_url = self.ct_channel_choices.get(self.ct_theme_channel.get())
        if not channel_url:
            self.ct_status.configure(text="Choisissez une chaîne pour le thème.",
                                     text_color=T_ALERTE)
            return
        self._run(self._do_ct_create_theme, title, channel_url)

    def _do_ct_create_theme(self, title, channel_url):
        """(Thread) Crée le thème côté serveur puis rafraîchit."""
        try:
            self.api.create_theme(title, channel_url)
            self._ui(self._log, f"Thème créé : {title}")
            self._ui(lambda: self.ct_new_theme_title.delete(0, "end"))
            self._do_ct_load()
        except Exception as e:
            self._ui(self.ct_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Création thème : {e}")

    # ── Modification ───────────────────────────────────────────────────────

    # ── Habillage (couleur, description, bannière) ─────────────────────────

    def _ct_habillage(self, element, genre: str):
        """Ouvre la fenêtre d'habillage d'une chaîne ou d'un thème.

        `genre` vaut "channel" ou "theme". Les deux partagent titre, description
        et bannière ; seule une chaîne possède en plus une couleur et un état de
        visibilité (vérifié par sonde sur l'API de l'instance).

        Le champ `style` (CSS libre) n'est volontairement PAS exposé : une
        erreur de syntaxe s'appliquerait à l'affichage public de la chaîne.
        """
        est_chaine = (genre == "channel")
        libelle = "chaîne" if est_chaine else "thème"

        win = ctk.CTkToplevel(self)
        win.title(f"Habillage — {element.get('title', '')}")
        win.geometry("560x560" if est_chaine else "560x470")
        _focus_toplevel(win, self)

        ctk.CTkLabel(win, text=f"🎨  Habillage de la {libelle}",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(
            anchor="w", padx=18, pady=(16, 2))
        ctk.CTkLabel(win, text=element.get("title", ""), text_color=T_SECONDAIRE,
                     font=ctk.CTkFont(size=12)).pack(anchor="w", padx=18, pady=(0, 10))

        corps = ctk.CTkFrame(win, fg_color=S_CARTE)
        corps.pack(fill="both", expand=True, padx=16, pady=(0, 8))
        corps.columnconfigure(1, weight=1)
        ligne = 0

        # — Titre —
        ctk.CTkLabel(corps, text="Titre :", width=110, anchor="e").grid(
            row=ligne, column=0, padx=8, pady=8)
        titre_entry = ctk.CTkEntry(corps)
        titre_entry.insert(0, element.get("title", "") or "")
        titre_entry.grid(row=ligne, column=1, columnspan=2, sticky="ew", padx=(0, 8), pady=8)
        ligne += 1

        # — Description —
        ctk.CTkLabel(corps, text="Description :", width=110, anchor="ne").grid(
            row=ligne, column=0, padx=8, pady=8, sticky="ne")
        desc_box = ctk.CTkTextbox(corps, height=90)
        desc_box.insert("1.0", element.get("description", "") or "")
        desc_box.grid(row=ligne, column=1, columnspan=2, sticky="ew", padx=(0, 8), pady=8)
        ligne += 1

        # — Couleur (chaînes uniquement) —
        couleur_entry = None
        apercu_couleur = None
        if est_chaine:
            ctk.CTkLabel(corps, text="Couleur :", width=110, anchor="e").grid(
                row=ligne, column=0, padx=8, pady=8)
            zone = ctk.CTkFrame(corps, fg_color="transparent")
            zone.grid(row=ligne, column=1, columnspan=2, sticky="ew", padx=(0, 8), pady=8)
            couleur_entry = ctk.CTkEntry(zone, width=120, placeholder_text="ex. 223333")
            couleur_entry.insert(0, (element.get("color") or "").lstrip("#"))
            couleur_entry.pack(side="left")
            apercu_couleur = ctk.CTkLabel(zone, text="   ", width=40, height=26,
                                          corner_radius=4)
            apercu_couleur.pack(side="left", padx=8)

            def rafraichir_couleur(*_):
                """Montre la couleur saisie, pour éviter les codes erronés."""
                val = (couleur_entry.get() or "").strip().lstrip("#")
                try:
                    if len(val) in (3, 6) and all(c in "0123456789abcdefABCDEF" for c in val):
                        apercu_couleur.configure(fg_color=f"#{val}")
                    else:
                        apercu_couleur.configure(fg_color=S_PUCE)
                except Exception:
                    pass
            couleur_entry.bind("<KeyRelease>", rafraichir_couleur)
            rafraichir_couleur()
            ctk.CTkButton(zone, text="Choisir…", width=90, fg_color=C_NEUTRE,
                          command=lambda: self._ct_choisir_couleur(couleur_entry,
                                                                   rafraichir_couleur), text_color=T_SUR_NEUTRE).pack(side="left")
            ligne += 1

        # — Bannière —
        ctk.CTkLabel(corps, text="Bannière :", width=110, anchor="e").grid(
            row=ligne, column=0, padx=8, pady=8)
        banniere = {"url": element.get("headband") or ""}
        banniere_lbl = ctk.CTkLabel(
            corps, anchor="w", font=ctk.CTkFont(size=11), text_color=T_SECONDAIRE,
            text=("Image définie" if banniere["url"] else "aucune"))
        banniere_lbl.grid(row=ligne, column=1, sticky="w", padx=(0, 6), pady=8)

        def choisir_banniere():
            """Ouvre la galerie et retient l'image choisie."""
            win.grab_release()
            picker = BannerPicker(self, f"Bannière de la {libelle}", banniere["url"])
            self.wait_window(picker)
            if picker.resultat is not None:
                banniere["url"] = picker.resultat
                banniere_lbl.configure(
                    text=("Image définie" if picker.resultat else "aucune (retirée)"),
                    text_color=("#22c55e" if picker.resultat else "#f59e0b"))
            try:
                win.grab_set()
            except Exception:
                pass

        ctk.CTkButton(corps, text="Choisir…", width=110,
                      fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=choisir_banniere, text_color=T_SUR_NEUTRE).grid(row=ligne, column=2, padx=8, pady=8)
        ligne += 1

        # — Visibilité (chaînes uniquement) —
        visible_var = None
        if est_chaine:
            visible_var = ctk.BooleanVar(value=bool(element.get("visible")))
            ctk.CTkCheckBox(corps, text="Chaîne visible publiquement",
                            variable=visible_var).grid(
                row=ligne, column=1, columnspan=2, sticky="w", padx=(0, 8), pady=8)
            ligne += 1

        msg = ctk.CTkLabel(win, text="", font=ctk.CTkFont(size=11),
                           wraplength=500, justify="left", anchor="w")
        msg.pack(fill="x", padx=18, pady=(0, 4))

        bas = ctk.CTkFrame(win, fg_color="transparent")
        bas.pack(fill="x", padx=16, pady=(0, 14))

        def enregistrer():
            """Compose le PATCH et l'applique."""
            payload = {
                "title": titre_entry.get().strip(),
                "description": desc_box.get("1.0", "end").strip(),
                "headband": banniere["url"] or None,
            }
            if not payload["title"]:
                msg.configure(text="Le titre ne peut pas être vide.", text_color=T_ALERTE)
                return
            if est_chaine:
                coul = (couleur_entry.get() or "").strip().lstrip("#")
                if coul and not (len(coul) in (3, 6)
                                 and all(c in "0123456789abcdefABCDEF" for c in coul)):
                    msg.configure(
                        text="Couleur invalide : attendu 3 ou 6 caractères hexadécimaux "
                             "(ex. 223333).", text_color=T_ALERTE)
                    return
                payload["color"] = coul
                payload["visible"] = bool(visible_var.get())
            msg.configure(text="⏳ Enregistrement…", text_color=T_SECONDAIRE)
            self._run(self._do_ct_habillage, element, genre, payload, win, msg)

        ctk.CTkButton(bas, text="Enregistrer", width=140, fg_color=C_SUCCES,
                      hover_color=C_SUCCES_SURV, command=enregistrer).pack(side="right")
        ctk.CTkButton(bas, text="Annuler", width=110, fg_color=C_NEUTRE,
                      hover_color=C_NEUTRE_SURV, command=win.destroy, text_color=T_SUR_NEUTRE).pack(side="right", padx=8)

    def _ct_choisir_couleur(self, champ, apres):
        """Ouvre le sélecteur de couleur du système et remplit le champ."""
        from tkinter import colorchooser
        actuel = (champ.get() or "").strip().lstrip("#")
        try:
            initial = f"#{actuel}" if len(actuel) in (3, 6) else "#1f4e79"
            choix = colorchooser.askcolor(color=initial, title="Couleur de la chaîne")
        except Exception:
            choix = (None, None)
        if choix and choix[1]:
            champ.delete(0, "end")
            champ.insert(0, str(choix[1]).lstrip("#"))
            apres()

    def _do_ct_habillage(self, element, genre: str, payload: dict, win, msg):
        """(Thread) Applique l'habillage puis rafraîchit la liste."""
        try:
            if genre == "channel":
                self.api.patch_channel(element.get("url") or element.get("id"), payload)
            else:
                self.api.patch_theme(element.get("url") or element.get("id"), payload)
            element.update({k: v for k, v in payload.items() if v is not None})
            self._ui(self._log,
                     f"🎨 Habillage enregistré : {payload.get('title', '')}.")
            self._ui(win.destroy)
            self._do_ct_load()          # recharge chaînes et thèmes
        except Exception as e:
            self._ui(msg.configure, text=f"❌ {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Habillage : {e}")

    def _ct_rename_channel(self, ch):
        """Renomme une chaîne (boîte de saisie)."""
        dlg = ctk.CTkInputDialog(text=f"Nouveau nom pour « {ch.get('title')} » :",
                                 title="Renommer la chaîne")
        new = dlg.get_input()              # bloque jusqu'à fermeture ; None si annulé
        if new and new.strip():
            self._run(self._do_ct_patch, "channel", ch.get("url"),
                      {"title": new.strip()}, f"Chaîne renommée : {new.strip()}")

    def _ct_rename_theme(self, th):
        """Renomme un thème (boîte de saisie)."""
        dlg = ctk.CTkInputDialog(text=f"Nouveau nom pour « {th.get('title')} » :",
                                 title="Renommer le thème")
        new = dlg.get_input()
        if new and new.strip():
            self._run(self._do_ct_patch, "theme", th.get("url"),
                      {"title": new.strip()}, f"Thème renommé : {new.strip()}")

    def _ct_manage_videos(self, ch):
        """Ouvre la gestion des vidéos d'une chaîne. Si la chaîne a des thèmes,
        propose de gérer la chaîne entière OU l'un de ses thèmes (Option 1).
        Scanne les vidéos au besoin (en arrière-plan)."""
        if not self.api:
            self.ct_status.configure(text="Connectez-vous d'abord.", text_color=T_ALERTE)
            return
        self.ct_status.configure(text="⏳  Chargement des vidéos…", text_color=T_SECONDAIRE)
        self._run(self._do_ct_prepare_organizer, ch)

    def _do_ct_prepare_organizer(self, ch):
        """(Thread) Scanne les vidéos puis ouvre soit le sélecteur de chaîne
        (si aucun thème), soit le petit menu chaîne/thèmes."""
        try:
            self.ensure_videos_sync()      # magasin partagé (chargé si besoin)
            curl = str(ch.get("url", "")).rstrip("/")
            # Thèmes appartenant à CETTE chaîne (cohérence : on ne propose que ceux-là)
            themes = [t for t in self.ct_themes
                      if str(t.get("channel")).rstrip("/") == curl]
            self._ui(self.ct_status.configure,
                     text=f"{len(self.videos)} vidéos chargées.", text_color=T_SECONDAIRE)
            if themes:
                self._ui(lambda: self._ct_organizer_dialog(ch, themes))
            else:
                # Pas de thème : on va directement au sélecteur de la chaîne entière
                self._ui(lambda: self._ct_open_channel_picker(ch))
        except Exception as e:
            self._ui(self.ct_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Chargement vidéos (chaîne) : {e}")

    def _ct_organizer_dialog(self, ch, themes):
        """Petit menu : gérer les vidéos de la chaîne entière ou d'un de ses thèmes."""
        win = ctk.CTkToplevel(self)
        win.title(f"Organiser « {ch.get('title')} »")
        win.geometry("440x420")
        _focus_toplevel(win, self)

        ctk.CTkLabel(win, text=f"Chaîne : {ch.get('title')}",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(padx=16, pady=(16, 4), anchor="w")
        ctk.CTkLabel(win, text="Choisissez ce que vous voulez organiser :",
                     text_color=T_SECONDAIRE, font=ctk.CTkFont(size=12)).pack(padx=16, anchor="w")

        # La chaîne entière (appartenance vidéo ↔ chaîne)
        row = ctk.CTkFrame(win, fg_color=S_LIGNE, corner_radius=6)
        row.pack(fill="x", padx=16, pady=(12, 4))
        ctk.CTkLabel(row, text="📁  La chaîne entière", anchor="w",
                     font=ctk.CTkFont(size=12)).pack(side="left", padx=10, pady=8, fill="x", expand=True)
        ctk.CTkButton(row, text="Gérer", width=80,
                      fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=lambda: (win.destroy(), self._ct_open_channel_picker(ch)), text_color=T_SUR_NEUTRE).pack(side="right", padx=8)

        # Un bloc par thème de la chaîne
        ctk.CTkLabel(win, text="Thèmes (rubriques de la chaîne) :",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(padx=16, pady=(10, 2), anchor="w")
        holder = ctk.CTkScrollableFrame(win, height=200, fg_color=S_CARTE)
        holder.pack(fill="both", expand=True, padx=16, pady=(0, 12))
        for t in themes:
            r = ctk.CTkFrame(holder, fg_color=S_LIGNE, corner_radius=6)
            r.pack(fill="x", pady=2)
            ctk.CTkLabel(r, text=f"🏷  {t.get('title')}", anchor="w",
                         font=ctk.CTkFont(size=12)).pack(side="left", padx=10, pady=6,
                                                         fill="x", expand=True)
            ctk.CTkButton(r, text="Gérer", width=80,
                          fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                          command=lambda th=t: (win.destroy(),
                                                self._ct_open_theme_picker(ch, th)), text_color=T_SUR_NEUTRE).pack(side="right", padx=8)

    def _videos_in_relation(self, field, url):
        """Dict {slug: titre} des vidéos dont le champ relation (`channel`/`theme`)
        contient l'URL donnée. Sert à pré-cocher le sélecteur."""
        target = str(url).rstrip("/")
        pre = {}
        for v in self.videos:
            rel = v.get(field) or []
            if isinstance(rel, str):
                rel = [rel]
            urls = [str(x.get("url") if isinstance(x, dict) else x).rstrip("/") for x in rel]
            if target in urls:
                pre[v.get("slug")] = v.get("title", "?")
        return pre

    def _ct_open_channel_picker(self, ch):
        """Ouvre le sélecteur pour gérer l'appartenance à la chaîne entière."""
        pre = self._videos_in_relation("channel", ch.get("url"))
        VideoPicker(self, self.videos,
                    on_done=lambda slugs: self._ct_apply_channel_videos(ch, slugs),
                    title=f"Vidéos de « {ch.get('title')} »", preselected=pre)

    def _ct_open_theme_picker(self, ch, theme):
        """Ouvre le sélecteur pour ranger des vidéos dans un thème de la chaîne."""
        pre = self._videos_in_relation("theme", theme.get("url"))
        VideoPicker(self, self.videos,
                    on_done=lambda slugs: self._ct_apply_theme_videos(ch, theme, slugs),
                    title=f"Vidéos du thème « {theme.get('title')} »", preselected=pre)

    def _ct_apply_channel_videos(self, ch, selected_slugs):
        """Callback du sélecteur : applique les ajouts/retraits en arrière-plan."""
        self._run(self._do_ct_apply_channel_videos, ch, set(selected_slugs))

    def _do_ct_apply_channel_videos(self, ch, desired):
        """(Thread) Met le champ `channel` à jour sur chaque vidéo ajoutée ou retirée.
        On préserve les AUTRES chaînes de chaque vidéo (on n'ajoute/retire que celle-ci)."""
        curl = ch.get("url", "")
        curl_n = str(curl).rstrip("/")
        by_slug = {v.get("slug"): v for v in self.videos}

        # Membres actuels de la chaîne (d'après le cache).
        # On passe par _rel_urls : le champ `channel` peut contenir des URLs
        # OU des objets imbriqués selon le sérialiseur de l'instance. Sans cette
        # normalisation, l'appartenance n'était pas détectée quand l'API renvoie
        # des objets → aucun retrait possible et risque de doublons.
        current = set()
        for v in self.videos:
            if curl_n in self._rel_urls(v.get("channel")):
                current.add(v.get("slug"))

        to_add = desired - current        # à rattacher à la chaîne
        to_remove = current - desired     # à détacher de la chaîne
        ok = fail = theme_cleaned = 0

        # URLs des thèmes appartenant à CETTE chaîne (pour purge à la cohérence) :
        # retirer une vidéo de la chaîne doit aussi la retirer des thèmes de
        # cette chaîne, sinon elle resterait dans un thème orphelin (incohérent).
        chan_theme_urls = {str(t.get("url")).rstrip("/")
                           for t in self.ct_themes
                           if curl_n in self._rel_urls(t.get("channel"))}

        for slug in (to_add | to_remove):
            v = by_slug.get(slug)
            if not v:
                continue
            # URLs des chaînes de la vidéo (objets imbriqués convertis en URLs) :
            # ce qui sera renvoyé tel quel dans le PATCH, d'où l'importance de
            # ne jamais y laisser un dictionnaire.
            chans = self._rel_urls(v.get("channel"), normalise=False)
            chans_n = [c.rstrip("/") for c in chans]
            theme_urls = None     # None = ne pas toucher au champ theme

            if slug in to_add and curl_n not in chans_n:
                chans.append(curl)                                     # ajout de cette chaîne
            if slug in to_remove:
                chans = [c for c in chans if c.rstrip("/") != curl_n]  # retrait de la chaîne
                # COHÉRENCE : purger les thèmes de cette chaîne sur la vidéo
                themes = self._rel_urls(v.get("theme"), normalise=False)
                kept = [t for t in themes if t.rstrip("/") not in chan_theme_urls]
                if len(kept) != len(themes):
                    theme_urls = kept            # on devra patcher le champ theme
                    theme_cleaned += 1
            try:
                # PATCH channel (+ theme si purge nécessaire) — préserve le reste
                self.api.assign_video_to_channels(v, chans, theme_urls=theme_urls)
                v["channel"] = chans                                   # MAJ cache local
                sync_payload = {"channel": chans}
                if theme_urls is not None:
                    v["theme"] = theme_urls
                    sync_payload["theme"] = theme_urls
                self._sync_video_caches(slug, sync_payload)
                ok += 1
            except Exception as e:
                fail += 1
                self._ui(self._log, f"❌ {slug} : {e}")

        msg = (f"✅  Chaîne « {ch.get('title')} » : "
               f"+{len(to_add)} / -{len(to_remove)} vidéo(s).")
        if theme_cleaned:
            msg += f"  ({theme_cleaned} retirée(s) aussi des thèmes pour cohérence.)"
        self._ui(self.ct_status.configure,
                 text=msg, text_color=T_SUCCES if not fail else "#f59e0b")
        self._ui(self._log,
                 f"Chaîne « {ch.get('title')} » : {len(to_add)} ajout(s), "
                 f"{len(to_remove)} retrait(s), {theme_cleaned} purgé(s) des thèmes, "
                 f"{fail} échec(s).")
        # Rafraîchir l'affichage : sans cela, la liste garderait l'état d'avant
        # l'action (nombre de vidéos par chaîne, etc.) jusqu'au prochain
        # « Actualiser » manuel.
        self._ui(self._render_ct)

    def _ct_apply_theme_videos(self, ch, theme, selected_slugs):
        """Callback du sélecteur de thème : applique les changements en arrière-plan."""
        self._run(self._do_ct_apply_theme_videos, ch, theme, set(selected_slugs))

    def _do_ct_apply_theme_videos(self, ch, theme, desired):
        """(Thread) Range les vidéos dans un thème (champ `theme`).
        GARDE-FOU DE COHÉRENCE : l'API n'empêche pas de mettre une vidéo dans un
        thème sans qu'elle soit dans la chaîne parente. On le corrige donc nous-
        mêmes : toute vidéo ajoutée à un thème est AUSSI ajoutée à la chaîne
        parente du thème si elle n'y est pas déjà. Le retrait d'un thème ne
        touche pas la chaîne (la vidéo reste dans la chaîne, hors rubrique)."""
        turl = theme.get("url", "")
        turl_n = str(turl).rstrip("/")
        curl = ch.get("url", "")
        curl_n = str(curl).rstrip("/")
        by_slug = {v.get("slug"): v for v in self.videos}

        # Membres actuels du thème (d'après le cache)
        current = set()
        for v in self.videos:
            if turl_n in [u.rstrip("/") for u in self._rel_urls(v.get("theme"), normalise=False)]:
                current.add(v.get("slug"))

        to_add = desired - current        # à ranger dans le thème
        to_remove = current - desired     # à sortir du thème
        ok = fail = forced = 0

        for slug in (to_add | to_remove):
            v = by_slug.get(slug)
            if not v:
                continue
            themes = self._rel_urls(v.get("theme"), normalise=False)
            themes_n = [u.rstrip("/") for u in themes]
            chans = self._rel_urls(v.get("channel"), normalise=False)
            chans_n = [u.rstrip("/") for u in chans]
            payload = {}

            if slug in to_add:
                if turl_n not in themes_n:
                    themes.append(turl)                       # ajout du thème
                    payload["theme"] = themes
                # COHÉRENCE : forcer l'appartenance à la chaîne parente
                if curl_n not in chans_n:
                    chans.append(curl)
                    payload["channel"] = chans
                    forced += 1
            if slug in to_remove:
                themes = [u for u in themes if u.rstrip("/") != turl_n]  # retrait du thème
                payload["theme"] = themes
                # On NE touche PAS à la chaîne au retrait (la vidéo y reste)

            if not payload:
                continue
            try:
                self.api.patch_video(v, payload)              # PATCH theme (+ channel si forcé)
                v.update(payload)                             # MAJ cache local
                self._sync_video_caches(slug, payload)
                ok += 1
            except Exception as e:
                fail += 1
                self._ui(self._log, f"❌ {slug} : {e}")

        msg = (f"✅  Thème « {theme.get('title')} » : "
               f"+{len(to_add)} / -{len(to_remove)} vidéo(s).")
        if forced:
            msg += f"  ({forced} ajoutée(s) aussi à la chaîne pour cohérence.)"
        self._ui(self.ct_status.configure,
                 text=msg, text_color=T_SUCCES if not fail else "#f59e0b")
        self._ui(self._log,
                 f"Thème « {theme.get('title')} » : {len(to_add)} ajout(s), "
                 f"{len(to_remove)} retrait(s), {forced} forcé(s) en chaîne, {fail} échec(s).")
        self._ui(self._render_ct)      # refléter le changement dans la liste

    def _ct_manage_owners(self, ch):
        """Ouvre un sélecteur de comptes pour gérer les ADMINISTRATEURS (owners)
        de la chaîne. Les owners sont des comptes individuels (pas des groupes)
        qui peuvent administrer la chaîne. Pré-coche les administrateurs actuels."""
        if not self.api:
            self.ct_status.configure(text="Connectez-vous d'abord.", text_color=T_ALERTE)
            return
        if not self.all_users:
            self.ct_status.configure(text="⏳  Chargement des comptes…", text_color=T_SECONDAIRE)
            self._run(lambda: (self._reload_users_for_admin(),
                               self._ui(lambda: self._ct_open_owner_picker(ch))))
            return
        self._ct_open_owner_picker(ch)

    def _ct_open_owner_picker(self, ch):
        """Construit la pré-sélection (owners actuels) et ouvre OwnerPicker."""
        owners = ch.get("owners") or []
        if isinstance(owners, str):
            owners = [owners]
        # Table URL de compte → libellé lisible (à partir des comptes chargés)
        label_by_url = {str(u.get("url", "")).rstrip("/"): self._user_label(u)
                        for u in (self.all_users or [])}
        pre = {}
        for o in owners:
            ourl = o.get("url") if isinstance(o, dict) else o
            pre[ourl] = label_by_url.get(str(ourl).rstrip("/"), ourl)
        OwnerPicker(self,
                    on_done=lambda urls, labels: self._ct_apply_owners(ch, urls),
                    title=f"Administrateurs de « {ch.get('title')} »",
                    preselected=pre)

    def _ct_apply_owners(self, ch, urls):
        """Callback : applique la nouvelle liste d'administrateurs en arrière-plan."""
        self._run(self._do_ct_apply_owners, ch, list(urls))

    def _do_ct_apply_owners(self, ch, urls):
        """(Thread) PATCH du champ `owners` de la chaîne (liste d'URLs de comptes)."""
        try:
            self.api.patch_channel(ch.get("url"), {"owners": urls})
            ch["owners"] = urls                       # MAJ cache local
            self._ui(self.ct_status.configure,
                     text=f"✅  Chaîne « {ch.get('title')} » : "
                          f"{len(urls)} administrateur(s).", text_color=T_SUCCES)
            self._ui(self._log,
                     f"Chaîne « {ch.get('title')} » : {len(urls)} administrateur(s) défini(s).")
            self._ui(self._render_ct)      # refléter le changement dans la liste
        except Exception as e:
            self._ui(self.ct_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Administrateurs « {ch.get('title')} » : {e}")

    def _ct_manage_groups(self, ch):
        """Restreint toutes les vidéos d'une chaîne à des groupes d'accès.
        Charge d'abord les vidéos (en arrière-plan) pour pré-cocher les groupes
        déjà appliqués, puis ouvre la fenêtre de sélection."""
        if not self.api:
            self.ct_status.configure(text="Connectez-vous d'abord.", text_color=T_ALERTE)
            return
        if not self.access_groups:
            self.ct_status.configure(text="Aucun groupe d'accès chargé.", text_color=T_ALERTE)
            return
        self.ct_status.configure(text="⏳  Lecture des restrictions…", text_color=T_SECONDAIRE)
        self._run(self._do_ct_prepare_groups, ch)

    def _do_ct_prepare_groups(self, ch):
        """(Thread) Analyse les restrictions des vidéos de la chaîne, puis ouvre
        la fenêtre de sélection.

        On calcule DEUX informations, et pas seulement l'intersection :
          • `common` : groupes présents sur TOUTES les vidéos → cases cochées ;
          • `counts` : pour chaque groupe, le NOMBRE de vidéos qui l'ont.

        Pourquoi : avec la seule intersection, il suffisait qu'UNE vidéo sur
        trente n'ait pas le groupe pour que la case apparaisse vide — laissant
        croire qu'aucune restriction n'existait. Le décompte permet d'afficher
        un état « partiel » (ex. « 29/30 ») au lieu de masquer l'information."""
        try:
            # Les données peuvent avoir changé depuis le dernier chargement
            # (modification faite dans un autre onglet) : on recharge si le
            # cache est vide, et on s'appuie sinon sur les caches synchronisés.
            self.ensure_videos_sync()      # magasin partagé (chargé si besoin)
            curl = str(ch.get("url", "")).rstrip("/")

            def in_chan(v):
                """Teste si une vidéo appartient à la chaîne filtrée."""
                return curl in self._rel_urls(v.get("channel"))

            vids = [v for v in self.videos if in_chan(v)]
            # Groupes communs à TOUTES les vidéos (intersection) = restriction
            # uniforme en vigueur ; sert de pré-cochage.
            common = None
            counts = {}                     # URL de groupe → nb de vidéos concernées
            for v in vids:
                gs = set(self._rel_urls(v.get("restrict_access_to_groups")))
                for g in gs:
                    counts[g] = counts.get(g, 0) + 1
                common = gs if common is None else (common & gs)
            common = common or set()
            self._ui(self.ct_status.configure,
                     text=f"{len(vids)} vidéo(s) dans la chaîne.", text_color=T_SECONDAIRE)
            self._ui(lambda: self._ct_groups_dialog(ch, common, counts, len(vids)))
        except Exception as e:
            self._ui(self.ct_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Lecture restrictions chaîne : {e}")

    def _ct_groups_dialog(self, ch, current_norm, counts=None, nb_videos=0):
        """Fenêtre de sélection des groupes de la chaîne.

        `current_norm` : groupes appliqués à TOUTES les vidéos → cases cochées.
        `counts`       : URL de groupe → nombre de vidéos concernées. Sert à
                         signaler les états PARTIELS (ex. « 29/30 vidéos »),
                         qui, sans cela, apparaîtraient comme « non restreint »
                         et induiraient l'utilisateur en erreur.
        `nb_videos`    : nombre total de vidéos dans la chaîne."""
        counts = counts or {}
        win = ctk.CTkToplevel(self)
        win.title(f"Restreindre « {ch.get('title')} »")
        win.geometry("470x520")
        _focus_toplevel(win, self)

        ctk.CTkLabel(win, text=f"Chaîne : {ch.get('title')}",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(padx=16, pady=(16, 2), anchor="w")
        ctk.CTkLabel(win, text=f"{nb_videos} vidéo(s) dans la chaîne. Cochez les groupes "
                               "autorisés.\nLa restriction sera appliquée à TOUTES les vidéos "
                               "(statut « Restreint »).\nSans groupe coché, les vidéos repassent "
                               "en public.",
                     text_color=T_SECONDAIRE, font=ctk.CTkFont(size=11),
                     justify="left").pack(padx=16, anchor="w")

        holder = ctk.CTkScrollableFrame(win, height=280, label_text="Groupes d'accès", fg_color=S_CARTE, label_anchor="w",
                                                  label_font=ctk.CTkFont(size=12, weight="bold"))
        holder.pack(fill="both", expand=True, padx=16, pady=10)
        gvars = {}
        partiels = 0
        for g in self.access_groups:
            gurl = g.get("url", "")
            gnorm = str(gurl).rstrip("/")
            n = counts.get(gnorm, 0)                 # nb de vidéos ayant ce groupe
            uniforme = gnorm in current_norm         # groupe présent partout
            var = ctk.BooleanVar(value=uniforme)
            gvars[gurl] = var
            ligne = ctk.CTkFrame(holder, fg_color="transparent")
            ligne.pack(fill="x", pady=1)
            ctk.CTkCheckBox(ligne, text=g.get("code_name", "?"),
                            variable=var).pack(side="left", anchor="w")
            # État PARTIEL : le groupe existe sur une partie seulement des
            # vidéos. La case reste décochée (l'action s'applique à tout), mais
            # on l'affiche pour que l'utilisateur ne croie pas à une absence.
            if n and not uniforme:
                partiels += 1
                ctk.CTkLabel(ligne, text=f"⚠ partiel : {n}/{nb_videos}",
                             text_color=T_ALERTE,
                             font=ctk.CTkFont(size=11)).pack(side="left", padx=8)
            elif uniforme:
                ctk.CTkLabel(ligne, text=f"toutes ({n}/{nb_videos})",
                             text_color=T_SUCCES,
                             font=ctk.CTkFont(size=11)).pack(side="left", padx=8)

        if partiels:
            ctk.CTkLabel(win,
                         text=f"⚠ {partiels} groupe(s) ne concernent qu'une PARTIE des vidéos. "
                              "Appliquer uniformisera la chaîne : les cases décochées seront "
                              "retirées de toutes les vidéos.",
                         text_color=T_ALERTE, font=ctk.CTkFont(size=11),
                         wraplength=430, justify="left").pack(padx=16, anchor="w")

        bar = ctk.CTkFrame(win, fg_color="transparent")
        bar.pack(fill="x", padx=16, pady=(0, 12))
        ctk.CTkButton(bar, text="Appliquer à la chaîne", fg_color=C_SUCCES,
                      hover_color=C_SUCCES_SURV,
                      command=lambda: (
                          win.destroy(),
                          self._ct_apply_groups(ch, [u for u, v in gvars.items() if v.get()]))
                      ).pack(side="left")
        ctk.CTkButton(bar, text="Annuler", fg_color=C_NEUTRE,
                      command=win.destroy, text_color=T_SUR_NEUTRE).pack(side="left", padx=8)

    def _ct_apply_groups(self, ch, group_urls):
        """Confirme puis propage la restriction (en arrière-plan)."""
        n_groups = len(group_urls)
        verb = (f"restreindre à {n_groups} groupe(s)" if n_groups
                else "retirer toute restriction de groupe")
        if not messagebox.askyesno(
                "Restreindre la chaîne",
                f"Appliquer « {verb} » à TOUTES les vidéos de « {ch.get('title')} » ?"):
            return
        self.ct_status.configure(text="⏳  Application en cours…", text_color=T_SECONDAIRE)
        self._run(self._do_ct_apply_groups, ch, list(group_urls))

    def _do_ct_apply_groups(self, ch, group_urls):
        """(Thread) Applique la restriction par groupe à chaque vidéo de la chaîne."""
        try:
            self.ensure_videos_sync()      # magasin partagé (chargé si besoin)
            curl = str(ch.get("url", "")).rstrip("/")
            # Vidéos appartenant à cette chaîne
            def in_chan(v):
                """Teste si une vidéo appartient à la chaîne filtrée."""
                chans = v.get("channel") or []
                if isinstance(chans, str):
                    chans = [chans]
                urls = [str(c.get("url") if isinstance(c, dict) else c).rstrip("/")
                        for c in chans]
                return curl in urls
            vids = [v for v in self.videos if in_chan(v)]
            ok = fail = 0
            for i, v in enumerate(vids, 1):
                try:
                    self.api.set_video_groups(v, group_urls)
                    v["restrict_access_to_groups"] = list(group_urls)
                    v["is_restricted"] = bool(group_urls)
                    sync_payload = {"restrict_access_to_groups": list(group_urls),
                                    "is_restricted": bool(group_urls)}
                    if group_urls:
                        v["is_draft"] = False     # cohérent avec set_video_groups
                        sync_payload["is_draft"] = False
                    self._sync_video_caches(v.get("slug"), sync_payload)
                    ok += 1
                except Exception as e:
                    fail += 1
                    self._ui(self._log, f"❌ {v.get('slug')} : {e}")
                self._ui(self.ct_status.configure,
                         text=f"⏳  {i}/{len(vids)}…", text_color=T_SECONDAIRE)
            msg = (f"✅  « {ch.get('title')} » : {ok} vidéo(s) "
                   f"{'restreinte(s)' if group_urls else 'rendue(s) publiques'}.")
            self._ui(self.ct_status.configure,
                     text=msg, text_color=T_SUCCES if not fail else "#f59e0b")
            self._ui(self._log,
                     f"Restriction chaîne « {ch.get('title')} » : {ok} OK, {fail} échec(s), "
                     f"{len(group_urls)} groupe(s).")
            self._ui(self._render_ct)  # refléter le changement dans la liste
        except Exception as e:
            self._ui(self.ct_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Restriction chaîne : {e}")

    def _ct_toggle_visible(self, ch):
        # Inverse la visibilité de la chaîne
        new_val = not bool(ch.get("visible"))
        self._run(self._do_ct_patch, "channel", ch.get("url"),
                  {"visible": new_val},
                  f"Chaîne « {ch.get('title')} » → {'visible' if new_val else 'masquée'}")

    def _do_ct_patch(self, kind, url, payload, logmsg):
        """(Thread) PATCH générique sur une chaîne ou un thème, puis rechargement."""
        try:
            if kind == "channel":
                self.api.patch_channel(url, payload)
            else:
                self.api.patch_theme(url, payload)
            self._ui(self._log, logmsg)
            self._do_ct_load()
            # Les autres onglets affichent des noms de chaîne : les rafraîchir.
            self._ui(self.schedule_refresh, channels=True)
        except Exception as e:
            self._ui(self.ct_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Modification {kind} : {e}")

    # ── Suppression (double confirmation) ──────────────────────────────────

    def _ct_delete_channel(self, ch):
        """Supprime une chaîne (après confirmation)."""
        if not messagebox.askyesno(
                "⚠️  Supprimer la chaîne",
                f"Supprimer la chaîne « {ch.get('title')} » ?\n\n"
                "Ses thèmes seront supprimés. Les vidéos ne sont pas supprimées "
                "mais perdront ce classement."):
            return
        if not messagebox.askyesno("Dernière confirmation",
                                   "Confirmez-vous la suppression de cette chaîne ?"):
            return
        self._run(self._do_ct_delete, "channel", ch.get("url"), ch.get("title"))

    def _ct_delete_theme(self, th):
        """Supprime un thème (après confirmation)."""
        if not messagebox.askyesno(
                "Supprimer le thème",
                f"Supprimer le thème « {th.get('title')} » ?"):
            return
        self._run(self._do_ct_delete, "theme", th.get("url"), th.get("title"))

    def _do_ct_delete(self, kind, url, label):
        """(Thread) DELETE d'une chaîne ou d'un thème, puis rechargement."""
        try:
            if kind == "channel":
                self.api.delete_channel(url)
            else:
                self.api.delete_theme(url)
            self._ui(self._log, f"{kind.capitalize()} supprimé(e) : {label}")
            self._do_ct_load()
        except Exception as e:
            self._ui(self.ct_status.configure, text=f"❌  {e}", text_color=T_ERREUR)
            self._ui(self._log, f"❌ Suppression {kind} : {e}")

    # ═════════════════════════════════════════════════════════════════════
    #  ONGLET JOURNAL
    # ═════════════════════════════════════════════════════════════════════

    def _build_tab_help(self):
        """Onglet « Aide » : mode d'emploi intégré de PodAdmin (console
        d'administration : token superutilisateur, compte véhicule pour le chunké
        des gros fichiers avec choix du propriétaire, onglets d'admin…)."""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tabs["help"] = frame

        ctk.CTkLabel(frame, text="❓  Aide — PodAdmin",
                     font=ctk.CTkFont(size=20, weight="bold"),
                     text_color=("#1d4ed8", "#60a5fa")).pack(anchor="w", padx=6, pady=(4, 2))
        ctk.CTkLabel(frame,
                     text="Mode d'emploi de la console d'administration. "
                          "Contact : support-pod@utoulouse.fr.",
                     font=ctk.CTkFont(size=12), text_color=T_DISCRET).pack(anchor="w", padx=6, pady=(0, 8))

        scroll = ctk.CTkScrollableFrame(frame, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=2, pady=2)

        def section(titre, corps, couleur_titre=("#1d4ed8", "#60a5fa")):
            """Ajoute une carte (titre + texte) à la page d'aide."""
            card = ctk.CTkFrame(scroll, fg_color=S_CARTE, corner_radius=10)
            card.pack(fill="x", padx=4, pady=6)
            ctk.CTkLabel(card, text=titre, font=ctk.CTkFont(size=14, weight="bold"),
                         text_color=couleur_titre, justify="left").pack(
                anchor="w", padx=14, pady=(12, 4))
            ctk.CTkLabel(card, text=corps.strip(), font=ctk.CTkFont(size=12),
                         text_color=("gray20", "gray85"), justify="left",
                         wraplength=760).pack(anchor="w", padx=14, pady=(0, 12))

        section(
            "🚀  Démarrage rapide",
            "1. Onglet Configuration : saisissez l'URL et le TOKEN d'un compte "
            "SUPERUTILISATEUR, puis « Tester & se connecter ».\n"
            "2. (Facultatif) Renseignez le compte VÉHICULE si vous téléverserez des "
            "fichiers de plus de 500 Mo.\n"
            "3. Onglet Téléversement : ajoutez des vidéos, choisissez le propriétaire "
            "et le type, puis lancez.")

        section(
            "🔑  Token superutilisateur & compte véhicule",
            "• TOKEN (obligatoire) : compte superutilisateur. Il donne accès à toute "
            "l'instance (comptes, vidéos, chaînes…) et sert à toutes les opérations API.\n"
            "• COMPTE VÉHICULE (facultatif) : un compte LOCAL, servant uniquement à "
            "ouvrir la session web du téléversement par morceaux (chunké). Requis "
            "seulement pour les fichiers > 500 Mo.\n\n"
            "Les deux sont stockés CHIFFRÉS dans le coffre-fort de l'OS, dans un espace "
            "séparé des autres applis Pod. « Oublier le token » efface tout du poste.")

        section(
            "📂  Téléverser des vidéos (et gros fichiers)",
            "• Choisissez le PROPRIÉTAIRE : les vidéos lui appartiendront.\n"
            "• Bascule automatique par taille : ≤ 500 Mo → envoi classique par token ; "
            "> 500 Mo → envoi par MORCEAUX via le compte véhicule, puis la vidéo est "
            "RÉATTRIBUÉE au propriétaire choisi (métadonnées + encodage ensuite). "
            "Transparent à l'usage.\n"
            "• Co-propriétaires et crédits (co-auteurs) disponibles comme d'habitude.")

        section(
            "⚠️  Gros fichiers : réattribution & finalisation serveur",
            "Deux points à connaître pour les fichiers > 500 Mo :\n"
            "• La vidéo naît d'abord au nom du compte véhicule, puis est réattribuée "
            "au propriétaire choisi. Si cette réattribution échoue, l'appli l'affiche "
            "en ROUGE (« ⚠️ NON réattribuée ») : la vidéo reste alors au véhicule, à "
            "corriger via l'onglet Réaffectation. Ce n'est jamais silencieux.\n"
            "• La finalisation d'un très gros fichier peut prendre plusieurs minutes ; "
            "si la passerelle affiche une erreur 504, l'appli attend (jusqu'à 30 min) "
            "que la vidéo apparaisse côté serveur, puis poursuit. Laissez-la travailler.",
            couleur_titre=("#b45309", "#f59e0b"))

        section(
            "🔒  Visibilité et restriction à un groupe",
            "La visibilité d'une CHAÎNE et celle d'une VIDÉO sont indépendantes : "
            "mettre une vidéo dans une chaîne cachée ne restreint PAS la vidéo.\n\n"
            "Pour restreindre des vidéos à un groupe :\n"
            "• Une par une : onglet Vidéos → « Restreindre à des groupes ».\n"
            "• En lot : onglet Vidéos → Ctrl+clic sur les vidéos → « 🔐 Restreindre "
            "au groupe… » → choisir le(s) groupe(s). Les vidéos passent en « Restreint » "
            "et sortent de brouillon.")

        section(
            "🎫  Donner un accès à un enseignant (jeton)",
            "Circuit complet, depuis l'onglet Comptes :\n"
            "1. L'enseignant se connecte une fois à Pod (cela CRÉE son compte) puis "
            "remplit le questionnaire.\n"
            "2. Filtrez son nom dans Comptes, puis activez l'interrupteur « Équipe » : "
            "ce statut est nécessaire pour que ses dépôts de gros fichiers soient "
            "réattribués correctement.\n"
            "3. Bouton « 🔑 Token » : ouvre les jetons de l'administration Pod dans votre "
            "navigateur, recherche déjà remplie au nom du compte.\n"
            "   • S'il a DÉJÀ un jeton : recopiez-le simplement.\n"
            "   • Sinon : « Ajouter », choisir le compte, « Enregistrer », copier.\n"
            "4. Bouton « ✉️ » : ouvre votre messagerie avec le message d'accueil déjà "
            "rédigé, avec le support en copie cachée. Remplacez le repère\n"
            "   « >>> COLLER ICI LA CLÉ <<< » par le jeton copié, puis envoyez.\n"
            "   (Le message parle de « clé d'activation » : le mot « token » ne dit "
            "rien aux enseignants.)\n\n"
            "⚠️ UN SEUL JETON PAR COMPTE : Pod n'en autorise qu'un. Tenter d'en créer "
            "un second échoue avec une erreur. Et surtout, ne supprimez PAS un jeton "
            "existant pour en refaire un : l'ancien devient aussitôt invalide et "
            "l'application de l'enseignant cesse de fonctionner jusqu'à ce qu'il saisisse "
            "le nouveau. À ne faire que sur demande (jeton perdu ou compromis), en "
            "prévenant la personne.\n\n"
            "Le jeton n'est pas créé par l'application : l'API ne le permet pas, et le "
            "faire supposerait d'y stocker un mot de passe superutilisateur. La création "
            "reste donc dans le navigateur, sous votre propre identité d'administrateur.\n"
            "Le bouton ✉️ est grisé si aucune adresse n'est connue pour le compte.")

        section(
            "👁  Statistiques de consultation",
            "L'onglet Inventaire affiche désormais les VUES, en plus de la volumétrie.\n\n"
            "Le nombre total de vues et la période couverte apparaissent dans les "
            "chiffres clés. Le menu « Répartition par » propose quatre vues "
            "supplémentaires :\n"
            "• Vues — vidéos : le classement des plus consultées, avec leur propriétaire ;\n"
            "• Vues — chaînes : quelles chaînes sont réellement regardées ;\n"
            "• Vues — utilisateurs : quels enseignants ont de l'audience ;\n"
            "• Vues — par mois : l'évolution dans le temps, avec une barre de "
            "proportion.\n\n"
            "Tout est repris dans l'export Excel, en quatre feuilles distinctes.\n\n"
            "À savoir : le total compte TOUTES les vues enregistrées, y compris "
            "celles de vidéos supprimées depuis ; le classement, lui, ne montre que "
            "les vidéos encore présentes. La somme du classement peut donc être "
            "inférieure au total — c'est normal.\n\n"
            "Tant que la plateforme n'est pas ouverte, ces chiffres restent faibles : "
            "c'est attendu.")

        section(
            "⌨  Raccourcis clavier",
            "Dans toutes les fenêtres secondaires (choix d'un propriétaire, d'une "
            "chaîne, d'une bannière, suivi d'un envoi) :\n"
            "• Échap — fermer sans valider ;\n"
            "• Entrée — déclencher l'action principale, quand la fenêtre en "
            "désigne une (« Valider »).\n\n"
            "Entrée ne valide JAMAIS au hasard : une fenêtre sans action "
            "principale évidente ne réagit pas à cette touche, pour éviter de "
            "déclencher par mégarde une opération de masse.")

        section(
            "🎨  Apparence : mode clair ou sombre",
            "Le bouton en bas de la barre latérale, juste au-dessus du numéro de "
            "version, bascule entre le mode sombre (par défaut) et le mode clair.\n\n"
            "Le choix est MÉMORISÉ : l'application rouvrira dans le mode retenu.\n\n"
            "Le mode clair convient mieux aux salles très éclairées et à la "
            "vidéoprojection ; le mode sombre fatigue moins en usage prolongé.")

        section(
            "🛠  Réglages non accessibles depuis l'application",
            "Certains paramètres de Pod ne sont pas exposés par l'API : l'application "
            "ne peut donc pas les modifier. C'est le cas de la PAGE D'ACCUEIL de la "
            "plateforme, qui relève de la configuration du serveur.\n\n"
            "L'onglet Configuration propose, tout en bas, une section « Réglages de "
            "l'instance » avec quatre boutons qui ouvrent l'administration de Pod "
            "directement à la bonne page.\n\n"
            "La page d'accueil se règle à DEUX endroits distincts :\n"
            "• « Accueil : texte » — le texte de présentation (page statique) ;\n"
            "• « Accueil : blocs » — les vignettes et encarts affichés.\n\n"
            "Vous devez y être connecté en administrateur — c'est votre session de "
            "navigateur qui vous authentifie, l'application ne détient aucun mot de "
            "passe privilégié.")

        section(
            "🎨  Habiller une chaîne ou un thème",
            "Bouton « 🎨 Habillage » sur chaque chaîne (et « 🎨 » sur chaque thème) "
            "de l'onglet Chaînes. On y règle :\n"
            "• le titre et la description ;\n"
            "• la couleur (chaînes seulement), avec aperçu et sélecteur ;\n"
            "• la bannière ;\n"
            "• la visibilité publique (chaînes seulement).\n\n"
            "Pour la bannière, deux possibilités :\n"
            "• BIBLIOTHÈQUE — réutiliser une image déjà présente sur la plateforme. "
            "Les vignettes générées automatiquement pour les vidéos sont masquées par "
            "défaut, sans quoi elles noieraient les vraies bannières ; décochez la case "
            "pour toutes les voir.\n"
            "• DEPUIS MON ORDINATEUR — déposer une nouvelle image (JPG ou PNG). Pod "
            "refuse les images trop petites ou mal formées.\n\n"
            "Le bouton « Retirer la bannière » enlève l'image sans en poser d'autre.\n\n"
            "À savoir : la page d'ACCUEIL de la plateforme n'est pas modifiable depuis "
            "l'application — elle relève de la configuration du serveur, et non de "
            "l'API.")

        section(
            "☑  Traiter plusieurs vidéos à la fois (onglet Vidéos)",
            "Dans la liste de l'onglet Vidéos :\n"
            "• Ctrl + clic pour ajouter ou retirer une vidéo de la sélection ;\n"
            "• Maj + clic pour sélectionner toute une plage.\n\n"
            "Les lignes retenues passent en bleu, et le panneau de droite bascule sur "
            "les actions groupées : mettre en brouillon, rendre public, rendre "
            "restreint, restreindre à des groupes, affecter à une ou plusieurs "
            "chaînes.\n\n"
            "Le bouton « ☑ Tout sélectionner », à droite du compteur, retient TOUTES "
            "les vidéos filtrées — y compris celles qui ne sont pas affichées, la "
            "liste étant limitée à 300 lignes. Le panneau indique alors combien sont "
            "hors affichage.\n\n"
            "Deux filtres de DÉTECTION complètent la recherche :\n"
            "• « Doublons de titre » — vidéos portant le même titre ;\n"
            "• « Vieux brouillons » — brouillons plus anciens que le nombre de mois "
            "indiqué à côté.\n\n"
            "La liste peut être triée par « Plus récentes » (ordre d'origine), "
            "« A → Z » ou « Z → A » — utile pour retrouver une vidéo dont on "
            "connaît le titre.\n\n"
            "Pour l'affectation à une chaîne, deux modes sont proposés :\n"
            "• AJOUTER — les vidéos restent dans leurs chaînes actuelles ;\n"
            "• REMPLACER — les affectations existantes sont perdues.\n"
            "L'ajout est proposé en premier : c'est l'option la moins destructrice, "
            "et une vidéo peut légitimement appartenir à plusieurs chaînes.\n\n"
            "La sélection est CONSERVÉE après une action : on peut enchaîner "
            "(restreindre à un groupe, puis affecter à une chaîne) sans tout "
            "recocher. « Annuler la sélection » revient au détail d'une vidéo.\n\n"

            "Pendant un traitement, le bouton « 🛑 Interrompre le traitement » "
            "s'active. L'arrêt est PROPRE : la vidéo en cours est menée à son "
            "terme, puis le traitement s'arrête — on ne coupe jamais au milieu "
            "d'une opération. Le bilan indique combien de vidéos n'ont pas été "
            "traitées.\n\n"
            "La suppression en masse figure dans la « ZONE SENSIBLE », séparée des "
            "autres actions en bas du panneau. Elle demande une DOUBLE "
            "confirmation : un avertissement, puis la saisie du nombre exact de "
            "vidéos concernées — recopier ce chiffre oblige à regarder combien on "
            "s'apprête à détruire.\n\n"
            "Rappel : Pod n'a pas de corbeille, la suppression est IRRÉVERSIBLE. "
            "Pour masquer des vidéos sans les perdre, préférez « Mettre en "
            "brouillon ».")

        section(
            "⚠️  Message « LISTE INCOMPLÈTE »",
            "Si ce message rouge apparaît, la lecture des vidéos s'est arrêtée sur une "
            "limite de sécurité : les totaux affichés sont FAUX et des vidéos manquent. "
            "Ce n'est pas un simple avertissement de confort — ne vous fiez pas aux "
            "chiffres de l'Inventaire dans ce cas, et signalez-le au support.")

        section(
            "🗂  Les onglets d'administration",
            "• Comptes : recherche, statut « Équipe », création de jeton et message "
            "d'accueil (voir la rubrique « Donner un accès »).\n"
            "• Réaffectation : changer le propriétaire de vidéos (par lot).\n"
            "  (L'ancien onglet Explorateur a été fusionné ici : ses filtres de "
            "détection, la sélection globale et la suppression en masse sont "
            "désormais dans l'onglet Vidéos.)\n"
            "• Chaînes : chaînes et thèmes, ajout de vidéos, restriction/visibilité.\n"
            "• Groupes d'accès : gestion des groupes.\n"

            "• Journal : historique horodaté — le premier endroit à consulter en cas de "
            "souci. Il est aussi ENREGISTRÉ SUR LE DISQUE (un fichier par mois) : le "
            "bouton « Effacer » ne vide que l'affichage, et « 📂 Ouvrir les journaux » "
            "donne accès à l'historique complet.")

        section(
            "🍎  macOS : « est endommagé » ou « impossible d'ouvrir »",
            "L'application N'EST PAS endommagée : macOS affiche ce message pour toute "
            "application diffusée hors de l'App Store. Deux méthodes :\n\n"
            "1) La plus simple — Réglages Système → Confidentialité et sécurité : après "
            "une tentative d'ouverture, un bouton « Ouvrir quand même » apparaît en bas.\n\n"
            "2) Si ce bouton n'apparaît pas (cas du message « est endommagé ») :\n"
            "   • Copiez d'abord l'application depuis le .dmg vers le dossier Applications "
            "(ou le Bureau) — impossible de la débloquer tant qu'elle est dans le .dmg, "
            "qui est en lecture seule. Éjectez ensuite le .dmg.\n"
            "   • Ouvrez le Terminal (⌘+Espace, « Terminal ») et tapez :\n"
            "       xattr -cr\n"
            "     puis un ESPACE, puis glissez l'application dans la fenêtre du Terminal "
            "(le chemin s'écrit tout seul) et appuyez sur Entrée.\n"
            "   • Si le message persiste, la signature du paquet a été abîmée pendant le "
            "transfert. Réparez-la avec :\n"
            "       codesign --force --deep --sign - \n"
            "     suivi du même glisser-déposer.\n\n"
            "À noter : transférer un .app par messagerie (Telegram, WhatsApp…) ou le "
            "recompresser avec un outil quelconque casse souvent sa signature. Préférez "
            "toujours le .dmg, et transmettez-le par un lien de téléchargement.")

        section(
            "🛠  Dépannage courant",
            "• Un .exe déjà compilé ne reflète pas les mises à jour du code : pour "
            "tester une nouvelle version, lancez « python app.py », ou recompilez.\n"
            "• Gros fichier refusé : vérifiez que le compte véhicule est renseigné et "
            "valide (onglet Configuration).\n"
            "• Erreur réseau pendant un envoi : chaque morceau est ré-essayé "
            "automatiquement ; utilisez « Relancer les échecs » au besoin.")

        section(
            "✉️  Support",
            "Une question, un bug, une amélioration ? Écrivez à "
            "support-pod@utoulouse.fr en joignant, si possible, le contenu de "
            "l'onglet Journal au moment du problème.",
            couleur_titre=("gray40", "gray70"))

    def _build_tab_about(self):
        """Onglet « À propos » : informations sur l'application, sa version,
        ses auteurs et le contact support (aligné sur Pod Téléverseur)."""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tabs["about"] = frame

        # Carte centrale
        card = ctk.CTkFrame(frame, fg_color=S_CARTE, corner_radius=12)
        card.pack(padx=20, pady=20, fill="x")

        # Logo (repli texte si absent), sur bandeau blanc
        try:
            logo_path = resource_path(os.path.join("assets", "logo_ut.png"))
            if os.path.exists(logo_path):
                from PIL import Image
                img = Image.open(logo_path)
                ratio = img.width / img.height if img.height else 3
                ctkimg = ctk.CTkImage(light_image=img, dark_image=img,
                                      size=(int(64 * ratio), 64))
                band = ctk.CTkFrame(card, fg_color="white", corner_radius=8)
                band.pack(pady=(20, 10))
                ctk.CTkLabel(band, image=ctkimg, text="").pack(padx=16, pady=8)
        except Exception:
            pass

        ctk.CTkLabel(card, text="PodAdmin",
                     font=ctk.CTkFont(size=26, weight="bold")).pack(pady=(6, 0))
        ctk.CTkLabel(card, text=f"Version {__version__}",
                     font=ctk.CTkFont(size=13), text_color=T_DISCRET).pack(pady=(0, 10))
        ctk.CTkLabel(
            card,
            text="Console d'administration pour l'instance Esup-Pod de\n"
                 "l'Université de Toulouse (videos.utoulouse.fr).\n"
                 "Téléversement, modération, chaînes, thèmes, groupes d'accès.",
            font=ctk.CTkFont(size=12), text_color=("gray30", "gray75"),
            justify="center").pack(pady=(0, 14))

        # Séparateur
        ctk.CTkFrame(card, height=1, fg_color=S_FILET).pack(fill="x", padx=40, pady=4)

        # Auteurs
        ctk.CTkLabel(card, text="Développé par",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(pady=(12, 2))
        for nom in ("Cédric MONNA", "Philippe BAQUÉ", "Michel JACOB"):
            ctk.CTkLabel(card, text=nom, font=ctk.CTkFont(size=12),
                         text_color=("gray20", "gray85")).pack()

        # Contact + institution
        ctk.CTkLabel(card, text="Université de Toulouse",
                     font=ctk.CTkFont(size=12), text_color=T_DISCRET).pack(pady=(12, 0))
        ctk.CTkLabel(card, text="support-pod@utoulouse.fr",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=("#1d4ed8", "#60a5fa")).pack(pady=(0, 14))

        # Mention légale : titulaire des droits, puis conditions d'utilisation.
        # Le copyright désigne l'auteur ; la licence dit ce qu'on a le droit
        # d'en faire — c'est elle qui encadre réellement la réutilisation.
        ctk.CTkLabel(card, text=__copyright__,
                     font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=T_DISCRET).pack(pady=(0, 2))
        ctk.CTkLabel(card, text=__license__,
                     font=ctk.CTkFont(size=10, slant="italic"),
                     text_color=T_DISCRET, wraplength=420,
                     justify="center").pack(padx=20, pady=(0, 18))

    def _build_tab_log(self):
        """Construit l'onglet Journal (zone de texte horodatée + bouton Effacer)."""
        frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tabs["log"] = frame
        top = ctk.CTkFrame(frame, fg_color="transparent")
        top.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(top, text="📋  Journal", font=ctk.CTkFont(size=20, weight="bold")).pack(side="left")
        ctk.CTkButton(top, text="🗑 Effacer", width=100, fg_color=C_NEUTRE,
                      hover_color=C_NEUTRE_SURV, command=self._clear_log, text_color=T_SUR_NEUTRE).pack(side="right")
        # Le journal est AUSSI écrit sur disque : « Effacer » ne vide que
        # l'affichage, l'historique complet reste consultable dans le fichier.
        ctk.CTkButton(top, text="📂 Ouvrir les journaux", width=170, fg_color=C_NEUTRE,
                      hover_color=C_NEUTRE_SURV,
                      command=self._open_journal_folder, text_color=T_SUR_NEUTRE).pack(side="right", padx=6)
        ctk.CTkLabel(frame,
                     text="Le journal est enregistré chaque mois dans "
                          "« Mes documents\\.podadmin » (dossier personnel). "
                          "« Effacer » ne vide que l'affichage.",
                     font=ctk.CTkFont(size=11), text_color=T_DISCRET,
                     anchor="w").pack(fill="x", pady=(0, 6))
        self.log_box = ctk.CTkTextbox(frame, font=ctk.CTkFont(family="Consolas", size=11))
        self.log_box.pack(fill="both", expand=True)
        self.log_box.configure(state="disabled")
        self._log("Application démarrée.")

    def _log(self, msg: str):
        """Ajoute une ligne horodatée au journal (écran ET fichier)."""
        maintenant = datetime.now()
        ts = maintenant.strftime("%H:%M:%S")
        self.log_box.configure(state="normal")
        self.log_box.insert("end", f"[{ts}]  {msg}\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")
        self._log_fichier(maintenant, msg)

    def _log_fichier(self, quand, msg: str):
        """Recopie la ligne de journal dans un fichier mensuel sur le disque.

        L'affichage à l'écran est effaçable en un clic et disparaît à la
        fermeture. Or cette application SUPPRIME définitivement des vidéos et
        réaffecte des comptes sur toute l'instance : il faut pouvoir retrouver
        après coup qui a fait quoi, et quand. Le fichier est en mode ajout, un
        par mois, dans le dossier personnel.

        Toute erreur d'écriture est ignorée : la journalisation ne doit jamais
        empêcher l'application de fonctionner.
        """
        try:
            dossier = os.path.join(os.path.expanduser("~"), ".podadmin")
            os.makedirs(dossier, exist_ok=True)
            chemin = os.path.join(dossier, f"journal-{quand.strftime('%Y-%m')}.log")
            with open(chemin, "a", encoding="utf-8") as f:
                f.write(f"{quand.strftime('%Y-%m-%d %H:%M:%S')}\t{msg}\n")
        except Exception:
            pass      # jamais bloquant

    def _journal_path(self) -> str:
        """Chemin du fichier de journal du mois en cours."""
        return os.path.join(os.path.expanduser("~"), ".podadmin",
                            f"journal-{datetime.now().strftime('%Y-%m')}.log")

    def _open_journal_folder(self):
        """Ouvre le dossier contenant les journaux dans l'explorateur de fichiers."""
        dossier = os.path.join(os.path.expanduser("~"), ".podadmin")
        try:
            os.makedirs(dossier, exist_ok=True)
            if sys.platform.startswith("win"):
                os.startfile(dossier)                     # noqa: S606 (Windows)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", dossier])
            else:
                subprocess.Popen(["xdg-open", dossier])
            self._log(f"Dossier des journaux ouvert : {dossier}")
        except Exception as e:
            self._log(f"❌ Ouverture du dossier des journaux : {e}")

    def _clear_log(self):
        """Vide le journal."""
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")


# ════════════════════════════════════════════════════════════════════════════
#  FENÊTRE : sélection de propriétaires additionnels
# ════════════════════════════════════════════════════════════════════════════

def _focus_toplevel(win, master=None):
    """Amène une fenêtre secondaire au premier plan, lui donne le focus et la
    rend modale (focus capturé jusqu'à fermeture). Corrige le cas où une
    CTkToplevel s'ouvre derrière la fenêtre principale.
    Les appels sont légèrement différés (after) car la fenêtre n'est pas encore
    dessinée à l'instant de sa création."""
    try:
        if master is not None:
            win.transient(master)          # la fenêtre reste au-dessus de son parent
    except Exception:
        pass

    # RACCOURCIS CLAVIER, posés ici parce que TOUTES les fenêtres secondaires
    # passent par cette fonction : un seul endroit à maintenir, aucune risque
    # d'en oublier une.
    #
    # Échap ferme la fenêtre. Entrée déclenche l'action principale si la fenêtre
    # en désigne une via `bouton_defaut` — on ne devine pas : valider au hasard
    # dans une fenêtre de suppression serait pire que pas de raccourci du tout.
    def _sur_echap(_e=None):
        """Ferme la fenêtre (Échap)."""
        try:
            annuler = getattr(win, "action_annuler", None)
            if callable(annuler):
                annuler()
            else:
                win.destroy()
        except Exception:
            pass
        return "break"

    def _sur_entree(_e=None):
        """Déclenche l'action principale, si la fenêtre en a désigné une."""
        bouton = getattr(win, "bouton_defaut", None)
        try:
            if bouton is not None and bouton.winfo_exists() \
                    and str(bouton.cget("state")) != "disabled":
                bouton.invoke()
        except Exception:
            pass
        return "break"

    try:
        win.bind("<Escape>", _sur_echap)
        win.bind("<Return>", _sur_entree)
        win.bind("<KP_Enter>", _sur_entree)     # pavé numérique
    except Exception:
        pass
    win.lift()
    win.attributes("-topmost", True)        # passe au-dessus, le temps de s'afficher
    # On retire 'topmost' juste après (sinon elle resterait au-dessus de TOUTES
    # les applications), puis on capture le focus.
    win.after(150, lambda: (win.attributes("-topmost", False), win.focus_force()))
    win.after(200, lambda: win.grab_set())  # modale : bloque la fenêtre principale


class ProgressModal(ctk.CTkToplevel):
    """Fenêtre MODALE de progression, pour les opérations longues à ne pas
    interrompre (remplacement d'un fichier source + ré-encodage).

    Pourquoi une modale : pendant un remplacement, toute autre manipulation
    (changer de vidéo, actualiser la liste, relancer l'action…) peut couper
    l'envoi en cours. Cette fenêtre capture le focus (`grab_set`) et neutralise
    la croix de fermeture tant que l'opération tourne : l'utilisateur ne peut
    donc rien faire d'autre que patienter, et voit l'avancement.

    Cycle de vie :
      • création (thread principal) → `set_phase()` / `set_progress()` pendant
        le travail (appelés depuis le thread via App._ui) ;
      • `finish(ok, message)` en fin d'opération : la fenêtre se déverrouille,
        affiche le résultat et propose un bouton « Fermer ».
    """

    def __init__(self, master, title: str = "Opération en cours",
                 intro: str = "", subtitle: str = ""):
        """Construit la fenêtre modale de progression (titre, sous-titre, étape initiale)."""
        super().__init__(master)
        self.master_app = master
        self._done = False                 # opération terminée ? (pilote la fermeture)
        self.title(title)
        # 470 x 250 était trop court : le contenu réclame environ 265 px au
        # départ, et jusqu'à 290 quand le message de fin est long (cas d'une
        # finalisation coupée par la passerelle). Le bouton « Fermer » se
        # retrouvait alors hors de la fenêtre, qui n'est pas redimensionnable.
        # On prévoit une marge, et la fenêtre s'ajustera d'elle-même si un
        # message plus long survient (voir _ajuster_hauteur).
        self.geometry("470x320")
        self.resizable(False, False)
        # Tant que l'opération tourne, la croix de fermeture est NEUTRALISÉE :
        # fermer la fenêtre laisserait un envoi orphelin en arrière-plan.
        self.protocol("WM_DELETE_WINDOW", self._on_close_attempt)

        ctk.CTkLabel(self, text="⏳  Veuillez patienter…",
                     font=ctk.CTkFont(size=17, weight="bold")).pack(
            anchor="w", padx=20, pady=(18, 2))

        self.subtitle_lbl = ctk.CTkLabel(
            self, text=subtitle, text_color=T_SECONDAIRE, font=ctk.CTkFont(size=12),
            wraplength=420, justify="left")
        self.subtitle_lbl.pack(anchor="w", padx=20, pady=(0, 8))

        # Phase courante (envoi / finalisation / ré-encodage…)
        self.phase_lbl = ctk.CTkLabel(self, text=intro, font=ctk.CTkFont(size=13),
                                      wraplength=420, justify="left")
        self.phase_lbl.pack(anchor="w", padx=20, pady=(0, 6))

        # Barre d'avancement (même code couleur que le téléversement par lot)
        self.bar = ctk.CTkProgressBar(self, progress_color="#16a34a")
        self.bar.pack(fill="x", padx=20)
        self.bar.set(0)

        # Détail chiffré sous la barre (Mo envoyés / Mo total)
        self.detail_lbl = ctk.CTkLabel(self, text="", text_color=T_SECONDAIRE,
                                       font=ctk.CTkFont(size=11))
        self.detail_lbl.pack(anchor="w", padx=20, pady=(4, 0))

        # Rappel : ne pas interrompre (masqué une fois l'opération finie)
        self.warn_lbl = ctk.CTkLabel(
            self, text="Ne fermez pas cette fenêtre et ne lancez pas d'autre action : "
                       "cela interromprait l'envoi.",
            text_color=T_ALERTE, font=ctk.CTkFont(size=11),
            wraplength=420, justify="left")
        self.warn_lbl.pack(anchor="w", padx=20, pady=(10, 0))

        # Bouton de fermeture : désactivé jusqu'à la fin de l'opération.
        # Échap referme la fenêtre — mais seulement si le bouton est actif,
        # c'est-à-dire une fois le traitement terminé.
        self.action_annuler = lambda: (
            self.close_btn.invoke()
            if str(self.close_btn.cget("state")) != "disabled" else None)
        self.close_btn = ctk.CTkButton(self, text="Fermer", width=110,
                                       fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                                       state="disabled", command=self._close_now, text_color=T_SUR_NEUTRE)
        self.close_btn.pack(anchor="e", padx=20, pady=(10, 14))

        _focus_toplevel(self, master)      # au premier plan + modale (grab_set)

    # ── Mises à jour (appelées depuis le thread de travail via App._ui) ────

    def _ajuster_hauteur(self):
        """Agrandit la fenêtre si son contenu dépasse la hauteur disponible.

        Les messages de fin varient beaucoup en longueur (une ligne pour un
        succès, cinq pour une finalisation interrompue). Plutôt que de figer une
        taille suffisante pour le pire cas — ce qui laisserait un grand vide la
        plupart du temps — on ajuste après coup, uniquement si nécessaire."""
        if not self.winfo_exists():
            return
        try:
            self.update_idletasks()
            requise = self.winfo_reqheight()
            actuelle = self.winfo_height()
            if requise > actuelle:
                self.geometry(f"470x{requise + 20}")
        except Exception:
            pass

    def set_phase(self, text: str, color: str = None):
        """Change le libellé de l'étape en cours (envoi, finalisation, encodage…)."""
        if not self.winfo_exists():
            return
        self.phase_lbl.configure(text=text, **({"text_color": color} if color else {}))

    def set_progress(self, fraction: float, detail: str = ""):
        """Positionne la barre (0 à 1) et le détail chiffré sous la barre."""
        if not self.winfo_exists():
            return
        self.bar.set(max(0.0, min(1.0, fraction)))
        if detail:
            self.detail_lbl.configure(text=detail)

    def set_indeterminate(self, on: bool = True):
        """Bascule en animation continue quand l'avancement n'est pas mesurable
        (finalisation côté serveur : on ne sait pas combien de temps il reste)."""
        if not self.winfo_exists():
            return
        try:
            if on:
                self.bar.configure(mode="indeterminate")
                self.bar.start()
            else:
                self.bar.stop()
                self.bar.configure(mode="determinate")
        except Exception:
            pass

    def finish(self, ok: bool, message: str):
        """Fin de l'opération : déverrouille la fenêtre et affiche le résultat."""
        if not self.winfo_exists():
            return
        self._done = True
        self.set_indeterminate(False)
        self.bar.set(1.0 if ok else self.bar.get())
        self.phase_lbl.configure(text=("✅  " if ok else "❌  ") + message,
                                 text_color=T_SUCCES if ok else "#ef4444")
        self.warn_lbl.configure(text="Opération terminée. Vous pouvez fermer cette fenêtre.",
                                text_color=T_SECONDAIRE)
        self.close_btn.configure(state="normal")
        self._ajuster_hauteur()            # le message de fin peut être long
        try:
            self.grab_release()            # rend la main à la fenêtre principale
        except Exception:
            pass

    def ensure_unlocked(self):
        """FILET DE SÉCURITÉ : déverrouille la fenêtre si l'opération s'est
        terminée sans passer par `finish()` (voie de sortie imprévue). Sans ce
        garde-fou, une modale restée « grabbed » figerait toute l'application."""
        if not self.winfo_exists() or self._done:
            return
        self.finish(False, "Opération terminée de façon inattendue. "
                           "Vérifiez la vidéo sur le site et le Journal.")

    # ── Fermeture ─────────────────────────────────────────────────────────

    def _on_close_attempt(self):
        """Clic sur la croix : ignoré tant que l'opération n'est pas terminée."""
        if self._done:
            self._close_now()

    def _close_now(self):
        """Ferme réellement la fenêtre (après la fin de l'opération)."""
        try:
            self.grab_release()
        except Exception:
            pass
        self.destroy()


class OwnerPicker(ctk.CTkToplevel):
    """Sélecteur multi-utilisateurs : même système que l'agent (liste + filtre + clic)."""

    def __init__(self, master: App, on_done, title="Propriétaires additionnels",
                 preselected: dict | None = None, single: bool = False, on_single=None):
        """Construit la fenêtre de sélection de propriétaires (liste + filtre)."""
        super().__init__(master)
        self.master_app = master
        self.on_done = on_done
        self.single = single
        self.on_single = on_single
        self.title(title)
        self.geometry("500x560")
        _focus_toplevel(self, master)
        self.selected: dict[str, str] = dict(preselected or {})   # url → libellé

        intro = ("Cliquez sur un utilisateur pour le choisir." if single else
                 "Cochez les comptes Pod à ajouter comme propriétaires\n"
                 "additionnels. Filtrez la liste puis cliquez pour (dé)cocher.")
        ctk.CTkLabel(self, text=intro, justify="left").pack(padx=14, pady=(14, 8), anchor="w")

        bar = ctk.CTkFrame(self, fg_color="transparent")
        bar.pack(fill="x", padx=14)
        self.filter = ctk.CTkEntry(bar, placeholder_text="🔍 nom / identifiant…")
        self.filter.pack(side="left", fill="x", expand=True, padx=(0, 6))
        # Temporisation : évite de reconstruire toute la liste à chaque caractère.
        self.filter.bind("<KeyRelease>", lambda e: self._render_differe())
        ctk.CTkButton(bar, text="🔄", width=40, fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=self._reload, text_color=T_SUR_NEUTRE).pack(side="left")

        self.count_lbl = ctk.CTkLabel(self, text="", text_color=T_SECONDAIRE, font=ctk.CTkFont(size=11))
        self.count_lbl.pack(anchor="w", padx=14, pady=(4, 0))

        self.listbox = ctk.CTkScrollableFrame(self, height=320, fg_color=S_CARTE)
        self.listbox.pack(fill="both", expand=True, padx=14, pady=8)

        self.chosen_lbl = ctk.CTkLabel(self, text="Sélection : aucun", text_color=T_SECONDAIRE,
                                       wraplength=460, justify="left")
        self.chosen_lbl.pack(padx=14, anchor="w")

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=14, pady=10)
        # Désigné comme action par défaut : Entrée déclenche « Valider ».
        self.bouton_defaut = ctk.CTkButton(
            btns, text="Valider", fg_color=C_SUCCES, hover_color=C_SUCCES_SURV,
            command=self._validate)
        self.bouton_defaut.pack(side="right")
        ctk.CTkButton(btns, text="Annuler", fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=self.destroy, text_color=T_SUR_NEUTRE).pack(side="right", padx=8)

        self.after(80, self._init_list)

    def _init_list(self):
        """Affiche la liste si les comptes sont déjà chargés, sinon déclenche un chargement."""
        if self.master_app.all_users:
            self._render()
            self._update_chosen()
        else:
            self.count_lbl.configure(text="⏳  Chargement des utilisateurs…")
            self._reload()

    def _reload(self):
        """(Thread) Charge la liste des comptes si nécessaire, puis rafraîchit l'affichage."""
        def work():
            """(Thread) Charge les données nécessaires puis rafraîchit l'affichage."""
            try:
                if not self.master_app.all_users:
                    users = self.master_app.api.get_all_users()
                    users.sort(key=lambda u: (u.get("username") or "").lower())
                    self.master_app.all_users = users
                self.after(0, self._render)
                self.after(0, self._update_chosen)
            except Exception as e:
                # PIÈGE PYTHON 3 : le nom `e` est SUPPRIMÉ à la sortie du bloc
                # `except`. Or ce lambda n'est exécuté que plus tard, par
                # `after(0, …)` : il lèverait alors NameError, et l'utilisateur
                # resterait devant « ⏳ Chargement… » sans jamais voir l'erreur.
                # On capture donc le message dans une variable ordinaire.
                msg = str(e)
                self.after(0, lambda m=msg: self.count_lbl.configure(
                    text=f"Erreur : {m}", text_color=T_ERREUR))
        threading.Thread(target=work, daemon=True).start()

    def _label(self, u: dict) -> str:
        """Libellé lisible d'un compte."""
        return f"{u.get('username','?')} — {u.get('first_name','')} {u.get('last_name','')}".strip()

    def _render_differe(self):
        """Replanifie l'affichage après une courte pause de frappe (voir
        App._debounce) : une seule reconstruction au lieu d'une par caractère."""
        job = getattr(self, "_render_job", None)
        if job:
            try:
                self.after_cancel(job)
            except Exception:
                pass
        self._render_job = self.after(FILTER_DELAY_MS, self._render)

    def _render(self):
        """Affiche la liste filtrée (cases à cocher)."""
        flt = self.filter.get().strip().lower()
        for w in self.listbox.winfo_children():
            w.destroy()
        users = self.master_app.all_users
        if not users:
            ctk.CTkLabel(self.listbox, text="Liste non disponible.", text_color=T_SECONDAIRE).pack(pady=10)
            return
        matches = [u for u in users if not flt or flt in self._label(u).lower()]
        CAP = 300
        for u in matches[:CAP]:
            url = u.get("url", "")
            sel = url in self.selected
            if self.single:
                prefix = "   "
            else:
                prefix = "☑  " if sel else "☐  "
            ctk.CTkButton(self.listbox, text=prefix + self._label(u), anchor="w",
                          fg_color=S_SELECTION if (sel and not self.single) else "transparent",
                          text_color=("gray10", "gray90"), hover_color=("gray75", "gray28"),
                          height=28, font=ctk.CTkFont(size=12),
                          command=lambda uu=u: self._toggle(uu)).pack(fill="x", pady=1)
        self.count_lbl.configure(text=f"{len(matches)} affiché(s) sur {len(users)} — "
                                      f"{len(self.selected)} sélectionné(s)", text_color=T_SECONDAIRE)
        if len(matches) > CAP:
            ctk.CTkLabel(self.listbox, text=f"… affinez le filtre ({len(matches) - CAP} de plus)",
                         text_color=T_SECONDAIRE).pack(pady=4)
        elif not matches:
            ctk.CTkLabel(self.listbox, text="Aucun résultat.", text_color=T_SECONDAIRE).pack(pady=8)

    def _toggle(self, u: dict):
        """Coche/décoche un compte (ou valide directement en mode sélection unique)."""
        if self.single:
            if self.on_single:
                self.on_single(u)
            self.destroy()
            return
        url = u.get("url", "")
        if not url:
            return
        if url in self.selected:
            del self.selected[url]
        else:
            self.selected[url] = self._label(u)
        self._render()
        self._update_chosen()

    def _update_chosen(self):
        """Met à jour le libellé récapitulant la sélection courante."""
        if self.selected:
            self.chosen_lbl.configure(text="Sélection : " + ", ".join(self.selected.values()),
                                      text_color=T_SUCCES)
        else:
            self.chosen_lbl.configure(text="Sélection : aucun", text_color=T_SECONDAIRE)

    def _validate(self):
        """Renvoie la sélection à l'appelant (on_done) puis ferme la fenêtre."""
        self.on_done(list(self.selected.keys()), list(self.selected.values()))
        self.destroy()


# ════════════════════════════════════════════════════════════════════════════


class VideoPicker(ctk.CTkToplevel):
    """Sélecteur multi-vidéos (recherche + cases à cocher).
    `videos` : liste de dicts {slug, title, …}. `on_done(slugs)` au Valider.
    `preselected` : dict {slug: titre} cochés au départ (membres actuels)."""

    def __init__(self, master, videos, on_done, title="Vidéos",
                 preselected: dict | None = None):
        """Construit la fenêtre de sélection de vidéos (liste + filtre)."""
        super().__init__(master)
        self.on_done = on_done
        self.videos = videos or []
        self.selected: dict[str, str] = dict(preselected or {})   # slug → titre
        self.title(title)
        self.geometry("520x560")
        _focus_toplevel(self, master)

        ctk.CTkLabel(self, text="Cochez les vidéos à inclure dans la chaîne "
                                "(décochez pour les retirer).",
                     justify="left", wraplength=480).pack(padx=14, pady=(14, 8), anchor="w")

        self.filter = ctk.CTkEntry(self, placeholder_text="🔍 titre / slug…")
        self.filter.pack(fill="x", padx=14)
        # Temporisation : évite de reconstruire toute la liste à chaque caractère.
        self.filter.bind("<KeyRelease>", lambda e: self._render_differe())

        self.listbox = ctk.CTkScrollableFrame(self, height=360, fg_color=S_CARTE)
        self.listbox.pack(fill="both", expand=True, padx=14, pady=8)

        self.chosen_lbl = ctk.CTkLabel(self, text="0 vidéo sélectionnée", text_color=T_SECONDAIRE)
        self.chosen_lbl.pack(padx=14, anchor="w")

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=14, pady=10)
        # Désigné comme action par défaut : Entrée déclenche « Valider ».
        self.bouton_defaut = ctk.CTkButton(
            btns, text="Valider", fg_color=C_SUCCES, hover_color=C_SUCCES_SURV,
            command=self._validate)
        self.bouton_defaut.pack(side="right")
        ctk.CTkButton(btns, text="Annuler", fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=self.destroy, text_color=T_SUR_NEUTRE).pack(side="right", padx=8)

        self._render()
        self._update_chosen()

    def _render_differe(self):
        """Replanifie l'affichage après une courte pause de frappe (voir
        App._debounce) : une seule reconstruction au lieu d'une par caractère."""
        job = getattr(self, "_render_job", None)
        if job:
            try:
                self.after_cancel(job)
            except Exception:
                pass
        self._render_job = self.after(FILTER_DELAY_MS, self._render)

    def _render(self):
        """Affiche la liste filtrée des vidéos (cases à cocher)."""
        flt = self.filter.get().strip().lower()
        for w in self.listbox.winfo_children():
            w.destroy()
        matches = [v for v in self.videos
                   if not flt or flt in f"{v.get('title','')} {v.get('slug','')}".lower()]
        CAP = 500
        for v in matches[:CAP]:
            slug = v.get("slug", "")
            sel = slug in self.selected
            title = (v.get("title") or "(sans titre)")[:54]
            ctk.CTkButton(self.listbox, text=("☑  " if sel else "☐  ") + f"{title}  ·  {slug}",
                          anchor="w", height=26,
                          fg_color=S_SELECTION if sel else "transparent",
                          text_color=("gray10", "gray90"), hover_color=("gray75", "gray28"),
                          font=ctk.CTkFont(size=12),
                          command=lambda vv=v: self._toggle(vv)).pack(fill="x", pady=1)
        if len(matches) > CAP:
            ctk.CTkLabel(self.listbox, text=f"… +{len(matches) - CAP} autres. Affinez le filtre.",
                         text_color=T_SECONDAIRE).pack(pady=4)
        elif not matches:
            ctk.CTkLabel(self.listbox, text="Aucune vidéo.", text_color=T_SECONDAIRE).pack(pady=8)

    def _toggle(self, v: dict):
        """Coche/décoche une vidéo dans la sélection."""
        slug = v.get("slug", "")
        if not slug:
            return
        if slug in self.selected:
            del self.selected[slug]
        else:
            self.selected[slug] = v.get("title", "?")
        self._render()
        self._update_chosen()

    def _update_chosen(self):
        """Met à jour le compteur de sélection."""
        n = len(self.selected)
        self.chosen_lbl.configure(
            text=f"{n} vidéo(s) sélectionnée(s)",
            text_color=T_SUCCES if n else "gray")

    def _validate(self):
        """Renvoie la liste des slugs sélectionnés à l'appelant puis ferme."""
        self.on_done(list(self.selected.keys()))
        self.destroy()


class BannerPicker(ctk.CTkToplevel):
    """Fenêtre de choix d'une bannière : bibliothèque de l'instance ou fichier local.

    Deux façons de poser une bannière :
      • BIBLIOTHÈQUE — réutiliser une image déjà présente sur l'instance ;
      • ORDINATEUR   — déposer un nouveau fichier.

    Deux difficultés traitées ici :

    1. LE BRUIT. Une instance accumule des centaines d'images, en très grande
       majorité des vignettes générées automatiquement pour les vidéos (noms du
       type « 0005-galerie-imagemp4_3 »). Une galerie brute serait inutilisable :
       un filtre masque ces vignettes par défaut.

    2. LA LENTEUR. Afficher toutes les vignettes signifierait autant de
       téléchargements. On plafonne donc l'affichage et on charge les images
       PROGRESSIVEMENT, en arrière-plan, sans figer la fenêtre.
    """

    # Au-delà, on n'affiche pas : il faut affiner le filtre.
    PLAFOND = 40

    def __init__(self, master, titre: str, image_actuelle: str = ""):
        """Prépare la fenêtre de choix de bannière.

        `image_actuelle` : URL de la bannière déjà posée, mise en évidence dans
        la galerie pour qu'on voie ce qu'on remplace."""
        super().__init__(master)
        self.master_app = master
        self.resultat = None          # URL choisie, "" pour retirer, None si annulé
        self.image_actuelle = str(image_actuelle or "")
        self.images = []              # catalogue complet (chargé une fois)
        self.vignettes = {}           # URL d'image → CTkImage (évite de retélécharger)
        self._job_filtre = None

        self.title(titre)
        self.geometry("760x620")
        _focus_toplevel(self, master)

        ctk.CTkLabel(self, text=titre,
                     font=ctk.CTkFont(size=15, weight="bold")).pack(
            anchor="w", padx=16, pady=(14, 2))

        self.onglets = ctk.CTkTabview(self, height=470)
        self.onglets.pack(fill="both", expand=True, padx=12, pady=6)
        self.onglets.add("Bibliothèque")
        self.onglets.add("Depuis mon ordinateur")

        self._build_bibliotheque(self.onglets.tab("Bibliothèque"))
        self._build_local(self.onglets.tab("Depuis mon ordinateur"))

        bas = ctk.CTkFrame(self, fg_color="transparent")
        bas.pack(fill="x", padx=16, pady=(0, 12))
        ctk.CTkButton(bas, text="Retirer la bannière", width=160,
                      fg_color=C_ALERTE, hover_color=C_ALERTE_SURV,
                      command=self._retirer).pack(side="left")
        ctk.CTkButton(bas, text="Annuler", width=110, fg_color=C_NEUTRE,
                      hover_color=C_NEUTRE_SURV, command=self._annuler, text_color=T_SUR_NEUTRE).pack(side="right")

        self._charger_catalogue()

    # ── Onglet bibliothèque ───────────────────────────────────────────────

    def _build_bibliotheque(self, parent):
        """Construit la galerie des images déjà présentes sur l'instance."""
        barre = ctk.CTkFrame(parent, fg_color="transparent")
        barre.pack(fill="x", padx=6, pady=(6, 2))
        self.filtre = ctk.CTkEntry(barre, placeholder_text="🔍 filtrer par nom…", width=260)
        self.filtre.pack(side="left")
        self.filtre.bind("<KeyRelease>", lambda e: self._filtrer_differe())

        # Les vignettes de vidéos sont écartées par défaut : ce sont elles qui
        # noient les vraies bannières.
        self.masquer_vignettes = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(barre, text="Masquer les vignettes de vidéos",
                        variable=self.masquer_vignettes,
                        command=self._filtrer).pack(side="left", padx=12)

        self.compteur = ctk.CTkLabel(parent, text="Chargement…", text_color=T_SECONDAIRE,
                                     font=ctk.CTkFont(size=11), anchor="w")
        self.compteur.pack(fill="x", padx=8, pady=(2, 4))

        self.galerie = ctk.CTkScrollableFrame(parent, height=350, fg_color=S_CARTE)
        self.galerie.pack(fill="both", expand=True, padx=6, pady=(0, 6))

    def _charger_catalogue(self):
        """(Thread) Récupère la liste des images de l'instance."""
        def travail():
            """(Thread) Lecture du catalogue d'images."""
            try:
                images = self.master_app.api.get_images()
            except Exception as e:
                self.master_app._ui(self.compteur.configure,
                                    text=f"❌ {e}", text_color=T_ERREUR)
                return
            self.images = images
            self.master_app._ui(self._filtrer)
        self.master_app._run(travail)

    @staticmethod
    def _est_vignette_video(nom: str) -> bool:
        """Ce nom ressemble-t-il à une vignette générée automatiquement ?

        Pod nomme ces images d'après le slug de la vidéo, avec un suffixe
        numérique : « 0005-galerie-imagemp4_3 ». On repère ce motif plutôt
        que de lister des cas particuliers."""
        import re as _re
        return bool(_re.match(r"^\d{3,5}-.*_\d+$", (nom or "").strip()))

    def _filtrer_differe(self):
        """Attend une pause dans la frappe avant de reconstruire la galerie."""
        if self._job_filtre:
            try:
                self.after_cancel(self._job_filtre)
            except Exception:
                pass
        self._job_filtre = self.after(FILTER_DELAY_MS, self._filtrer)

    def _filtrer(self):
        """Applique les filtres puis réaffiche la galerie."""
        self._job_filtre = None
        texte = (self.filtre.get() or "").strip().lower()
        retenues = []
        for img in self.images:
            nom = str(img.get("name", ""))
            if self.masquer_vignettes.get() and self._est_vignette_video(nom):
                continue
            if texte and texte not in nom.lower():
                continue
            retenues.append(img)
        self._afficher(retenues)

    def _afficher(self, images: list):
        """Dessine les vignettes (plafonnées) et lance leur chargement."""
        for w in self.galerie.winfo_children():
            w.destroy()

        total = len(images)
        montrees = images[:self.PLAFOND]
        masquees = len(self.images) - total
        detail = f"{total} image(s)"
        if masquees > 0:
            detail += f" — {masquees} vignette(s) de vidéos masquée(s)"
        if total > self.PLAFOND:
            detail += f" — {self.PLAFOND} affichées, affinez le filtre"
        self.compteur.configure(text=detail, text_color=T_SECONDAIRE)

        if not montrees:
            ctk.CTkLabel(self.galerie, text="Aucune image ne correspond.",
                         text_color=T_SECONDAIRE).pack(pady=20)
            return

        # Grille de 4 colonnes
        for i, img in enumerate(montrees):
            ligne, col = divmod(i, 4)
            case = ctk.CTkFrame(self.galerie, width=165, height=140, fg_color=S_LIGNE)
            case.grid(row=ligne, column=col, padx=6, pady=6)
            case.grid_propagate(False)

            actuelle = str(img.get("url", "")).rstrip("/") == self.image_actuelle.rstrip("/")
            apercu = ctk.CTkLabel(case, text="…", width=150, height=80,
                                  fg_color=S_PUCE, corner_radius=4)
            apercu.pack(padx=6, pady=(6, 2))

            nom = str(img.get("name", "?"))
            ctk.CTkLabel(case, text=(("✅ " if actuelle else "") + nom[:22]),
                         font=ctk.CTkFont(size=10),
                         text_color=(T_SUCCES if actuelle else T_SECONDAIRE)).pack()
            ctk.CTkButton(case, text="Choisir", height=22, width=140,
                          font=ctk.CTkFont(size=11), fg_color=C_ACTION, hover_color=C_ACTION_SURV,
                          command=lambda u=img.get("url", ""): self._choisir(u)).pack(pady=(2, 6))

            # Chargement de la vignette en arrière-plan, une par une.
            self._charger_vignette(img, apercu)

    def _charger_vignette(self, img: dict, cible):
        """(Thread) Télécharge et affiche une vignette, sans bloquer la fenêtre."""
        url_fichier = str(img.get("file", ""))
        if not url_fichier:
            return
        deja = self.vignettes.get(url_fichier)
        if deja is not None:
            cible.configure(image=deja, text="")
            return

        def travail():
            """(Thread) Téléchargement puis mise à l'échelle de l'image."""
            try:
                import io
                from PIL import Image as _Img
                r = self.master_app.api.session.get(url_fichier, timeout=20)
                if r.status_code != 200:
                    return
                pil = _Img.open(io.BytesIO(r.content))
                pil.thumbnail((150, 80))
                ctkimg = ctk.CTkImage(light_image=pil, dark_image=pil, size=pil.size)
                self.vignettes[url_fichier] = ctkimg
                self.master_app._ui(self._poser_vignette, cible, ctkimg)
            except Exception:
                pass          # une vignette illisible ne doit rien interrompre
        self.master_app._run(travail)

    @staticmethod
    def _poser_vignette(cible, image):
        """Affiche la vignette téléchargée (thread principal)."""
        try:
            if cible.winfo_exists():
                cible.configure(image=image, text="")
        except Exception:
            pass

    # ── Onglet fichier local ──────────────────────────────────────────────

    def _build_local(self, parent):
        """Construit l'onglet de dépôt d'un fichier depuis le poste."""
        ctk.CTkLabel(parent,
                     text="Déposer une nouvelle image sur la plateforme, puis "
                          "l'utiliser comme bannière.",
                     font=ctk.CTkFont(size=12), wraplength=650,
                     justify="left").pack(anchor="w", padx=10, pady=(12, 6))

        cadre = ctk.CTkFrame(parent, fg_color=S_CARTE)
        cadre.pack(fill="x", padx=10, pady=6)
        cadre.columnconfigure(1, weight=1)

        ctk.CTkLabel(cadre, text="Fichier :", width=90, anchor="e").grid(
            row=0, column=0, padx=8, pady=8)
        self.chemin_lbl = ctk.CTkEntry(cadre, placeholder_text="aucun fichier choisi")
        self.chemin_lbl.grid(row=0, column=1, sticky="ew", padx=(0, 6), pady=8)
        ctk.CTkButton(cadre, text="Parcourir…", width=110,
                      fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=self._parcourir, text_color=T_SUR_NEUTRE).grid(row=0, column=2, padx=8, pady=8)

        ctk.CTkLabel(cadre, text="Nom :", width=90, anchor="e").grid(
            row=1, column=0, padx=8, pady=8)
        self.nom_entry = ctk.CTkEntry(cadre, placeholder_text="libellé de l'image")
        self.nom_entry.grid(row=1, column=1, columnspan=2, sticky="ew", padx=(0, 8), pady=8)

        ctk.CTkLabel(cadre, text="Dossier :", width=90, anchor="e").grid(
            row=2, column=0, padx=8, pady=8)
        self.dossier_menu = ctk.CTkOptionMenu(cadre, values=["(chargement…)"], **STYLE_CHAMP)
        self.dossier_menu.grid(row=2, column=1, columnspan=2, sticky="ew",
                               padx=(0, 8), pady=8)

        self.msg_local = ctk.CTkLabel(parent, text="", font=ctk.CTkFont(size=11),
                                      wraplength=650, justify="left", anchor="w")
        self.msg_local.pack(fill="x", padx=10, pady=(4, 0))

        ctk.CTkButton(parent, text="⬆  Déposer et utiliser comme bannière",
                      height=34, fg_color=C_SUCCES, hover_color=C_SUCCES_SURV,
                      command=self._deposer).pack(padx=10, pady=12)

        ctk.CTkLabel(parent,
                     text="Formats acceptés : JPG, PNG. Pod refuse les images trop "
                          "petites ou mal formées.",
                     font=ctk.CTkFont(size=10), text_color=T_DISCRET,
                     wraplength=650, justify="left").pack(anchor="w", padx=10)

        self._charger_dossiers()

    def _charger_dossiers(self):
        """(Thread) Remplit la liste des dossiers de rangement."""
        def travail():
            """(Thread) Lecture des dossiers."""
            try:
                dossiers = self.master_app.api.get_folders()
            except Exception:
                dossiers = []
            self._dossiers = {f"{d.get('name', '?')}": d.get("url", "") for d in dossiers}
            valeurs = list(self._dossiers.keys()) or ["(aucun dossier)"]
            self.master_app._ui(self.dossier_menu.configure, values=valeurs)
            self.master_app._ui(self.dossier_menu.set, valeurs[0])
        self.master_app._run(travail)

    def _parcourir(self):
        """Ouvre le sélecteur de fichier et pré-remplit le nom."""
        from tkinter import filedialog
        chemin = filedialog.askopenfilename(
            title="Choisir une image",
            filetypes=[("Images", "*.jpg *.jpeg *.png"), ("Tous les fichiers", "*.*")])
        if not chemin:
            return
        self.chemin_lbl.delete(0, "end")
        self.chemin_lbl.insert(0, chemin)
        if not self.nom_entry.get().strip():
            import os as _os
            self.nom_entry.insert(0, _os.path.splitext(_os.path.basename(chemin))[0])

    def _deposer(self):
        """Dépose l'image choisie puis la retient comme bannière."""
        chemin = (self.chemin_lbl.get() or "").strip()
        if not chemin:
            self.msg_local.configure(text="Choisissez d'abord un fichier.",
                                     text_color=T_ALERTE)
            return
        dossier_url = getattr(self, "_dossiers", {}).get(self.dossier_menu.get(), "")
        if not dossier_url:
            self.msg_local.configure(
                text="Aucun dossier de rangement disponible : impossible de déposer.",
                text_color=T_ERREUR)
            return
        # `created_by` : le compte connecté à PodAdmin, c'est-à-dire celui qui
        # dépose réellement l'image.
        createur = getattr(self.master_app, "vehicle_owner_url", "") or ""
        if not createur:
            for u in (getattr(self.master_app, "all_users", None) or []):
                createur = u.get("url", "")
                break
        if not createur:
            self.msg_local.configure(
                text="Impossible de déterminer le compte déposant.", text_color=T_ERREUR)
            return

        self.msg_local.configure(text="⏳ Dépôt en cours…", text_color=T_SECONDAIRE)

        def travail():
            """(Thread) Dépôt de l'image sur l'instance."""
            try:
                img = self.master_app.api.upload_image(
                    chemin, self.nom_entry.get().strip(), dossier_url, createur)
                url = img.get("url", "")
                self.master_app._ui(self._choisir, url)
            except Exception as e:
                self.master_app._ui(self.msg_local.configure,
                                    text=f"❌ {e}", text_color=T_ERREUR)
        self.master_app._run(travail)

    # ── Sortie ────────────────────────────────────────────────────────────

    def _choisir(self, url: str):
        """Retient l'image et ferme la fenêtre."""
        self.resultat = url
        self._fermer()

    def _retirer(self):
        """Demande le retrait de la bannière (la chaîne n'en aura plus)."""
        self.resultat = ""
        self._fermer()

    def _annuler(self):
        """Ferme sans rien changer."""
        self.resultat = None
        self._fermer()

    def _fermer(self):
        """Libère la fenêtre."""
        try:
            self.grab_release()
        except Exception:
            pass
        self.destroy()


class ChannelPicker(ctk.CTkToplevel):
    """Sélecteur multi-chaînes (sur le modèle d'OwnerPicker).
    `channels` : liste de dicts {url, title}. `on_done(urls, labels)` au Valider."""

    def __init__(self, master, channels, on_done, title="Chaînes",
                 preselected: dict | None = None):
        """Construit la fenêtre de sélection de chaînes (liste + filtre)."""
        super().__init__(master)
        self.on_done = on_done
        self.channels = channels or []
        self.selected: dict[str, str] = dict(preselected or {})   # url → titre
        self.title(title)
        self.geometry("460x520")
        _focus_toplevel(self, master)

        ctk.CTkLabel(self, text="Cochez les chaînes où la vidéo doit apparaître.",
                     justify="left").pack(padx=14, pady=(14, 8), anchor="w")

        self.filter = ctk.CTkEntry(self, placeholder_text="🔍 titre…")
        self.filter.pack(fill="x", padx=14)
        # Temporisation : évite de reconstruire toute la liste à chaque caractère.
        self.filter.bind("<KeyRelease>", lambda e: self._render_differe())

        self.listbox = ctk.CTkScrollableFrame(self, height=320, fg_color=S_CARTE)
        self.listbox.pack(fill="both", expand=True, padx=14, pady=8)

        self.chosen_lbl = ctk.CTkLabel(self, text="Sélection : aucune", text_color=T_SECONDAIRE,
                                       wraplength=420, justify="left")
        self.chosen_lbl.pack(padx=14, anchor="w")

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=14, pady=10)
        # Désigné comme action par défaut : Entrée déclenche « Valider ».
        self.bouton_defaut = ctk.CTkButton(
            btns, text="Valider", fg_color=C_SUCCES, hover_color=C_SUCCES_SURV,
            command=self._validate)
        self.bouton_defaut.pack(side="right")
        ctk.CTkButton(btns, text="Annuler", fg_color=C_NEUTRE, hover_color=C_NEUTRE_SURV,
                      command=self.destroy, text_color=T_SUR_NEUTRE).pack(side="right", padx=8)

        self._render()
        self._update_chosen()

    def _render_differe(self):
        """Replanifie l'affichage après une courte pause de frappe (voir
        App._debounce) : une seule reconstruction au lieu d'une par caractère."""
        job = getattr(self, "_render_job", None)
        if job:
            try:
                self.after_cancel(job)
            except Exception:
                pass
        self._render_job = self.after(FILTER_DELAY_MS, self._render)

    def _render(self):
        """Affiche la liste filtrée (cases à cocher)."""
        flt = self.filter.get().strip().lower()
        for w in self.listbox.winfo_children():
            w.destroy()
        matches = [c for c in self.channels
                   if not flt or flt in (c.get("title", "")).lower()]
        for c in matches:
            url = c.get("url", "")
            sel = url in self.selected
            ctk.CTkButton(self.listbox, text=("☑  " if sel else "☐  ") + c.get("title", "?"),
                          anchor="w", height=28,
                          fg_color=S_SELECTION if sel else "transparent",
                          text_color=("gray10", "gray90"), hover_color=("gray75", "gray28"),
                          font=ctk.CTkFont(size=12),
                          command=lambda cc=c: self._toggle(cc)).pack(fill="x", pady=1)
        if not matches:
            ctk.CTkLabel(self.listbox, text="Aucune chaîne.", text_color=T_SECONDAIRE).pack(pady=8)

    def _toggle(self, c: dict):
        """Coche/décoche une chaîne dans la sélection."""
        url = c.get("url", "")
        if not url:
            return
        if url in self.selected:
            del self.selected[url]
        else:
            self.selected[url] = c.get("title", "?")
        self._render()
        self._update_chosen()

    def _update_chosen(self):
        """Met à jour le libellé récapitulant la sélection courante."""
        if self.selected:
            self.chosen_lbl.configure(text="Sélection : " + ", ".join(self.selected.values()),
                                      text_color=T_SUCCES)
        else:
            self.chosen_lbl.configure(text="Sélection : aucune", text_color=T_SECONDAIRE)

    def _validate(self):
        """Renvoie la sélection à l'appelant (on_done) puis ferme la fenêtre."""
        self.on_done(list(self.selected.keys()), list(self.selected.values()))
        self.destroy()


# ════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # Thème retenu lors de la dernière utilisation. Sombre par défaut : c'est
    # le mode d'origine de l'application, et celui auquel les utilisateurs
    # actuels sont habitués.
    try:
        _theme = (cfg.load_config() or {}).get("theme", "dark")
    except Exception:
        _theme = "dark"
    ctk.set_appearance_mode(_theme if _theme in ("dark", "light") else "dark")
    ctk.set_default_color_theme("blue")
    app = App()
    app.mainloop()
